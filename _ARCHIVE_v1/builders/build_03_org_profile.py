"""
Builder du module 03 - TenderPilot Organization Profile.

Produit : dist/TenderPilot_Toolkit/03_ORGANIZATION_PROFILE/
          03_TenderPilot_Organization_Profile.xlsx

Principe central : un registre unique de champs (data/org_profile_fields.csv)
genere a la fois le formulaire de saisie, la vue normalisee PROFILE_SUMMARY
et le calcul de completude. Ajouter un champ = ajouter une ligne au CSV.

Le selecteur Candidate_Type (PME, CABINET, ONG, CONSULTANT, CONSORTIUM) ne
masque pas les champs - un tableur ne le fait pas de facon fiable - il les
qualifie : Requis / Optionnel / Sans objet.

Dependances : schema/columns.py, data/org_profile_fields.csv
"""

import csv
import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from schema import columns as S
from builders.build_02_command_center import (
    BORDER, FILL_GREEN, FILL_GREY, FILL_HEAD, FILL_ORANGE, FILL_RED,
    FILL_YELLOW, F_BODY, F_H2, F_HEAD, F_KPI, F_LABEL, F_MUTED, F_TITLE,
    dxf_fill,
)

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT_DIR = ROOT / "dist" / "TenderPilot_Toolkit" / "03_ORGANIZATION_PROFILE"
OUT_FILE = OUT_DIR / "03_TenderPilot_Organization_Profile.xlsx"

MODULE_VERSION = "0.1.0"
MAX_DOC_ROWS = 200

CANDIDATE_TYPES = ["PME", "CABINET", "ONG", "CONSULTANT", "CONSORTIUM"]

SECTION_TITLES = {
    "IDENTITY": "Identite legale et contacts",
    "FINANCE": "Capacites financieres",
    "EXPERIENCE": "Experience",
    "TEAM": "Ressources humaines",
    "CAPABILITIES": "Capacites techniques et politiques internes",
}

FORM_COLS = ["Champ", "Valeur", "Requis ?", "Aide", "Applicable_A",
             "Obligatoire_Registre"]

# Categories de documents - section 12 de la documentation fonctionnelle.
DOCUMENT_TYPES = [
    ("RCCM", "01_Legal"),
    ("Statuts / acte constitutif", "01_Legal"),
    ("Recepisse ou agrement", "01_Legal"),
    ("Accord de consortium", "01_Legal"),
    ("IFU / identifiant fiscal", "02_Fiscal"),
    ("Attestation fiscale", "02_Fiscal"),
    ("Attestation sociale", "02_Fiscal"),
    ("Releve d'identite bancaire", "03_Financial"),
    ("Etats financiers N-1", "03_Financial"),
    ("Etats financiers N-2", "03_Financial"),
    ("Rapport d'audit", "03_Financial"),
    ("Attestation de bonne execution", "04_References"),
    ("CV du personnel cle", "05_CV"),
    ("Certification qualite", "06_Certifications"),
    ("Plaquette de presentation", "07_Company_Profile"),
]

DOC_COLS = ["Document", "Categorie", "Disponible", "Date_Emission",
            "Date_Expiration", "Jours_Avant_Expiration", "Statut",
            "Lien_Drive", "Responsable", "Notes"]

SUMMARY_COLS = ["Field_Key", "Section", "Label", "Valeur", "Requis", "Rempli"]

NOT_PROVIDED = "Non fourni"


def read_fields():
    with open(DATA / "org_profile_fields.csv", encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))
    corrupted = [r for r in rows if None in r or any(v is None for v in r.values())]
    if corrupted:
        raise ValueError(f"Registre de champs corrompu : {corrupted}")
    return rows


def header_row(ws, headers, row, widths):
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 26


# ------------------------------------------------------------ START_HERE --
def sheet_start_here(wb, today, n_fields):
    ws = wb.active
    ws.title = "START_HERE"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 108
    ws.sheet_view.showGridLines = False

    r = 2
    ws.cell(row=r, column=2, value="TENDERPILOT - ORGANIZATION PROFILE").font = F_TITLE
    r += 1
    ws.cell(row=r, column=2,
            value=f"Module 03 - version {MODULE_VERSION} - schema {S.SCHEMA_VERSION} - "
                  f"genere le {today.isoformat()}").font = F_MUTED
    r += 2

    def block(title, lines):
        nonlocal r
        ws.cell(row=r, column=2, value=title).font = F_H2
        r += 1
        for line in lines:
            c = ws.cell(row=r, column=2, value=line)
            c.font = F_BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1

    block("A QUOI SERT CE CLASSEUR", [
        "Il decrit une fois pour toutes ce que votre organisation est capable de prouver.",
        f"{n_fields} champs repartis en 5 sections, plus un suivi des documents "
        "administratifs et de leurs dates d'expiration.",
        "Il alimente le Go/No-Go (module 04) : sans profil rempli, aucune decision "
        "d'eligibilite n'est fiable.",
    ])
    block("PREMIERE ETAPE - CHOISIR LE TYPE DE CANDIDAT", [
        "Ouvrir l'onglet IDENTITY et renseigner la cellule Type de candidat en haut.",
        "Valeurs possibles : PME, CABINET, ONG, CONSULTANT, CONSORTIUM.",
        "La colonne Requis ? de chaque onglet se met alors a jour :",
        "   Requis      = champ obligatoire pour ce type de candidat.",
        "   Optionnel   = utile mais non bloquant.",
        "   Sans objet  = ne vous concerne pas, laisser vide.",
        "Les champs ne sont volontairement pas masques : un tableur ne masque pas de "
        "facon fiable, et un champ cache est un champ oublie.",
    ])
    block("REGLE DE SAISIE", [
        "Ne jamais inventer une valeur pour remplir une case.",
        "Un champ inconnu se laisse vide : PROFILE_SUMMARY affichera Non fourni.",
        "Un profil honnete avec des trous reste exploitable. Un profil invente produit "
        "un Go/No-Go faux et une offre rejetee.",
    ])
    block("COMPLETUDE N'EST PAS ELIGIBILITE", [
        "L'onglet COMPLETENESS mesure le pourcentage de champs requis renseignes.",
        "C'est un indicateur de preparation, PAS une eligibilite a un marche.",
        "Un profil complet a 100 % peut etre inelegible a un appel d'offres precis.",
        "L'eligibilite se decide dans le module 04 (Go/No-Go), face aux exigences "
        "d'un DAO reel.",
    ])
    block("ONGLETS", [
        "IDENTITY, FINANCE, EXPERIENCE, TEAM, CAPABILITIES : saisie.",
        "DOCUMENTS : disponibilite et expiration des pieces administratives.",
        "PROFILE_SUMMARY : vue normalisee, lisible par le Go/No-Go et par l'IA.",
        "COMPLETENESS : taux de remplissage par section.",
    ])
    return ws


# ------------------------------------------------------ onglets de saisie --
def sheet_section(wb, section, fields, ct_ref, is_identity):
    ws = wb.create_sheet(section)
    widths = [42, 34, 16, 46, 26, 20]

    r = 1
    ws.cell(row=r, column=1, value=SECTION_TITLES[section]).font = F_TITLE
    r += 1

    ct_row = None
    if is_identity:
        ws.cell(row=r, column=1, value="Type de candidat").font = F_LABEL
        c = ws.cell(row=r, column=2)
        c.font = Font(name="Calibri", size=11, bold=True, color="FF1F3A5F")
        c.border = BORDER
        ws.cell(row=r, column=4,
                value="Determine les champs Requis dans tous les onglets. "
                      "A renseigner en premier.").font = F_MUTED
        dv = DataValidation(type="list",
                            formula1='"' + ",".join(CANDIDATE_TYPES) + '"',
                            allow_blank=False)
        dv.errorTitle = "Type non valide"
        dv.error = "Choisir : " + ", ".join(CANDIDATE_TYPES)
        ws.add_data_validation(dv)
        dv.add(c)
        ct_row = r
        r += 2

    header_row(ws, FORM_COLS, r, widths)
    ws.freeze_panes = ws.cell(row=r + 1, column=1)
    r += 1
    first = r

    positions = {}
    for f in fields:
        ws.cell(row=r, column=1, value=f["Label"]).font = F_BODY
        v = ws.cell(row=r, column=2)
        v.font = F_BODY
        v.border = BORDER
        v.alignment = Alignment(vertical="top", wrap_text=True)

        # Requis ? depend du type de candidat saisi dans IDENTITY.
        req = ws.cell(row=r, column=3, value=(
            f'=IF({ct_ref}="","Type non defini",'
            f'IF(OR($E{r}="ALL",ISNUMBER(SEARCH({ct_ref},$E{r}))),'
            f'IF($F{r}="OUI","Requis","Optionnel"),"Sans objet"))'
        ))
        req.font = F_BODY
        req.alignment = Alignment(horizontal="center")

        ws.cell(row=r, column=4, value=f["Help"]).font = F_MUTED
        ws.cell(row=r, column=5, value=f["Applies_To"]).font = F_MUTED
        ws.cell(row=r, column=6, value=f["Required"]).font = F_MUTED
        positions[f["Field_Key"]] = r
        r += 1
    last = r - 1

    # Colonnes techniques du registre : presentes pour les formules, masquees.
    for letter in ("E", "F"):
        ws.column_dimensions[letter].hidden = True

    ws.conditional_formatting.add(
        f"A{first}:C{last}",
        FormulaRule(formula=[f'=$C{first}="Sans objet"'], fill=FILL_GREY,
                    stopIfTrue=True))
    ws.conditional_formatting.add(
        f"A{first}:C{last}",
        FormulaRule(formula=[f'=AND($C{first}="Requis",$B{first}="")'],
                    fill=FILL_RED))
    return ws, positions, ct_row


# ------------------------------------------------------------- DOCUMENTS --
def sheet_documents(wb):
    ws = wb.create_sheet("DOCUMENTS")
    widths = [38, 22, 16, 16, 16, 22, 24, 34, 20, 34]
    ws.cell(row=1, column=1, value="Documents administratifs").font = F_TITLE
    ws.cell(row=2, column=1,
            value="La colonne Categorie reprend l'arborescence du Document Vault "
                  "(module 05). Les lignes pre-remplies sont des TYPES de documents, "
                  "pas des donnees.").font = F_MUTED
    head = 3
    header_row(ws, DOC_COLS, head, widths)
    ws.freeze_panes = ws.cell(row=head + 1, column=1)
    first = head + 1

    for i, (name, category) in enumerate(DOCUMENT_TYPES):
        rr = first + i
        ws.cell(row=rr, column=1, value=name).font = F_BODY
        ws.cell(row=rr, column=2, value=category).font = F_BODY
        ws.cell(row=rr, column=3, value=NOT_PROVIDED).font = F_BODY

    last = first + MAX_DOC_ROWS - 1
    for rr in range(first, last + 1):
        d = ws.cell(row=rr, column=6)
        d.value = f'=IF($E{rr}="","",$E{rr}-TODAY())'
        d.font = F_BODY
        d.alignment = Alignment(horizontal="center")

        s = ws.cell(row=rr, column=7)
        s.value = (
            f'=IF($A{rr}="","",'
            f'IF($C{rr}<>"OUI","Manquant",'
            f'IF($E{rr}="","Sans expiration",'
            f'IF($F{rr}<0,"Expire",'
            f'IF($F{rr}<=15,"Expire sous 15 jours",'
            f'IF($F{rr}<=30,"Expire sous 30 jours","Valide"))))))'
        )
        s.font = F_BODY
        for col in (4, 5):
            ws.cell(row=rr, column=col).number_format = "yyyy-mm-dd"

    dv = DataValidation(type="list", formula1='"OUI,NON,Non fourni,En cours"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C{first}:C{last}")

    rng = f"A{first}:{get_column_letter(len(DOC_COLS))}{last}"
    for formula, fill in [
        (f'=$G{first}="Expire"', FILL_RED),
        (f'=$G{first}="Manquant"', dxf_fill("FFF5D0D0")),
        (f'=$G{first}="Expire sous 15 jours"', FILL_ORANGE),
        (f'=$G{first}="Expire sous 30 jours"', FILL_YELLOW),
        (f'=$G{first}="Valide"', FILL_GREEN),
    ]:
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=[formula], fill=fill, stopIfTrue=True))

    ws.auto_filter.ref = f"A{head}:{get_column_letter(len(DOC_COLS))}{last}"
    return ws, first, last


# ------------------------------------------------------- PROFILE_SUMMARY --
def sheet_summary(wb, fields, positions):
    ws = wb.create_sheet("PROFILE_SUMMARY")
    ws.cell(row=1, column=1, value="Profil normalise").font = F_TITLE
    ws.cell(row=2, column=1,
            value="Vue consommee par le Go/No-Go (module 04) et par l'analyseur IA "
                  "(module P). Aucune saisie ici : tout est repris des onglets de "
                  "saisie.").font = F_MUTED
    head = 3
    header_row(ws, SUMMARY_COLS, head, [26, 18, 42, 40, 16, 12])
    ws.freeze_panes = ws.cell(row=head + 1, column=1)

    r = head + 1
    first = r
    for f in fields:
        src = f["Section"]
        srow = positions[src][f["Field_Key"]]
        ws.cell(row=r, column=1, value=f["Field_Key"]).font = F_BODY
        ws.cell(row=r, column=2, value=src).font = F_BODY
        ws.cell(row=r, column=3, value=f["Label"]).font = F_BODY
        ws.cell(row=r, column=4,
                value=f'=IF({src}!$B${srow}="","{NOT_PROVIDED}",'
                      f"{src}!$B${srow})").font = F_BODY
        ws.cell(row=r, column=5, value=f"={src}!$C${srow}").font = F_BODY
        ws.cell(row=r, column=6,
                value=f'=IF({src}!$B${srow}="","NON","OUI")').font = F_BODY
        r += 1
    last = r - 1

    ws.conditional_formatting.add(
        f"A{first}:F{last}",
        FormulaRule(formula=[f'=$E{first}="Sans objet"'], fill=FILL_GREY,
                    stopIfTrue=True))
    ws.conditional_formatting.add(
        f"A{first}:F{last}",
        FormulaRule(formula=[f'=AND($E{first}="Requis",$F{first}="NON")'],
                    fill=FILL_RED))
    ws.auto_filter.ref = f"A{head}:F{last}"
    return ws, first, last


# ---------------------------------------------------------- COMPLETENESS --
def sheet_completeness(wb, sections, first, last, doc_first, doc_last):
    ws = wb.create_sheet("COMPLETENESS")
    for letter, width in (("A", 44), ("B", 16), ("C", 14), ("D", 14)):
        ws.column_dimensions[letter].width = width

    ws.cell(row=1, column=1, value="Completude du profil").font = F_TITLE
    warn = ws.cell(row=2, column=1,
                   value="Ce score mesure le remplissage du profil. Ce n'est PAS "
                         "une eligibilite.")
    warn.font = Font(name="Calibri", size=11, bold=True, color="FF9B1C1C")
    ws.cell(row=3, column=1,
            value="L'eligibilite se decide dans le module 04, face aux exigences d'un "
                  "DAO reel. Un profil complet a 100 % peut etre inelegible.").font = F_MUTED

    sec = f"PROFILE_SUMMARY!$B${first}:$B${last}"
    req = f"PROFILE_SUMMARY!$E${first}:$E${last}"
    fil = f"PROFILE_SUMMARY!$F${first}:$F${last}"

    r = 5
    header_row(ws, ["Section", "Requis", "Remplis", "Taux"], r, [44, 16, 14, 14])
    r += 1
    body_first = r
    for section in sections:
        ws.cell(row=r, column=1, value=SECTION_TITLES[section]).font = F_BODY
        ws.cell(row=r, column=2,
                value=f'=COUNTIFS({sec},"{section}",{req},"Requis")').font = F_BODY
        ws.cell(row=r, column=3,
                value=f'=COUNTIFS({sec},"{section}",{req},"Requis",'
                      f'{fil},"OUI")').font = F_BODY
        c = ws.cell(row=r, column=4, value=f'=IF(B{r}=0,"-",C{r}/B{r})')
        c.font = F_BODY
        c.number_format = "0%"
        r += 1
    body_last = r - 1

    r += 1
    total_row = r
    ws.cell(row=r, column=1, value="TOTAL CHAMPS REQUIS").font = F_LABEL
    ws.cell(row=r, column=2, value=f"=SUM(B{body_first}:B{body_last})").font = F_LABEL
    ws.cell(row=r, column=3, value=f"=SUM(C{body_first}:C{body_last})").font = F_LABEL
    r += 1
    ws.cell(row=r, column=1, value="SCORE DE COMPLETUDE").font = F_LABEL
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{total_row}=0,"Choisir le type de candidat",'
                      f"C{total_row}/B{total_row})")
    c.font = F_KPI
    c.number_format = "0%"
    r += 2

    ws.cell(row=r, column=1, value="CHAMPS REQUIS NON FOURNIS").font = F_H2
    r += 1
    ws.cell(row=r, column=1,
            value=f'=COUNTIFS({req},"Requis",{fil},"NON")').font = F_KPI
    ws.cell(row=r, column=2,
            value="Filtrer PROFILE_SUMMARY sur Requis = Requis et Rempli = NON "
                  "pour obtenir la liste.").font = F_MUTED
    r += 2

    ws.cell(row=r, column=1, value="DOCUMENTS").font = F_H2
    r += 1
    doc_status = f"DOCUMENTS!$G${doc_first}:$G${doc_last}"
    for label, status in [
        ("Documents manquants", "Manquant"),
        ("Documents expires", "Expire"),
        ("Expirent sous 15 jours", "Expire sous 15 jours"),
        ("Expirent sous 30 jours", "Expire sous 30 jours"),
        ("Documents valides", "Valide"),
    ]:
        ws.cell(row=r, column=1, value=label).font = F_LABEL
        ws.cell(row=r, column=2,
                value=f'=COUNTIF({doc_status},"{status}")').font = F_BODY
        r += 1
    return ws


# ------------------------------------------------------------------ docs --
README_TEMPLATE = """# Module 03 - TenderPilot Organization Profile

Version {version} - schema {schema} - genere le {date}

## Contenu

`03_TenderPilot_Organization_Profile.xlsx` - 9 onglets :
START_HERE, IDENTITY, FINANCE, EXPERIENCE, TEAM, CAPABILITIES, DOCUMENTS,
PROFILE_SUMMARY, COMPLETENESS.

{n_fields} champs de profil, {n_docs} types de documents pre-listes.

## Premiere etape

Ouvrir IDENTITY et renseigner **Type de candidat** (cellule B2) :
PME, CABINET, ONG, CONSULTANT ou CONSORTIUM.

Tant que cette cellule est vide, la colonne `Requis ?` de tous les onglets
affiche "Type non defini" et le score de completude ne se calcule pas.

## Le selecteur Candidate_Type

Il ne masque pas les champs, il les qualifie :

| Valeur affichee | Signification |
|-----------------|---------------|
| Requis | obligatoire pour ce type de candidat |
| Optionnel | utile, non bloquant |
| Sans objet | ne concerne pas ce type, laisser vide (ligne grisee) |

Le masquage dynamique de lignes n'est pas fiable d'un tableur a l'autre, et un
champ masque est un champ oublie. Les lignes "Sans objet" sont grisees, les
champs "Requis" encore vides sont surlignes en rouge.

## Completude n'est pas eligibilite

COMPLETENESS mesure le pourcentage de champs requis renseignes : un indicateur
de preparation interne.

L'eligibilite se decide dans le module 04 (Go/No-Go), face aux exigences d'un
DAO reel. Un profil rempli a 100 % peut etre inelegible a un marche precis, et
un profil rempli a 60 % peut etre parfaitement eligible a un autre. Les deux
notions ne doivent jamais etre affichees comme un chiffre unique.

## Regle de saisie

Ne jamais inventer une valeur pour remplir une case. Un champ inconnu reste
vide et PROFILE_SUMMARY affiche `{not_provided}`. C'est la contrainte du guide
agent : aucune donnee fabriquee.

## PROFILE_SUMMARY

Une ligne par champ :
`Field_Key | Section | Label | Valeur | Requis | Rempli`

`Field_Key` est un identifiant technique stable, consomme par le Go/No-Go
(module 04) et l'analyseur IA (module P). Ne pas le renommer sans mettre a jour
`data/org_profile_fields.csv`.

## DOCUMENTS

Statut calcule a partir de Disponible et Date_Expiration : Manquant,
Sans expiration, Expire, Expire sous 15 jours, Expire sous 30 jours, Valide.
Couleurs correspondantes, compteurs repris dans COMPLETENESS.

Les {n_docs} lignes pre-remplies sont des **types** de documents courants :
aucune date, aucun numero, aucun lien n'est fourni.

## Tests

    python tests/test_03_org_profile.py

## Limites connues

- Les champs "Sans objet" sont grises, pas masques (choix assume, voir plus haut).
- La liste des champs requis non fournis n'a pas d'onglet dedie : filtrer
  PROFILE_SUMMARY sur Requis = Requis et Rempli = NON.
- Le tableau DOCUMENTS couvre {n_docs_max} lignes.
- Les references detaillees et les experts vivent dans les modules 06 et 07 ;
  ce classeur ne stocke que les compteurs declares.

## Ajouter ou modifier un champ

Editer `data/org_profile_fields.csv` puis relancer `python build.py`.
Formulaire, PROFILE_SUMMARY et completude se regenerent ensemble : ils ne
peuvent pas diverger.

## Changelog

### {version} - {date}
- Version initiale.
- Registre de {n_fields} champs pilote par Candidate_Type.
- PROFILE_SUMMARY normalise et score de completude par section.
- Suivi des documents avec statut d'expiration calcule.
"""


def write_docs(today, n_fields):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / "README.md"
    path.write_text(
        README_TEMPLATE.format(
            version=MODULE_VERSION, schema=S.SCHEMA_VERSION, date=today.isoformat(),
            n_fields=n_fields, n_docs=len(DOCUMENT_TYPES), n_docs_max=MAX_DOC_ROWS,
            not_provided=NOT_PROVIDED,
        ),
        encoding="utf-8",
    )
    return path


# ----------------------------------------------------------------- build ---
def build():
    today = dt.date.today()
    fields = read_fields()
    sections = list(SECTION_TITLES)

    wb = Workbook()
    sheet_start_here(wb, today, len(fields))

    # IDENTITY!B2 porte le Type de candidat : c'est la reference de toutes les
    # formules "Requis ?" des autres onglets.
    ct_ref = "IDENTITY!$B$2"
    positions = {}
    for section in sections:
        sec_fields = [f for f in fields if f["Section"] == section]
        is_identity = section == "IDENTITY"
        _ws, pos, ct_row = sheet_section(wb, section, sec_fields, ct_ref, is_identity)
        if is_identity and ct_row != 2:
            raise AssertionError(
                f"Type de candidat attendu en IDENTITY!B2, trouve ligne {ct_row}")
        positions[section] = pos

    _doc_ws, doc_first, doc_last = sheet_documents(wb)
    _sum_ws, first, last = sheet_summary(wb, fields, positions)
    sheet_completeness(wb, sections, first, last, doc_first, doc_last)

    order = ["START_HERE"] + sections + ["DOCUMENTS", "PROFILE_SUMMARY",
                                         "COMPLETENESS"]
    wb._sheets = [wb[name] for name in order]
    wb.active = 0

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)
    write_docs(today, len(fields))
    return OUT_FILE


if __name__ == "__main__":
    print(f"OK  {build()}")
