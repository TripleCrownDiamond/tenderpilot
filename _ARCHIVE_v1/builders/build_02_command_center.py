"""
Builder du module 02 - TenderPilot Command Center.

Produit : dist/TenderPilot_Toolkit/02_COMMAND_CENTER/02_TenderPilot_Command_Center.xlsx

Le classeur est utilisable tel quel dans Excel ou importe dans Google Sheets
(Fichier > Importer). Aucune formule proprietaire a un seul tableur n'est
utilisee dans les cellules de calcul.

Dependances : schema/columns.py, data/*.csv, data/demo/*.csv
"""

import csv
import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from schema import columns as S

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT_DIR = ROOT / "dist" / "TenderPilot_Toolkit" / "02_COMMAND_CENTER"
OUT_FILE = OUT_DIR / "02_TenderPilot_Command_Center.xlsx"

MODULE_VERSION = "0.1.0"
MAX_ROWS = 1000  # lignes couvertes par validations, formules et couleurs

# ---------------------------------------------------------------- styles --
C_INK = "FF1F2933"
C_MUTED = "FF6B7280"
C_HEADER_BG = "FF1F3A5F"
C_BAND = "FFEEF2F7"

def argb(color):
    """Force le canal alpha a FF (opaque).

    openpyxl complete un code a 6 chiffres avec "00" - soit une couleur
    entierement TRANSPARENTE. Le fichier contient alors des remplissages que
    ni Excel ni Sheets n'affichent : les regles existent, elles peignent en
    invisible.
    """
    color = str(color).upper().lstrip("#")
    return color if len(color) == 8 else "FF" + color


F_TITLE = Font(name="Calibri", size=18, bold=True, color=C_HEADER_BG)
F_H2 = Font(name="Calibri", size=12, bold=True, color=C_HEADER_BG)
F_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
F_BODY = Font(name="Calibri", size=10, color=C_INK)
F_MUTED = Font(name="Calibri", size=9, color=C_MUTED, italic=True)
F_KPI = Font(name="Calibri", size=20, bold=True, color=C_HEADER_BG)
F_LABEL = Font(name="Calibri", size=10, bold=True, color=C_INK)

FILL_HEAD = PatternFill("solid", fgColor=argb(C_HEADER_BG))
FILL_BAND = PatternFill("solid", fgColor=argb(C_BAND))

# Palette deadlines - section 5 de la documentation fonctionnelle.
#
# Attention : une mise en forme CONDITIONNELLE est un "differential style".
# Excel y peint la couleur portee par bgColor, pas fgColor. Un fill construit
# avec fgColor seul s'ouvre sans aucune couleur visible. On renseigne donc les
# deux extremites pour ces regles.
def dxf_fill(color):
    return PatternFill(start_color=argb(color), end_color=argb(color),
                       fill_type="solid")


FILL_GREEN = dxf_fill("FFD8F3DC")
FILL_YELLOW = dxf_fill("FFFFF3BF")
FILL_ORANGE = dxf_fill("FFFFE0C2")
FILL_RED = dxf_fill("FFFFD6D6")
FILL_GREY = dxf_fill("FFECECEC")

THIN = Side(style="thin", color="FFD5DBE3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WIDTHS = {
    "Title": 46, "Organization": 26, "Notes": 40, "Next_Action": 30,
    "Missing_Documents": 30, "Source_URL": 34, "PDF_URL": 30, "Owner": 18,
    "Opportunity_Type": 20, "Subsector": 20, "Sector": 18, "Country": 16,
    "Region": 18, "Source_ID": 14, "Opportunity_ID": 16, "Source_Name": 32,
    "Details": 44, "Action": 26, "Module": 18,
}

OPP = "OPPORTUNITIES"


# ------------------------------------------------------------ utilitaires -
def read_csv(path):
    with open(path, encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh))


def offset_date(base, raw):
    """Convertit un offset DEMO exprime en jours en date reelle.

    Les jeux DEMO stockent des offsets et non des dates fixes, pour que les
    couleurs de deadline restent demonstratives quel que soit le jour de build.
    """
    if raw is None or str(raw).strip() == "":
        return None
    return base + dt.timedelta(days=int(raw))


def num(raw):
    if raw is None or str(raw).strip() == "":
        return None
    try:
        return float(raw) if "." in str(raw) else int(raw)
    except ValueError:
        return raw


def write_header(ws, headers, row=1, freeze=True):
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(name, 15)
    ws.row_dimensions[row].height = 28
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)


def add_list_validation(ws, letter, list_range):
    dv = DataValidation(type="list", formula1=list_range, allow_blank=True)
    dv.errorTitle = "Valeur non autorisee"
    dv.error = "Valeur hors liste. Utiliser la liste deroulante."
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}{MAX_ROWS + 1}")


def title_block(ws, title, subtitle=None, row=1):
    ws.cell(row=row, column=1, value=title).font = F_TITLE
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle).font = F_MUTED
    return row + 3


def kv_row(ws, row, label, value, note=None, fmt=None):
    ws.cell(row=row, column=1, value=label).font = F_LABEL
    c = ws.cell(row=row, column=2, value=value)
    c.font = F_BODY
    if fmt:
        c.number_format = fmt
    if note:
        ws.cell(row=row, column=3, value=note).font = F_MUTED
    return row + 1


# ---------------------------------------------------------------- LISTS ---
def sheet_lists(wb, sectors, countries, regions, currencies):
    """Onglet technique masque alimentant toutes les listes deroulantes."""
    ws = wb.create_sheet("LISTS")
    # Les en-tetes viennent du schema : le formulaire du module 12 y lit ses
    # listes deroulantes.
    lst = S.LISTS_COLUMNS
    cols = [
        (lst["status"], S.STATUSES),
        (lst["priority"], S.PRIORITIES),
        (lst["type"], S.OPPORTUNITY_TYPES),
        (lst["statusClosed"], S.STATUSES_CLOSED),
        (lst["yesNo"], S.YES_NO),
        (lst["sector"], sectors),
        (lst["country"], countries),
        (lst["region"], regions),
        (lst["currency"], currencies),
        (lst["language"], ["Francais", "Anglais", "Bilingue", "Autre"]),
        (lst["automationLevel"], S.AUTOMATION_LEVELS),
        (lst["linkStatus"], S.LINK_STATUSES),
    ]
    ranges = {}
    for i, (name, values) in enumerate(cols, start=1):
        letter = get_column_letter(i)
        h = ws.cell(row=1, column=i, value=name)
        h.font = F_HEAD
        h.fill = FILL_HEAD
        for j, v in enumerate(values, start=2):
            ws.cell(row=j, column=i, value=v).font = F_BODY
        ws.column_dimensions[letter].width = 22
        ranges[name] = (letter, len(values))
    ws.cell(row=1, column=len(cols) + 2,
            value="Onglet technique TenderPilot - ne pas supprimer, ne pas renommer.").font = F_MUTED
    ws.sheet_state = "hidden"
    return ranges


# -------------------------------------------------------- OPPORTUNITIES ---
def sheet_opportunities(wb, ranges, demo_rows, today):
    ws = wb.create_sheet(OPP)
    write_header(ws, S.OPPORTUNITIES)

    def col(name):
        return S.col_letter(S.OPPORTUNITIES, name)

    wrap_cols = ("Title", "Notes", "Missing_Documents", "Next_Action")
    for r, row in enumerate(demo_rows, start=2):
        vals = {
            "Opportunity_ID": row["Opportunity_ID"],
            "Added_At": today,
            "Title": row["Title"],
            "Organization": row["Organization"],
            "Country": row["Country"],
            "Region": row["Region"],
            "Sector": row["Sector"],
            "Subsector": row["Subsector"],
            "Opportunity_Type": row["Opportunity_Type"],
            "Source_ID": row["Source_ID"],
            "Source_URL": row["Source_URL"],
            "PDF_URL": row["PDF_URL"],
            "Publication_Date": offset_date(today, row["Publication_Offset_Days"]),
            "Deadline_Date": offset_date(today, row["Deadline_Offset_Days"]),
            "Deadline_Time": row["Deadline_Time"],
            "Budget": num(row["Budget"]),
            "Currency": row["Currency"],
            "Language": row["Language"],
            "Relevance_Score": num(row["Relevance_Score"]),
            "Eligibility_Score": num(row["Eligibility_Score"]),
            "Readiness_Score": num(row["Readiness_Score"]),
            "Eliminatory_Criterion": row["Eliminatory_Criterion"],
            "Status": row["Status"],
            "Priority": row["Priority"],
            "Owner": row["Owner"],
            "Next_Action": row["Next_Action"],
            "Next_Action_Date": offset_date(today, row["Next_Action_Offset_Days"]),
            "Missing_Documents": row["Missing_Documents"],
            "Notes": row["Notes"],
        }
        for i, name in enumerate(S.OPPORTUNITIES, start=1):
            if name == "Days_Remaining":
                continue
            c = ws.cell(row=r, column=i, value=vals.get(name))
            c.font = F_BODY
            c.alignment = Alignment(vertical="top", wrap_text=name in wrap_cols)
            if name.endswith("_Date") or name == "Added_At":
                c.number_format = "yyyy-mm-dd"
            if name == "Budget":
                c.number_format = "#,##0"

    # Days_Remaining : formule sur toute la plage, pas seulement sur le DEMO.
    dl = col("Deadline_Date")
    idx_days = S.OPPORTUNITIES.index("Days_Remaining") + 1
    for r in range(2, MAX_ROWS + 2):
        c = ws.cell(row=r, column=idx_days)
        c.value = '=IF(${d}{r}="","",${d}{r}-TODAY())'.format(d=dl, r=r)
        c.font = F_BODY
        c.alignment = Alignment(horizontal="center")

    def rng(key):
        letter, n = ranges[key]
        return "=LISTS!${l}$2:${l}${n}".format(l=letter, n=n + 1)

    for field, key in [
        ("Status", "Status"), ("Priority", "Priority"),
        ("Opportunity_Type", "Opportunity_Type"), ("Sector", "Sector"),
        ("Country", "Country"), ("Region", "Region"),
        ("Currency", "Currency"), ("Language", "Language"),
        ("Eliminatory_Criterion", "Yes_No"),
    ]:
        add_list_validation(ws, col(field), rng(key))

    for name in ("Relevance_Score", "Eligibility_Score", "Readiness_Score"):
        dv = DataValidation(type="whole", operator="between", formula1=0,
                            formula2=100, allow_blank=True)
        dv.errorTitle = "Score invalide"
        dv.error = "Un score doit etre compris entre 0 et 100."
        ws.add_data_validation(dv)
        letter = col(name)
        dv.add(f"{letter}2:{letter}{MAX_ROWS + 1}")

    # --- mise en forme conditionnelle ------------------------------------
    # L'ordre d'ajout EST la priorite. stopIfTrue empeche une regle plus
    # basse de repeindre une ligne deja traitee.
    st = col("Status")
    dr = col("Days_Remaining")
    last_col = get_column_letter(len(S.OPPORTUNITIES))
    full = f"A2:{last_col}{MAX_ROWS + 1}"

    # Les statuts clos sont enumeres dans la formule plutot que lus dans
    # l'onglet LISTS : Google Sheets refuse toute reference a une autre
    # feuille dans une mise en forme conditionnelle, et abandonne la regle
    # silencieusement a l'import. La liste vient du schema, elle ne peut donc
    # pas diverger de l'onglet LISTS.
    closed_test = "OR(" + ",".join(
        f'${st}2="{status}"' for status in S.STATUSES_CLOSED) + ")"

    rules = [
        # 1. Dossier clos, soumis ou NO-GO : jamais affiche comme urgent.
        (f"={closed_test}", FILL_GREY),
        # 2. Deadline depassee.
        (f'=AND(${dl}2<>"",${dr}2<0)', FILL_GREY),
        # 3 a 6. Paliers J-2 / J-7 / J-15 / au-dela.
        (f'=AND(${dl}2<>"",${dr}2<=2)', FILL_RED),
        (f'=AND(${dl}2<>"",${dr}2<=7)', FILL_ORANGE),
        (f'=AND(${dl}2<>"",${dr}2<=15)', FILL_YELLOW),
        (f'=AND(${dl}2<>"",${dr}2>15)', FILL_GREEN),
    ]
    for formula, fill in rules:
        ws.conditional_formatting.add(
            full, FormulaRule(formula=[formula], fill=fill, stopIfTrue=True)
        )

    # Critere eliminatoire : signal visuel distinct du score (doc section 10).
    el = col("Eliminatory_Criterion")
    ws.conditional_formatting.add(
        f"{el}2:{el}{MAX_ROWS + 1}",
        FormulaRule(formula=[f'=${el}2="OUI"'],
                    font=Font(bold=True, color="FF9B1C1C"), fill=FILL_RED),
    )

    ws.auto_filter.ref = f"A1:{last_col}{MAX_ROWS + 1}"
    return ws


def abs_range(sheet, letter, last, first=2):
    """Reference absolue complete : SHEET!$P$2:$P$1001.

    La borne de fin doit reporter sa lettre de colonne. Ecrire "$P$2:$1001"
    produit une reference que ni Excel ni Sheets n acceptent : la cellule
    affiche #ERROR! et tout ce qui en depend tombe avec elle.
    """
    return f"{sheet}!${letter}${first}:${letter}${last}"


# ------------------------------------------------------------- formules ---
def f_count_status(status):
    st = S.col_letter(S.OPPORTUNITIES, "Status")
    return f'=COUNTIF({OPP}!${st}$2:${st}${MAX_ROWS + 1},"{status}")'


def f_total():
    return f"=COUNTA({OPP}!$A$2:$A${MAX_ROWS + 1})"


def f_closed_count(ranges):
    st = S.col_letter(S.OPPORTUNITIES, "Status")
    letter, n = ranges["Status_Closed"]
    return (f"=SUMPRODUCT(COUNTIF({OPP}!${st}$2:${st}${MAX_ROWS + 1},"
            f"LISTS!${letter}$2:${letter}${n + 1}))")


def f_open_between(ranges, low, high):
    """Compte les lignes ouvertes dont Days_Remaining est dans [low, high]."""
    letter, n = ranges["Status_Closed"]
    last = MAX_ROWS + 1
    days = abs_range(OPP, S.col_letter(S.OPPORTUNITIES, "Days_Remaining"), last)
    status = abs_range(OPP, S.col_letter(S.OPPORTUNITIES, "Status"), last)
    closed = abs_range("LISTS", letter, n + 1)
    return (f'=SUMPRODUCT(({days}<>"")*({days}>={low})*({days}<={high})'
            f"*(COUNTIF({closed},{status})=0))")


def f_open_overdue(ranges):
    letter, n = ranges["Status_Closed"]
    last = MAX_ROWS + 1
    days = abs_range(OPP, S.col_letter(S.OPPORTUNITIES, "Days_Remaining"), last)
    status = abs_range(OPP, S.col_letter(S.OPPORTUNITIES, "Status"), last)
    closed = abs_range("LISTS", letter, n + 1)
    return (f'=SUMPRODUCT(({days}<>"")*({days}<0)'
            f"*(COUNTIF({closed},{status})=0))")


# ------------------------------------------------------------ DASHBOARD ---
def sheet_dashboard(wb, ranges):
    ws = wb.create_sheet("DASHBOARD")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 62
    r = title_block(ws, "TenderPilot - Dashboard",
                    "Valeurs calculees depuis l'onglet OPPORTUNITIES. Ne rien saisir ici.")

    ws.cell(row=r, column=1, value="INDICATEURS").font = F_H2
    r += 1
    kpis = [
        ("Opportunites enregistrees", f_total(), "Toutes lignes de OPPORTUNITIES."),
        ("Opportunites actives",
         "=" + f_total()[1:] + "-" + f_closed_count(ranges)[1:],
         "Total moins les statuts clos (Soumis, Gagne, Perdu, Expire, Archive, NO-GO)."),
        ("Nouvelles cette semaine",
         f'=COUNTIFS({OPP}!$B$2:$B${MAX_ROWS + 1},">="&TODAY()-7)',
         "Base sur Added_At."),
        ("Deadlines dans moins de 7 jours", f_open_between(ranges, 0, 7),
         "Statuts clos exclus - regle produit."),
        ("Deadlines depassees non traitees", f_open_overdue(ranges),
         "A archiver ou a passer en Expire."),
        ("GO", f_count_status("GO"), None),
        ("NO-GO", f_count_status("NO-GO"), None),
        ("En preparation", f_count_status("En preparation"), None),
        ("Pret a soumettre", f_count_status("Pret a soumettre"), None),
        ("Soumis", f_count_status("Soumis"), None),
        ("Gagne", f_count_status("Gagne"), None),
        ("Perdu", f_count_status("Perdu"), None),
    ]
    first_kpi = r
    for label, formula, note in kpis:
        r = kv_row(ws, r, label, formula, note)
    row_win = first_kpi + 10   # ligne du KPI "Gagne"
    row_loss = first_kpi + 11  # ligne du KPI "Perdu"

    r += 1
    ws.cell(row=r, column=1, value="Win rate").font = F_LABEL
    c = ws.cell(row=r, column=2,
                value=f'=IF(B{row_win}+B{row_loss}=0,"n/a",'
                      f"B{row_win}/(B{row_win}+B{row_loss}))")
    c.font = F_KPI
    c.number_format = "0%"
    ws.cell(row=r, column=3,
            value="Gagne / (Gagne + Perdu). Les dossiers en cours ne comptent pas.").font = F_MUTED
    r += 2

    # Repartitions : listes completes avec compteurs.
    sec = S.col_letter(S.OPPORTUNITIES, "Sector")
    cty = S.col_letter(S.OPPORTUNITIES, "Country")
    for title, list_key, opp_col in [
        ("REPARTITION PAR SECTEUR", "Sector", sec),
        ("REPARTITION PAR PAYS", "Country", cty),
    ]:
        ws.cell(row=r, column=1, value=title).font = F_H2
        r += 1
        letter, n = ranges[list_key]
        for i in range(2, n + 2):
            ws.cell(row=r, column=1, value=f"=LISTS!${letter}${i}").font = F_BODY
            ws.cell(row=r, column=2,
                    value=f"=COUNTIF({OPP}!${opp_col}$2:${opp_col}${MAX_ROWS + 1},A{r})").font = F_BODY
            r += 1
        r += 1

    ws.cell(row=r, column=1, value="REPARTITION PAR SOURCE").font = F_H2
    r += 1
    ws.cell(row=r, column=1,
            value="Voir onglet SOURCES, colonne Opportunites_Comptees.").font = F_MUTED
    return ws


# ------------------------------------------------------------- PIPELINE ---
def sheet_pipeline(wb):
    ws = wb.create_sheet("PIPELINE")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    r = title_block(ws, "Pipeline par statut",
                    "Vue de conversion. Statuts definis dans schema/columns.py.")
    write_header(ws, ["Statut", "Nombre", "Part"], row=r, freeze=False)
    r += 1
    total_ref = f_total()[1:]
    for status in S.STATUSES:
        ws.cell(row=r, column=1, value=status).font = F_BODY
        ws.cell(row=r, column=2, value=f_count_status(status)).font = F_BODY
        c = ws.cell(row=r, column=3, value=f'=IF({total_ref}=0,"",B{r}/{total_ref})')
        c.font = F_BODY
        c.number_format = "0%"
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = F_LABEL
    ws.cell(row=r, column=2, value=f_total()).font = F_LABEL
    return ws


# ------------------------------------------------------------ DEADLINES ---
def sheet_deadlines(wb, ranges):
    ws = wb.create_sheet("DEADLINES")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 70
    r = title_block(ws, "Deadlines",
                    "Compteurs par palier. Les statuts clos sont exclus de tous les paliers.")
    write_header(ws, ["Palier", "Nombre", "Regle"], row=r, freeze=False)
    r += 1
    for label, formula, note in [
        ("Deadline depassee", f_open_overdue(ranges),
         "Gris - a archiver ou passer en Expire."),
        ("J-1 (0 a 2 jours)", f_open_between(ranges, 0, 2), "Rouge - critique."),
        ("J-3 (3 a 7 jours)", f_open_between(ranges, 3, 7), "Orange - urgent."),
        ("J-7 (8 a 15 jours)", f_open_between(ranges, 8, 15), "Jaune - a preparer."),
        ("J-14 (plus de 15 jours)", f_open_between(ranges, 16, 3650),
         "Vert - a surveiller."),
    ]:
        ws.cell(row=r, column=1, value=label).font = F_LABEL
        c = ws.cell(row=r, column=2, value=formula)
        c.font = F_KPI
        ws.cell(row=r, column=3, value=note).font = F_BODY
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="LISTE DETAILLEE").font = F_H2
    r += 1
    for line in [
        "Limite connue : un fichier .xlsx ne peut pas generer dynamiquement une liste filtree.",
        "Trois facons d'obtenir la liste detaillee :",
        "1. Filtrer l'onglet OPPORTUNITIES sur Days_Remaining (le filtre est deja actif).",
        "2. Dans Google Sheets, coller la formule ci-dessous dans cet onglet.",
        "3. Laisser Apps Script (module 12) remplir cet onglet chaque jour.",
    ]:
        ws.cell(row=r, column=1, value=line).font = F_BODY
        r += 1
    r += 1
    dr = S.col_letter(S.OPPORTUNITIES, "Days_Remaining")
    ws.cell(row=r, column=1, value="Formule Google Sheets a coller :").font = F_LABEL
    r += 1
    c = ws.cell(row=r, column=1, value=(
        f'QUERY({OPP}!A2:AD; "select A,C,D,E,N,{dr},X where {dr} is not null '
        f'and {dr} <= 15 order by {dr} asc"; 0)'
    ))
    c.font = Font(name="Consolas", size=9, color=C_INK)
    ws.cell(row=r + 1, column=1,
            value="Ajouter le signe = devant en collant. Volontairement stockee "
                  "en texte : QUERY n'existe pas dans Excel et y produirait une "
                  "erreur.").font = F_MUTED
    return ws


# ------------------------------------------------------------ WATCHLIST ---
def sheet_watchlist(wb, ranges, countries, sectors):
    ws = wb.create_sheet("WATCHLIST")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 60
    r = title_block(ws, "Watchlist",
                    "Definit ce qui est pertinent pour VOUS. Alimente le moteur de "
                    "scoring (module O) et les notifications.")

    # Libelles issus de schema/columns.py : le moteur de pertinence du
    # module 12 retrouve chaque parametre par son libelle.
    wl = S.WATCHLIST_LABELS
    for title, fields in [
        ("GEOGRAPHIE", [
            (wl["region"], "", "Ex : Afrique de l'Ouest."),
            (wl["international"], "OUI", "OUI / NON."),
        ]),
        ("PARAMETRES", [
            (wl["min_budget"], "", "Laisser vide si sans objet."),
            (wl["max_budget"], "", "Laisser vide si sans objet."),
            (wl["min_days"], 7,
             "En dessous, l'opportunite est jugee non preparable."),
            (wl["language"], "Francais", None),
        ]),
        ("MOTS-CLES", [
            (wl["positive_keywords"], "", "Separes par des points-virgules."),
            (wl["negative_keywords"], "",
             "Exclusion immediate. Separes par des points-virgules."),
        ]),
    ]:
        ws.cell(row=r, column=1, value=title).font = F_H2
        r += 1
        for label, value, note in fields:
            r = kv_row(ws, r, label, value, note)
        r += 1

    ws.cell(row=r, column=1, value=S.WATCHLIST_TARGETS_TITLE).font = F_H2
    r += 1
    ws.cell(row=r, column=1,
            value="Tout est decoche au depart. Cochez ce que vous voulez voir "
                  "remonter : rien d'autre ne sera signale.").font = F_MUTED
    r += 1
    header = r
    start = header + 1

    # Une colonne par dimension, suivie de sa case a cocher.
    sources = {"countries": countries, "sectors": sectors,
               "types": S.OPPORTUNITY_TYPES}
    for bloc in S.WATCHLIST_BLOCKS:
        c0 = bloc["column"]
        values = sources[bloc["key"]]

        for offset, (title, width) in enumerate(
                ((bloc["header"], 30), (S.WATCHLIST_CHECK_HEADER, 10))):
            cell = ws.cell(row=header, column=c0 + offset, value=title)
            cell.font = F_HEAD
            cell.fill = FILL_HEAD
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(c0 + offset)].width = width

        for i, value in enumerate(values):
            row = start + i
            v = ws.cell(row=row, column=c0, value=value)
            v.font = F_BODY
            v.border = BORDER
            box = ws.cell(row=row, column=c0 + 1, value="NON")
            box.font = F_BODY
            box.border = BORDER
            box.alignment = Alignment(horizontal="center")

        letter = get_column_letter(c0 + 1)
        last = start + len(values) - 1
        dv = DataValidation(type="list", formula1='"OUI,NON"', allow_blank=True)
        dv.errorTitle = "Valeur non autorisee"
        dv.error = "Cocher revient a saisir OUI."
        ws.add_data_validation(dv)
        dv.add(f"{letter}{start}:{letter}{last}")

        # Une ligne cochee se voit d'un coup d'oeil.
        ws.conditional_formatting.add(
            f"{get_column_letter(c0)}{start}:{letter}{last}",
            FormulaRule(formula=[f'=${letter}{start}="OUI"'], fill=FILL_GREEN))

        # Colonne de separation entre deux blocs.
        if c0 + 2 <= 7:
            ws.column_dimensions[get_column_letter(c0 + 2)].width = 3

    return ws


# ------------------------------------------------------------- SETTINGS ---
def sheet_settings(wb):
    ws = wb.create_sheet("SETTINGS")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 70
    r = title_block(ws, "Parametres",
                    "Lus par les modules d'automatisation (Apps Script, Telegram, IA).")

    # Les libelles viennent de schema/columns.py : le projet Apps Script les
    # utilise pour retrouver chaque reglage. Les modifier ici seulement
    # casserait le script sans prevenir.
    lbl = S.SETTINGS_LABELS
    for title, fields in [
        ("GENERAL", [
            (lbl["org_name"], "", "Affiche dans les notifications."),
            (lbl["timezone"], "Africa/Porto-Novo",
             "Format IANA. Utilise pour les rappels."),
            (lbl["language"], "Francais", None),
            (lbl["schema_version"], S.SCHEMA_VERSION,
             "Ne pas modifier manuellement."),
            (lbl["module_version"], MODULE_VERSION,
             "Ne pas modifier manuellement."),
        ]),
        ("SEUILS", [
            (lbl["threshold_review"], 60,
             "Score Relevance en dessous duquel on ignore."),
            (lbl["threshold_ai"], 80,
             "L'IA n'est appelee qu'au-dessus. Evite de bruler des tokens "
             "(doc section 7)."),
            (lbl["document_alert_days"], 30, None),
        ]),
        ("NOTIFICATIONS", [
            (lbl["notify_email"], "", "Laisser vide pour desactiver."),
            (lbl["reminders_enabled"], "OUI", "OUI / NON."),
            (lbl["reminder_days"], "14;7;3;1",
             "Separes par des points-virgules."),
            (lbl["telegram_enabled"], "NON", "Module optionnel L."),
            (lbl["gmail_label"], "TenderPilot",
             "Le script lit les emails portant ce label (module 12)."),
        ]),
        ("INTELLIGENCE ARTIFICIELLE", [
            (lbl["ai_provider"], "",
             "Mistral / Gemini / OpenAI / Claude. Vide = IA desactivee."),
            (lbl["ai_model"], "", None),
            (lbl["ai_temperature"], 0.2, "Basse pour l'extraction factuelle."),
            (lbl["ai_max_tokens"], 4000, None),
        ]),
    ]:
        ws.cell(row=r, column=1, value=title).font = F_H2
        r += 1
        for label, value, note in fields:
            r = kv_row(ws, r, label, value, note)
        r += 1

    warn = ws.cell(row=r, column=1, value="NE PAS stocker de cle API dans ce fichier.")
    warn.font = Font(name="Calibri", size=11, bold=True, color="FF9B1C1C")
    ws.cell(row=r + 1, column=1,
            value="La cle se place dans les Proprietes du script Apps Script ou dans une "
                  "variable d'environnement. Un classeur partage expose la cle a tous ses "
                  "lecteurs (checklist de lancement, section Securite).").font = F_BODY
    return ws


# -------------------------------------------------------------- SOURCES ---
def sheet_sources(wb, ranges, demo_sources):
    ws = wb.create_sheet("SOURCES")
    headers = S.SOURCES + ["Opportunites_Comptees"]
    write_header(ws, headers)
    for r, row in enumerate(demo_sources, start=2):
        for i, name in enumerate(S.SOURCES, start=1):
            c = ws.cell(row=r, column=i, value=row.get(name, ""))
            c.font = F_BODY
            c.alignment = Alignment(vertical="top", wrap_text=name == "Notes")

    src = S.col_letter(S.OPPORTUNITIES, "Source_ID")
    last = len(headers)
    for r in range(2, MAX_ROWS + 2):
        c = ws.cell(row=r, column=last)
        c.value = (f'=IF($A{r}="","",'
                   f"COUNTIF({OPP}!${src}$2:${src}${MAX_ROWS + 1},$A{r}))")
        c.font = F_BODY
    ws.column_dimensions[get_column_letter(last)].width = 22

    def rng(key):
        letter, n = ranges[key]
        return "=LISTS!${l}$2:${l}${n}".format(l=letter, n=n + 1)

    for field, key in [
        ("Country", "Country"), ("Region", "Region"),
        ("Opportunity_Type", "Opportunity_Type"),
        ("Automation_Level", "Automation_Level"),
        ("Email_Alerts", "Yes_No"), ("RSS_Available", "Yes_No"),
        ("API_Available", "Yes_No"), ("Registration_Required", "Yes_No"),
        ("Link_Status", "Link_Status"), ("Active", "Yes_No"),
    ]:
        add_list_validation(ws, S.col_letter(S.SOURCES, field), rng(key))

    ls = S.col_letter(S.SOURCES, "Link_Status")
    full = f"A2:{get_column_letter(last)}{MAX_ROWS + 1}"
    ws.conditional_formatting.add(
        full, FormulaRule(formula=[f'=${ls}2="Casse"'], fill=FILL_RED, stopIfTrue=True))
    ws.conditional_formatting.add(
        full, FormulaRule(formula=[f'=${ls}2="A verifier"'], fill=FILL_YELLOW))
    ws.auto_filter.ref = f"A1:{get_column_letter(last)}{MAX_ROWS + 1}"
    return ws


# ----------------------------------------------------------------- LOGS ---
def sheet_logs(wb, today):
    ws = wb.create_sheet("LOGS")
    write_header(ws, S.LOGS)
    demo = ["LOG-0001", today, "INFO", "BUILD", "Creation du classeur", "",
            "builder", f"Genere par build_02_command_center.py v{MODULE_VERSION}"]
    for i, v in enumerate(demo, start=1):
        c = ws.cell(row=2, column=i, value=v)
        c.font = F_BODY
        if i == 2:
            c.number_format = "yyyy-mm-dd"
    ws.cell(row=4, column=1,
            value="Rempli automatiquement par Apps Script. Ne pas supprimer les entetes.").font = F_MUTED
    return ws


# ----------------------------------------------------------- START_HERE ---
def sheet_start_here(wb, today):
    ws = wb.active
    ws.title = "START_HERE"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110
    ws.sheet_view.showGridLines = False

    r = 2
    ws.cell(row=r, column=2, value="TENDERPILOT - COMMAND CENTER").font = F_TITLE
    r += 1
    ws.cell(row=r, column=2,
            value=f"Module 02 - version {MODULE_VERSION} - schema {S.SCHEMA_VERSION} - "
                  f"genere le {today.isoformat()}").font = F_MUTED
    r += 2

    def block(title, lines):
        nonlocal r
        ws.cell(row=r, column=2, value=title).font = F_H2
        r += 1
        for line in lines:
            c = ws.cell(row=r, column=2, value=line)
            c.font = F_BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1

    block("A QUOI SERT CE CLASSEUR", [
        "C'est le cockpit central de TenderPilot : toutes les opportunites y sont suivies, "
        "de la detection a la soumission.",
        "Il fonctionne seul, sans Python, sans Telegram et sans IA. Les modules "
        "d'automatisation ne font qu'accelerer ce qu'il permet deja.",
    ])
    block("PREMIERE UTILISATION - 8 ETAPES", [
        "1. Supprimer les lignes DEMO de OPPORTUNITIES (identifiants commencant par DEMO-).",
        "2. Supprimer les lignes DEMO de SOURCES et saisir vos sources reelles verifiees.",
        "3. Remplir WATCHLIST : pays, secteurs, types, mots-cles, budget.",
        "4. Remplir SETTINGS : organisation, fuseau horaire, seuils, email.",
        "5. Ajouter votre premiere opportunite reelle dans OPPORTUNITIES.",
        "6. Verifier que la couleur de la ligne correspond au delai restant.",
        "7. Consulter DASHBOARD et DEADLINES.",
        "8. Optionnel : installer le module 12 (Apps Script) pour les rappels automatiques.",
    ])
    block("LE MENU TENDERPILOT", [
        "Ce classeur fonctionne entierement sans menu : couleurs, formules, "
        "listes deroulantes et tableau de bord marchent des l'ouverture.",
        "Le menu TenderPilot - formulaire de saisie, rappels automatiques, "
        "import des alertes email - est un module supplementaire (module 12).",
        "Il n'existe QUE dans Google Sheets, et seulement apres installation "
        "du script. Dans Excel, il n'apparaitra jamais : Apps Script est une "
        "technologie Google, Excel ne sait pas l'executer.",
        "Si vous travaillez dans Google Sheets et que le menu n'apparait pas, "
        "c'est que le script n'a pas encore ete installe. Voir MISE_EN_LIGNE.md.",
    ])
    block("CODE COULEUR DES DEADLINES", [
        "Vert : plus de 15 jours restants.",
        "Jaune : 8 a 15 jours.",
        "Orange : 3 a 7 jours.",
        "Rouge : 0 a 2 jours.",
        "Gris : deadline depassee, OU dossier clos (Soumis, Gagne, Perdu, Expire, "
        "Archive, NO-GO).",
        "Regle importante : un dossier deja soumis n'apparait jamais en rouge, meme la "
        "veille de la deadline.",
    ])
    block("LES TROIS SCORES SONT SEPARES", [
        "Relevance_Score : est-ce interessant pour nous ?",
        "Eligibility_Score : sommes-nous eligibles ?",
        "Readiness_Score : avons-nous les documents et l'equipe des maintenant ?",
        "Ils ne doivent jamais etre fusionnes en un chiffre unique : cela masquerait un "
        "blocage reel.",
        "Eliminatory_Criterion = OUI l'emporte sur n'importe quel score. Un score de 95 "
        "avec un critere eliminatoire reste un NO-GO.",
    ])
    block("ONGLETS", [
        "START_HERE : cette page.",
        "DASHBOARD : indicateurs et repartitions, en lecture seule.",
        "OPPORTUNITIES : la table de travail principale.",
        "PIPELINE : conversion par statut.",
        "DEADLINES : compteurs par palier.",
        "WATCHLIST : vos criteres de pertinence.",
        "SETTINGS : parametres techniques.",
        "SOURCES : vos sources de veille actives.",
        "LOGS : journal des actions automatisees.",
        "LISTS : onglet technique masque. Ne pas supprimer, ne pas renommer.",
    ])
    block("AVERTISSEMENTS", [
        "Les donnees DEMO sont fictives (organisations DEMO, URL example.org). Elles "
        "servent uniquement a verifier que les formules et les couleurs fonctionnent.",
        "Ne jamais stocker de cle API dans ce classeur.",
        "Ne pas renommer les colonnes de OPPORTUNITIES : les modules d'automatisation "
        "s'y referent.",
    ])
    return ws


# ------------------------------------------------------ doc de livraison --
README_TEMPLATE = """# Module 02 - TenderPilot Command Center

Version {version} - schema {schema} - genere le {date}

## Contenu

`02_TenderPilot_Command_Center.xlsx` - 10 onglets :
START_HERE, DASHBOARD, OPPORTUNITIES, PIPELINE, DEADLINES, WATCHLIST,
SETTINGS, SOURCES, LOGS, et LISTS (technique, masque).

## Installation

### Excel
Ouvrir le fichier. Rien d'autre a faire.

### Google Sheets (recommande pour l'automatisation)
1. Google Drive > Nouveau > Importer le fichier.
2. Ouvrir avec Google Sheets.
3. Fichier > Parametres > verifier le fuseau horaire.
4. Verifier que l'onglet OPPORTUNITIES affiche bien des couleurs de ligne.

L'onglet LISTS reste masque apres import. Ne pas le supprimer : toutes les
listes deroulantes y font reference.

## Premiere configuration

1. Supprimer les 10 lignes DEMO de OPPORTUNITIES (identifiants `DEMO-*`).
2. Supprimer les 3 lignes DEMO de SOURCES, saisir vos sources reelles.
3. Remplir WATCHLIST puis SETTINGS.
4. Ajouter une premiere opportunite reelle et verifier la couleur de ligne.

## Jeu de demonstration

Les 10 lignes `DEMO-*` sont entierement fictives : organisations nommees
"DEMO Organisation X", URL en `example.org`. Elles existent pour prouver que
les formules et les couleurs fonctionnent, et couvrent :

| Ligne | Ce qu'elle teste |
|-------|------------------|
| DEMO-001 | cas normal, deadline lointaine (vert) |
| DEMO-002 | palier 8-15 jours (jaune) |
| DEMO-003 | palier 3-7 jours (orange) + documents manquants |
| DEMO-004 | palier 0-2 jours (rouge) |
| DEMO-005 | deadline depassee (gris) |
| DEMO-006 | statut Soumis a J+2 : ne doit PAS apparaitre en rouge |
| DEMO-007 | dossier gagne (win rate) |
| DEMO-008 | dossier perdu (win rate) |
| DEMO-009 | score 85 mais critere eliminatoire : reste NO-GO |
| DEMO-010 | champs manquants (scores, budget, devise vides) |

Les deadlines DEMO sont recalculees a chaque build a partir d'offsets en
jours : les couleurs restent demonstratives quelle que soit la date.

## Tests

    python tests/test_02_command_center.py

40 verifications. Les tests relisent les formules et les regles de couleur
reellement ecrites dans le .xlsx, puis les evaluent sur le jeu DEMO.

## Regles produit implementees

- Un dossier clos (Soumis, Gagne, Perdu, Expire, Archive, NO-GO) n'est jamais
  affiche comme urgent, meme la veille de la deadline. Cette regle a la
  priorite maximale sur toutes les couleurs.
- Les trois scores (Relevance, Eligibility, Readiness) restent separes. Aucun
  chiffre unique ne les fusionne.
- `Eliminatory_Criterion` dispose de son propre signal visuel, independant des
  scores.
- Aucun champ de saisie de cle API. SETTINGS porte un avertissement explicite.

## Limites connues

- Un fichier .xlsx ne peut pas generer dynamiquement une liste filtree :
  l'onglet DEADLINES affiche des compteurs, pas la liste detaillee. Trois
  solutions sont documentees dans l'onglet lui-meme.
- Formules et couleurs couvrent les lignes 2 a {max_rows}. Au-dela, etendre
  la plage ou regenerer le fichier avec un `MAX_ROWS` superieur.
- Les 3 sources DEMO sont fictives. Le fichier de sources reelles est le
  module 01, a produire separement.
- Les rappels automatiques, la lecture Gmail et Telegram relevent des modules
  12 (Apps Script) et L (Telegram), optionnels.

## Regenerer

    python build.py

Ne jamais modifier ce classeur a la main pour corriger une formule : modifier
`schema/columns.py`, `data/*.csv` ou le builder, puis relancer le build.

## Changelog

### {version} - {date}
- Version initiale.
- 10 onglets, 30 colonnes OPPORTUNITIES, listes deroulantes, bornes de score.
- 6 regles de couleur exclusives + regle dediee au critere eliminatoire.
- Dashboard (12 indicateurs + win rate), pipeline, compteurs de deadlines.
- Jeu DEMO de 10 lignes et suite de 40 tests.
"""


def write_docs(today):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / "README.md"
    path.write_text(
        README_TEMPLATE.format(version=MODULE_VERSION, schema=S.SCHEMA_VERSION,
                               date=today.isoformat(), max_rows=MAX_ROWS + 1),
        encoding="utf-8",
    )
    return path


# ---------------------------------------------------------------- build ---
def build():
    today = dt.date.today()
    sectors = [r["Sector"] for r in read_csv(DATA / "sectors.csv")]
    countries_rows = read_csv(DATA / "countries.csv")
    countries = [r["Country"] for r in countries_rows]
    regions = sorted({r["Region"] for r in countries_rows})
    currencies = [r["Currency"] for r in read_csv(DATA / "currencies.csv")]
    demo_opps = read_csv(DATA / "demo" / "opportunities_demo.csv")
    demo_srcs = read_csv(DATA / "demo" / "sources_demo.csv")

    wb = Workbook()
    sheet_start_here(wb, today)
    ranges = sheet_lists(wb, sectors, countries, regions, currencies)
    sheet_dashboard(wb, ranges)
    sheet_opportunities(wb, ranges, demo_opps, today)
    sheet_pipeline(wb)
    sheet_deadlines(wb, ranges)
    sheet_watchlist(wb, ranges, countries, sectors)
    sheet_settings(wb)
    sheet_sources(wb, ranges, demo_srcs)
    sheet_logs(wb, today)

    order = ["START_HERE", "DASHBOARD", OPP, "PIPELINE", "DEADLINES",
             "WATCHLIST", "SETTINGS", "SOURCES", "LOGS", "LISTS"]
    wb._sheets = [wb[name] for name in order]
    wb.active = 0

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)
    write_docs(today)
    return OUT_FILE


if __name__ == "__main__":
    print(f"OK  {build()}")
