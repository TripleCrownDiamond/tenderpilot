"""
Builder du module 04 - TenderPilot Go / No-Go Engine.

Produit : dist/TenderPilot_Toolkit/04_GO_NO_GO/04_TenderPilot_Go_NoGo.xlsx

Regle centrale du module, non negociable :

    un critere ELIMINATOIRE non satisfait ne peut JAMAIS etre compense
    par un score, aussi eleve soit-il.

Le verdict n'est donc pas un seuil sur une moyenne. C'est une cascade dont
les criteres eliminatoires occupent le premier etage, avant meme que le
score soit regarde.

Les trois scores (Relevance, Eligibility, Readiness) sont affiches
separement et ne sont jamais additionnes en un chiffre unique.

Dependances : schema/columns.py, data/demo/gonogo_demo.csv
"""

import csv
import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from schema import columns as S
from builders.build_02_command_center import (
    abs_range, BORDER, FILL_GREEN, FILL_GREY, FILL_HEAD, FILL_ORANGE, FILL_RED,
    FILL_YELLOW, F_BODY, F_H2, F_HEAD, F_KPI, F_LABEL, F_MUTED, F_TITLE,
)

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT_DIR = ROOT / "dist" / "TenderPilot_Toolkit" / "04_GO_NO_GO"
OUT_FILE = OUT_DIR / "04_TenderPilot_Go_NoGo.xlsx"

MODULE_VERSION = "0.1.0"
MAX_ROWS = 300
DEFAULT_GO_THRESHOLD = 70

WIDTHS = {
    "Criterion_ID": 14, "Category": 16, "Requirement": 56, "Source_Page": 16,
    "Mandatory": 12, "Eliminatory": 14, "Candidate_Evidence": 40,
    "Evidence_Link": 26, "Result": 22, "Risk": 12, "Weight": 10, "Score": 10,
    "Action": 40, "Owner": 18, "Poids_Effectif": 14, "Points": 12,
    "Poids_Readiness": 16, "Points_Readiness": 16,
}

TECHNICAL_COLS = ["Poids_Effectif", "Points", "Poids_Readiness",
                  "Points_Readiness"]

CRIT = "CRITERIA"


def col(name):
    return S.col_letter(S.GO_NOGO, name)


def read_demo():
    with open(DATA / "demo" / "gonogo_demo.csv", encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))
    corrupted = [r for r in rows if None in r or any(v is None for v in r.values())]
    if corrupted:
        raise ValueError(f"Jeu DEMO corrompu : {corrupted}")
    return rows


# --------------------------------------------------------------------------
# Logique de decision, en Python.
#
# Elle est la reference : le classeur en est la traduction en formules, et
# les tests comparent les deux sur une serie de scenarios. Toute evolution de
# la regle se fait ici d'abord.
# --------------------------------------------------------------------------
def verdict_for(criteria, go_threshold=DEFAULT_GO_THRESHOLD):
    """Retourne (verdict, motif, eligibility_score)."""
    assessed = [c for c in criteria if c.get("Criterion_ID")]
    if not assessed:
        return "Aucun critere saisi", "Aucun critere n'a ete renseigne.", None

    unassessed = [c for c in assessed if not c.get("Result")]
    elim_failed = [c for c in assessed
                   if c["Eliminatory"] == "OUI"
                   and c.get("Result") == "Non satisfait"]
    elim_unclear = [c for c in assessed
                    if c["Eliminatory"] == "OUI"
                    and c.get("Result") in ("Partiellement satisfait",
                                            "A verifier")]
    mandatory_gaps = [c for c in assessed
                      if c["Mandatory"] == "OUI" and c["Eliminatory"] == "NON"
                      and c.get("Result") in ("Non satisfait",
                                              "Partiellement satisfait",
                                              "A verifier")]

    total_weight = sum(int(c["Weight"]) for c in assessed
                       if c.get("Result") and c.get("Weight"))
    points = sum(int(c["Weight"]) * S.RESULT_SCORES[c["Result"]]
                 for c in assessed if c.get("Result") and c.get("Weight"))
    score = round(points / total_weight, 1) if total_weight else None

    # L'ordre de ces tests EST la regle produit.
    if elim_failed:
        ids = ", ".join(c["Criterion_ID"] for c in elim_failed)
        return ("NO_GO",
                f"Critere eliminatoire non satisfait ({ids}). "
                f"Aucun score ne peut compenser.", score)
    if elim_unclear:
        ids = ", ".join(c["Criterion_ID"] for c in elim_unclear)
        return ("NO_GO_CONDITIONAL",
                f"Critere eliminatoire non tranche ({ids}). "
                f"A clarifier avant toute decision.", score)
    if unassessed:
        return ("NO_GO_CONDITIONAL",
                f"{len(unassessed)} critere(s) sans resultat. "
                f"Evaluation incomplete.", score)
    if score is None or score < go_threshold:
        return ("NO_GO",
                f"Score d'eligibilite insuffisant ({score} < {go_threshold}).",
                score)
    if mandatory_gaps:
        return ("GO_WITH_ACTIONS",
                f"{len(mandatory_gaps)} exigence(s) obligatoire(s) a corriger "
                f"avant depot.", score)
    return "GO", "Tous les criteres obligatoires sont satisfaits.", score


# ------------------------------------------------------------------ LISTS --
def sheet_lists(wb):
    ws = wb.create_sheet("LISTS")
    cols = [
        ("Category", S.CRITERION_CATEGORIES),
        ("Result", S.CRITERION_RESULTS),
        ("Risk", S.RISK_LEVELS),
        ("Yes_No", ["OUI", "NON"]),
        ("Readiness_Categories", S.READINESS_CATEGORIES),
    ]
    ranges = {}
    for i, (name, values) in enumerate(cols, start=1):
        letter = get_column_letter(i)
        h = ws.cell(row=1, column=i, value=name)
        h.font = F_HEAD
        h.fill = FILL_HEAD
        for j, v in enumerate(values, start=2):
            ws.cell(row=j, column=i, value=v).font = F_BODY
        ws.column_dimensions[letter].width = 24
        ranges[name] = (letter, len(values))
    ws.cell(row=1, column=len(cols) + 2,
            value="Onglet technique - ne pas supprimer.").font = F_MUTED
    ws.sheet_state = "hidden"
    return ranges


# --------------------------------------------------------------- CRITERIA --
def sheet_criteria(wb, ranges, demo):
    ws = wb.create_sheet(CRIT)
    for i, name in enumerate(S.GO_NOGO, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[name]
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = ws.cell(row=2, column=4)

    computed = set(TECHNICAL_COLS) | {"Score"}
    for r, row in enumerate(demo, start=2):
        for i, name in enumerate(S.GO_NOGO, start=1):
            if name in computed:
                continue
            value = row.get(name, "")
            if name == "Weight" and value:
                value = int(value)
            c = ws.cell(row=r, column=i, value=value)
            c.font = F_BODY
            c.alignment = Alignment(vertical="top",
                                    wrap_text=name in ("Requirement",
                                                       "Candidate_Evidence",
                                                       "Action"))

    cid, cat = col("Criterion_ID"), col("Category")
    res, wgt, sco = col("Result"), col("Weight"), col("Score")
    readiness_test = "OR(" + ",".join(
        f'${cat}{{r}}="{c}"' for c in S.READINESS_CATEGORIES) + ")"

    for r in range(2, MAX_ROWS + 2):
        # Score : derive du Result, jamais saisi. Un chiffre libre serait la
        # porte ouverte a l'arrangement du resultat.
        ws.cell(row=r, column=S.GO_NOGO.index("Score") + 1, value=(
            f'=IF(${res}{r}="","",'
            f'IF(${res}{r}="Satisfait",100,'
            f'IF(${res}{r}="Partiellement satisfait",50,0)))'
        )).font = F_BODY

        # Colonnes techniques : 0 sur les lignes vides, pour que les moyennes
        # ponderees ne propagent pas d'erreur.
        empty = f'OR(${cid}{r}="",${res}{r}="",${wgt}{r}="")'
        ready = readiness_test.format(r=r)
        ws.cell(row=r, column=S.GO_NOGO.index("Poids_Effectif") + 1,
                value=f"=IF({empty},0,${wgt}{r})").font = F_BODY
        ws.cell(row=r, column=S.GO_NOGO.index("Points") + 1,
                value=f"=IF({empty},0,${wgt}{r}*${sco}{r})").font = F_BODY
        ws.cell(row=r, column=S.GO_NOGO.index("Poids_Readiness") + 1,
                value=f"=IF({empty},0,IF({ready},${wgt}{r},0))").font = F_BODY
        ws.cell(row=r, column=S.GO_NOGO.index("Points_Readiness") + 1,
                value=f"=IF({empty},0,IF({ready},${wgt}{r}*${sco}{r},0))").font = F_BODY

    def rng(key):
        letter, n = ranges[key]
        return "=LISTS!${l}$2:${l}${n}".format(l=letter, n=n + 1)

    for field, key in [("Category", "Category"), ("Result", "Result"),
                       ("Risk", "Risk"), ("Mandatory", "Yes_No"),
                       ("Eliminatory", "Yes_No")]:
        dv = DataValidation(type="list", formula1=rng(key), allow_blank=True)
        dv.errorTitle = "Valeur non autorisee"
        dv.error = "Utiliser la liste deroulante."
        ws.add_data_validation(dv)
        letter = col(field)
        dv.add(f"{letter}2:{letter}{MAX_ROWS + 1}")

    dv = DataValidation(type="whole", operator="between", formula1=1,
                        formula2=5, allow_blank=True)
    dv.errorTitle = "Poids invalide"
    dv.error = "Le poids va de 1 (accessoire) a 5 (determinant)."
    ws.add_data_validation(dv)
    dv.add(f"{wgt}2:{wgt}{MAX_ROWS + 1}")

    # Un critere eliminatoire non satisfait doit se voir de loin.
    elim = col("Eliminatory")
    full = f"A2:{col('Owner')}{MAX_ROWS + 1}"
    for formula, fill in [
        (f'=AND(${elim}2="OUI",${res}2="Non satisfait")', FILL_RED),
        (f'=AND(${elim}2="OUI",OR(${res}2="A verifier",'
         f'${res}2="Partiellement satisfait"))', FILL_ORANGE),
        (f'=${res}2="Non satisfait"', FILL_YELLOW),
        (f'=${res}2="Satisfait"', FILL_GREEN),
    ]:
        ws.conditional_formatting.add(
            full, FormulaRule(formula=[formula], fill=fill, stopIfTrue=True))

    for name in TECHNICAL_COLS:
        ws.column_dimensions[col(name)].hidden = True

    ws.auto_filter.ref = f"A1:{col('Owner')}{MAX_ROWS + 1}"
    return ws


# --------------------------------------------------------------- DECISION --
def sheet_decision(wb):
    ws = wb.create_sheet("DECISION")
    for letter, width in (("A", 44), ("B", 22), ("C", 76)):
        ws.column_dimensions[letter].width = width

    last = MAX_ROWS + 1
    cid = abs_range(CRIT, col("Criterion_ID"), last)
    man = abs_range(CRIT, col("Mandatory"), last)
    eli = abs_range(CRIT, col("Eliminatory"), last)
    res = abs_range(CRIT, col("Result"), last)
    rsk = abs_range(CRIT, col("Risk"), last)
    act = abs_range(CRIT, col("Action"), last)
    pe = abs_range(CRIT, col("Poids_Effectif"), last)
    pts = abs_range(CRIT, col("Points"), last)
    pr = abs_range(CRIT, col("Poids_Readiness"), last)
    ptr = abs_range(CRIT, col("Points_Readiness"), last)

    r = 1
    ws.cell(row=r, column=1, value="Decision Go / No-Go").font = F_TITLE
    r += 1
    ws.cell(row=r, column=1,
            value="Resume executif calcule depuis l'onglet CRITERIA. "
                  "Seules les cellules encadrees se saisissent.").font = F_MUTED
    r += 2

    ws.cell(row=r, column=1, value="OPPORTUNITE").font = F_H2
    r += 1
    for label, note in [
        ("Opportunity_ID", "Reprendre l'identifiant du Command Center (module 02)."),
        ("Titre", None),
        ("Organisation", None),
        ("Pays", None),
        ("Deadline", "Format aaaa-mm-jj."),
    ]:
        ws.cell(row=r, column=1, value=label).font = F_LABEL
        c = ws.cell(row=r, column=2)
        c.font = F_BODY
        c.border = BORDER
        if note:
            ws.cell(row=r, column=3, value=note).font = F_MUTED
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="PARAMETRE").font = F_H2
    r += 1
    ws.cell(row=r, column=1, value="Seuil de GO (score d'eligibilite)").font = F_LABEL
    thr = ws.cell(row=r, column=2, value=DEFAULT_GO_THRESHOLD)
    thr.font = F_BODY
    thr.border = BORDER
    ws.cell(row=r, column=3,
            value="En dessous, le verdict est NO_GO. Ce seuil ne peut jamais "
                  "annuler un critere eliminatoire.").font = F_MUTED
    threshold_row = r
    r += 2

    ws.cell(row=r, column=1, value="LES TROIS SCORES RESTENT SEPARES").font = F_H2
    r += 1
    warn = ws.cell(row=r, column=1,
                   value="Ne jamais les additionner ni en faire une moyenne.")
    warn.font = Font(name="Calibri", size=10, bold=True, color="FF9B1C1C")
    ws.cell(row=r, column=3,
            value="Une moyenne masquerait un blocage reel derriere deux bons "
                  "scores.").font = F_MUTED
    r += 1

    ws.cell(row=r, column=1, value="Relevance Score").font = F_LABEL
    rel = ws.cell(row=r, column=2)
    rel.font = F_BODY
    rel.border = BORDER
    ws.cell(row=r, column=3,
            value="SAISIE. Vient du Command Center (module 02) ou du moteur de "
                  "scoring (module O). Mesure l'interet, pas "
                  "l'eligibilite.").font = F_MUTED
    r += 1

    ws.cell(row=r, column=1, value="Eligibility Score").font = F_LABEL
    elig = ws.cell(row=r, column=2,
                   value=f'=IF(SUM({pe})=0,"n/a",SUM({pts})/SUM({pe}))')
    elig.font = F_KPI
    elig.number_format = "0.0"
    elig_row = r
    ws.cell(row=r, column=3,
            value="CALCULE. Moyenne ponderee de tous les criteres "
                  "evalues.").font = F_MUTED
    r += 1

    ws.cell(row=r, column=1, value="Readiness Score").font = F_LABEL
    rd = ws.cell(row=r, column=2,
                 value=f'=IF(SUM({pr})=0,"n/a",SUM({ptr})/SUM({pr}))')
    rd.font = F_KPI
    rd.number_format = "0.0"
    ws.cell(row=r, column=3,
            value="CALCULE sur les categories "
                  + ", ".join(S.READINESS_CATEGORIES)
                  + " : sommes-nous prets MAINTENANT ?").font = F_MUTED
    r += 2

    ws.cell(row=r, column=1, value="BLOCAGES").font = F_H2
    r += 1
    counters = {}
    for label, formula, note in [
        ("Criteres evalues", f'=COUNTIF({cid},"<>")', None),
        ("Criteres sans resultat",
         f'=SUMPRODUCT(({cid}<>"")*({res}=""))',
         "Une evaluation incomplete ne peut pas produire un GO."),
        ("Eliminatoires non satisfaits",
         f'=COUNTIFS({eli},"OUI",{res},"Non satisfait")',
         "Un seul suffit a imposer NO_GO."),
        ("Eliminatoires non tranches",
         f'=COUNTIFS({eli},"OUI",{res},"A verifier")'
         f'+COUNTIFS({eli},"OUI",{res},"Partiellement satisfait")',
         "Imposent NO_GO_CONDITIONAL tant qu'ils ne sont pas clarifies."),
        ("Obligatoires non satisfaits",
         f'=COUNTIFS({man},"OUI",{eli},"NON",{res},"Non satisfait")'
         f'+COUNTIFS({man},"OUI",{eli},"NON",{res},"Partiellement satisfait")'
         f'+COUNTIFS({man},"OUI",{eli},"NON",{res},"A verifier")',
         "Corrigibles avant depot : conduisent a GO_WITH_ACTIONS."),
        ("Criteres a risque eleve", f'=COUNTIF({rsk},"Eleve")', None),
        ("Actions a mener avant decision", f'=COUNTIF({act},"<>")',
         "Filtrer CRITERIA sur la colonne Action pour obtenir la liste."),
    ]:
        ws.cell(row=r, column=1, value=label).font = F_LABEL
        ws.cell(row=r, column=2, value=formula).font = F_BODY
        if note:
            ws.cell(row=r, column=3, value=note).font = F_MUTED
        counters[label] = r
        r += 1
    r += 1

    n_crit = counters["Criteres evalues"]
    n_none = counters["Criteres sans resultat"]
    n_ko = counters["Eliminatoires non satisfaits"]
    n_unclear = counters["Eliminatoires non tranches"]
    n_gaps = counters["Obligatoires non satisfaits"]

    # La cascade reproduit verdict_for(). L'ordre EST la regle produit : les
    # criteres eliminatoires sont evalues AVANT le score.
    verdict = (
        f'=IF(B{n_crit}=0,"Aucun critere saisi",'
        f'IF(B{n_ko}>0,"NO_GO",'
        f'IF(B{n_unclear}>0,"NO_GO_CONDITIONAL",'
        f'IF(B{n_none}>0,"NO_GO_CONDITIONAL",'
        f'IF(B{elig_row}<B{threshold_row},"NO_GO",'
        f'IF(B{n_gaps}>0,"GO_WITH_ACTIONS","GO"))))))'
    )
    motif = (
        f'=IF(B{n_crit}=0,"Aucun critere n a ete renseigne.",'
        f'IF(B{n_ko}>0,"Critere eliminatoire non satisfait. '
        f'Aucun score ne peut compenser.",'
        f'IF(B{n_unclear}>0,"Critere eliminatoire non tranche. '
        f'A clarifier avant toute decision.",'
        f'IF(B{n_none}>0,"Evaluation incomplete : des criteres sont '
        f'sans resultat.",'
        f'IF(B{elig_row}<B{threshold_row},"Score d eligibilite insuffisant.",'
        f'IF(B{n_gaps}>0,"Exigences obligatoires a corriger avant depot.",'
        f'"Tous les criteres obligatoires sont satisfaits."))))))'
    )

    ws.cell(row=r, column=1, value="VERDICT").font = F_H2
    r += 1
    v = ws.cell(row=r, column=2, value=verdict)
    v.font = Font(name="Calibri", size=22, bold=True, color="FF1F3A5F")
    v.alignment = Alignment(horizontal="center", vertical="center")
    v.border = BORDER
    ws.row_dimensions[r].height = 42
    verdict_row = r
    m = ws.cell(row=r, column=3, value=motif)
    m.font = F_BODY
    m.alignment = Alignment(wrap_text=True, vertical="center")
    r += 2

    cell = f"B{verdict_row}"
    for formula, fill in [
        (f'={cell}="NO_GO"', FILL_RED),
        (f'={cell}="NO_GO_CONDITIONAL"', FILL_ORANGE),
        (f'={cell}="GO_WITH_ACTIONS"', FILL_YELLOW),
        (f'={cell}="GO"', FILL_GREEN),
        (f'={cell}="Aucun critere saisi"', FILL_GREY),
    ]:
        ws.conditional_formatting.add(
            f"B{verdict_row}:C{verdict_row}",
            FormulaRule(formula=[formula], fill=fill, stopIfTrue=True))

    ws.cell(row=r, column=1, value="LECTURE DU VERDICT").font = F_H2
    r += 1
    for verdict_name, meaning in [
        ("GO", "Soumissionner. Aucun blocage identifie."),
        ("GO_WITH_ACTIONS",
         "Soumissionner, mais des exigences obligatoires doivent etre "
         "corrigees avant le depot."),
        ("NO_GO_CONDITIONAL",
         "Ne pas decider tout de suite : un point eliminatoire n'est pas "
         "tranche, ou l'evaluation est incomplete. Clarifier puis re-evaluer."),
        ("NO_GO",
         "Ne pas soumissionner. Soit un critere eliminatoire est non "
         "satisfait, soit le score est sous le seuil."),
    ]:
        ws.cell(row=r, column=1, value=verdict_name).font = F_LABEL
        c = ws.cell(row=r, column=2, value=meaning)
        c.font = F_BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 1
    return ws


# ------------------------------------------------------------ START_HERE --
def sheet_start_here(wb, today, n_demo):
    ws = wb.active
    ws.title = "START_HERE"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 108
    ws.sheet_view.showGridLines = False

    r = 2
    ws.cell(row=r, column=2, value="TENDERPILOT - GO / NO-GO ENGINE").font = F_TITLE
    r += 1
    ws.cell(row=r, column=2,
            value=f"Module 04 - version {MODULE_VERSION} - schema "
                  f"{S.SCHEMA_VERSION} - genere le {today.isoformat()}").font = F_MUTED
    r += 2

    def block(title, lines):
        nonlocal r
        ws.cell(row=r, column=2, value=title).font = F_H2
        r += 1
        for line in lines:
            c = ws.cell(row=r, column=2, value=line)
            c.font = F_BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1

    block("A QUOI SERT CE CLASSEUR", [
        "Decider vite, et pour de bonnes raisons, s'il faut repondre a une "
        "opportunite.",
        "Un classeur = une opportunite. Dupliquer le fichier pour chaque DAO.",
    ])
    block("LA REGLE QUI COMPTE", [
        "Un critere ELIMINATOIRE non satisfait ne peut JAMAIS etre compense "
        "par un score.",
        "Le verdict regarde d'abord les criteres eliminatoires, ensuite "
        "seulement le score.",
        "C'est l'erreur classique : voir 80/100 et se lancer, alors qu'une "
        "certification obligatoire manque et que l'offre sera ecartee sans "
        "meme etre lue.",
    ])
    block("UTILISATION EN 6 ETAPES", [
        "1. Renseigner l'opportunite en haut de l'onglet DECISION.",
        "2. Saisir le Relevance Score : il vient du Command Center (module 02).",
        "3. Remplir l'onglet CRITERIA, une ligne par exigence du DAO.",
        "4. Pour chaque ligne : Obligatoire ? Eliminatoire ? Poids de 1 a 5 ?",
        "5. Renseigner la preuve de votre cote, puis le resultat.",
        "6. Lire le verdict sur l'onglet DECISION.",
    ])
    block("REMPLIR LA COLONNE RESULT", [
        "Satisfait : la preuve existe et est disponible aujourd'hui.",
        "Partiellement satisfait : partiellement couvert, ou couvert sous "
        "condition.",
        "Non satisfait : nous ne remplissons pas l'exigence.",
        "A verifier : nous ne savons pas encore. Ce n'est pas un Non, mais "
        "cela bloque le GO tant que ce n'est pas tranche.",
        "La colonne Score se calcule seule : Satisfait 100, Partiellement 50, "
        "les deux autres 0. Elle n'est volontairement pas saisissable.",
    ])
    block("LES TROIS SCORES", [
        "Relevance : est-ce interessant pour nous ? (saisi, vient du module 02)",
        "Eligibility : sommes-nous eligibles ? (calcule ici)",
        "Readiness : sommes-nous prets MAINTENANT ? (calcule sur les "
        "categories " + ", ".join(S.READINESS_CATEGORIES) + ")",
        "Ils sont affiches separement et ne doivent jamais etre moyennes.",
    ])
    block("LE JEU DE DEMONSTRATION", [
        f"L'onglet CRITERIA contient {n_demo} criteres DEMO entierement "
        "fictifs.",
        "Ils produisent volontairement le cas le plus instructif : un score "
        "d'eligibilite AU-DESSUS du seuil de GO, et pourtant un verdict "
        "NO_GO, parce qu'une certification eliminatoire manque (CRIT-003).",
        "Pour voir le mecanisme : passer CRIT-003 a Satisfait et observer le "
        "verdict basculer en GO_WITH_ACTIONS.",
        "Supprimer ces lignes avant tout usage reel.",
    ])
    return ws


# ------------------------------------------------------------------ docs --
README_TEMPLATE = """# Module 04 - TenderPilot Go / No-Go Engine

Version {version} - schema {schema} - genere le {date}

## Contenu

`04_TenderPilot_Go_NoGo.xlsx` - 4 onglets :
START_HERE, DECISION, CRITERIA, LISTS (technique, masque).

Un classeur = une opportunite. Dupliquer le fichier pour chaque DAO.

## La regle du module

> Un critere eliminatoire non satisfait ne peut jamais etre compense par un
> score.

Le verdict n'est pas un seuil sur une moyenne, c'est une cascade :

1. Aucun critere saisi -> pas de verdict.
2. Un eliminatoire **non satisfait** -> `NO_GO`. Definitif.
3. Un eliminatoire **non tranche** (A verifier / Partiellement) ->
   `NO_GO_CONDITIONAL`.
4. Des criteres **sans resultat** -> `NO_GO_CONDITIONAL` (evaluation
   incomplete).
5. Score d'eligibilite **sous le seuil** -> `NO_GO`.
6. Des obligatoires non satisfaits -> `GO_WITH_ACTIONS`.
7. Sinon -> `GO`.

Les etapes 2 et 3 passent AVANT toute lecture du score. C'est cette inversion
qui fait la valeur du module.

## Les trois scores

| Score | Origine | Mesure |
|-------|---------|--------|
| Relevance | saisi, vient du module 02 | est-ce interessant pour nous ? |
| Eligibility | calcule ici | sommes-nous eligibles ? |
| Readiness | calcule sur {readiness} | sommes-nous prets maintenant ? |

Ils ne sont jamais additionnes ni moyennes. Un chiffre unique masquerait un
blocage derriere deux bons resultats.

## La colonne Score n'est pas saisissable

Elle se derive du Result : Satisfait 100, Partiellement satisfait 50,
Non satisfait 0, A verifier 0. Un score saisi librement serait la porte
ouverte a l'arrangement du resultat par celui qui veut soumissionner.

## Jeu de demonstration

{n_demo} criteres fictifs, prefixes DEMO dans la colonne Requirement.

Ils produisent le cas le plus instructif du produit :

- score d'eligibilite calcule : **{demo_score}/100**
- seuil de GO : **{threshold}**
- le score PASSE le seuil
- verdict : **{demo_verdict}**

parce que CRIT-003 (certification eliminatoire) est non satisfait.

Pour la demonstration : passer CRIT-003 a `Satisfait`, le verdict bascule en
`GO_WITH_ACTIONS` (il reste deux exigences obligatoires a corriger, CRIT-006
et CRIT-009).

Supprimer ces lignes avant tout usage reel.

## Tests

    python tests/test_04_gonogo.py

Les tests comparent la cascade Python (`verdict_for()`) et la cascade en
formules du classeur sur une serie de scenarios : eliminatoire echoue avec
score parfait, eliminatoire non tranche, evaluation incomplete, score sous le
seuil, ecarts obligatoires seuls, et cas entierement satisfait.

## Limites connues

- La liste des ecarts et des actions n'a pas d'onglet dedie : filtrer CRITERIA
  sur Result ou sur Action.
- Le tableau couvre {max_rows} criteres.
- Le Relevance Score n'est pas recalcule ici : il est saisi et provient du
  Command Center ou du moteur de scoring.
- Les criteres doivent etre extraits du DAO. Le module 08 (Compliance Matrix)
  et l'analyseur IA (module P) automatisent partiellement cette extraction ;
  ici la saisie est manuelle.

## Changelog

### {version} - {date}
- Version initiale.
- Matrice de {n_cols} colonnes, dont 4 techniques masquees.
- Cascade de decision a 4 verdicts, criteres eliminatoires prioritaires.
- Trois scores affiches separement, Score par critere non saisissable.
- Jeu DEMO de {n_demo} criteres illustrant le cas score suffisant / NO_GO.
"""


def write_docs(today, demo):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    verdict, _motif, score = verdict_for(demo)
    path = OUT_DIR / "README.md"
    path.write_text(
        README_TEMPLATE.format(
            version=MODULE_VERSION, schema=S.SCHEMA_VERSION,
            date=today.isoformat(), n_demo=len(demo), demo_score=score,
            demo_verdict=verdict, threshold=DEFAULT_GO_THRESHOLD,
            max_rows=MAX_ROWS, n_cols=len(S.GO_NOGO),
            readiness=", ".join(S.READINESS_CATEGORIES),
        ),
        encoding="utf-8",
    )
    return path


# ----------------------------------------------------------------- build ---
def build():
    today = dt.date.today()
    demo = read_demo()

    wb = Workbook()
    sheet_start_here(wb, today, len(demo))
    ranges = sheet_lists(wb)
    sheet_criteria(wb, ranges, demo)
    sheet_decision(wb)

    wb._sheets = [wb[name] for name in ["START_HERE", "DECISION", CRIT, "LISTS"]]
    wb.active = 0

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)
    write_docs(today, demo)
    return OUT_FILE


if __name__ == "__main__":
    print(f"OK  {build()}")
