"""
Tests du module 03 - Organization Profile.

Comme pour le module 02, les tests relisent les formules reellement ecrites
dans le .xlsx. Ils evaluent en plus la logique Candidate_Type pour les 5 types
de candidats, en interpretant la formule "Requis ?" telle qu'elle est ecrite.

    python tests/test_03_org_profile.py
"""

import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tests"))

from builders.build_03_org_profile import (  # noqa: E402
    CANDIDATE_TYPES, DOCUMENT_TYPES, MAX_DOC_ROWS, NOT_PROVIDED, OUT_FILE,
    SECTION_TITLES, read_fields,
)

FAILURES = []
CHECKS = 0


def check(label, condition, detail=""):
    global CHECKS
    CHECKS += 1
    if condition:
        print(f"  ok   {label}")
    else:
        print(f"  FAIL {label}  {detail}")
        FAILURES.append(label)


def requis_for(applies_to, required, candidate_type):
    """Reproduit la formule 'Requis ?' du classeur.

    =IF(ct="","Type non defini",
       IF(OR(applies="ALL",ISNUMBER(SEARCH(ct,applies))),
          IF(required="OUI","Requis","Optionnel"),
          "Sans objet"))
    """
    if not candidate_type:
        return "Type non defini"
    if applies_to == "ALL" or candidate_type in applies_to:
        return "Requis" if required == "OUI" else "Optionnel"
    return "Sans objet"


def main():
    print(f"\nFichier teste : {OUT_FILE}")
    if not OUT_FILE.exists():
        print("FAIL : le fichier n'existe pas. Lancer d'abord le builder.")
        return 1

    wb = load_workbook(OUT_FILE, data_only=False)
    fields = read_fields()

    # ---------------------------------------------------------- structure --
    print("\n[1] Structure")
    expected = (["START_HERE"] + list(SECTION_TITLES)
                + ["DOCUMENTS", "PROFILE_SUMMARY", "COMPLETENESS"])
    check("les 9 onglets sont presents et dans l'ordre",
          wb.sheetnames == expected, f"trouve {wb.sheetnames}")

    ident = wb["IDENTITY"]
    check("le Type de candidat est bien en IDENTITY!B2",
          ident["A2"].value == "Type de candidat")
    dv_types = [dv for dv in ident.data_validations.dataValidation
                if dv.type == "list" and "B2" in str(dv.sqref)]
    check("le Type de candidat a une liste fermee", len(dv_types) == 1)
    if dv_types:
        listed = dv_types[0].formula1.strip('"').split(",")
        check("les 5 types de candidat sont proposes",
              listed == CANDIDATE_TYPES, str(listed))

    # ----------------------------------------------- registre -> classeur --
    print("\n[2] Le registre de champs est integralement transcrit")
    summary = wb["PROFILE_SUMMARY"]
    # La premiere ligne de donnees est localisee, jamais codee en dur : la
    # mise en page de l'entete peut bouger sans invalider le test.
    head_row = next(r for r in range(1, 12)
                    if summary.cell(row=r, column=1).value == "Field_Key")
    sum_first = head_row + 1
    keys_in_file = [summary.cell(row=r, column=1).value
                    for r in range(sum_first, sum_first + len(fields))]
    check(f"les {len(fields)} champs du CSV sont dans PROFILE_SUMMARY",
          keys_in_file == [f["Field_Key"] for f in fields],
          f"{len(keys_in_file)} lignes trouvees")

    # Chaque ligne du resume doit pointer sur la bonne cellule de saisie.
    mismatches = []
    for i, f in enumerate(fields):
        row = sum_first + i
        value_formula = summary.cell(row=row, column=4).value or ""
        m = re.search(r"([A-Z_]+)!\$B\$(\d+)", value_formula)
        if not m:
            mismatches.append((f["Field_Key"], "formule illisible"))
            continue
        sheet, srow = m.group(1), int(m.group(2))
        if sheet != f["Section"]:
            mismatches.append((f["Field_Key"], f"pointe sur {sheet}"))
            continue
        label_in_form = wb[sheet].cell(row=srow, column=1).value
        if label_in_form != f["Label"]:
            mismatches.append((f["Field_Key"], f"libelle {label_in_form!r}"))
    check("chaque ligne du resume pointe sur la bonne cellule de saisie",
          not mismatches, str(mismatches[:5]))

    check("la valeur absente affiche 'Non fourni' et non une valeur inventee",
          NOT_PROVIDED in (summary.cell(row=sum_first, column=4).value or ""))

    # ------------------------------------------- logique Candidate_Type ---
    print("\n[3] Logique Candidate_Type (formule relue et evaluee)")
    ident_fields = [f for f in fields if f["Section"] == "IDENTITY"]
    sample_row = 5  # premiere ligne de saisie de IDENTITY
    formula = ident.cell(row=sample_row, column=3).value or ""
    check("la formule Requis ? reference bien IDENTITY!$B$2",
          "IDENTITY!$B$2" in formula, formula)
    check("la formule gere le cas ou le type n'est pas choisi",
          '"Type non defini"' in formula)
    check("la formule distingue les 3 etats",
          all(s in formula for s in ('"Requis"', '"Optionnel"', '"Sans objet"')))

    # Cas metier concrets, evalues pour chaque type de candidat.
    by_key = {f["Field_Key"]: f for f in fields}
    cases = [
        ("rccm", "PME", "Requis"),
        ("rccm", "CONSULTANT", "Sans objet"),
        ("ngo_registration", "ONG", "Requis"),
        ("ngo_registration", "PME", "Sans objet"),
        ("consortium_agreement", "CONSORTIUM", "Requis"),
        ("consortium_agreement", "CABINET", "Sans objet"),
        ("years_experience", "CONSULTANT", "Requis"),
        ("years_experience", "ONG", "Sans objet"),
        ("meal_system", "ONG", "Requis"),
        ("employees_count", "CONSULTANT", "Sans objet"),
        ("org_name", "CONSULTANT", "Requis"),
        ("website", "PME", "Optionnel"),
    ]
    for key, ctype, want in cases:
        f = by_key[key]
        got = requis_for(f["Applies_To"], f["Required"], ctype)
        check(f"{key} pour un {ctype} -> {want}", got == want, f"obtenu {got}")

    # Un consultant individuel ne doit pas se voir imposer des champs societe.
    company_only = ["rccm", "employees_count", "turnover_n1", "legal_status"]
    consultant_reqs = [k for k in company_only
                       if requis_for(by_key[k]["Applies_To"],
                                     by_key[k]["Required"], "CONSULTANT") == "Requis"]
    check("un consultant individuel n'a aucun champ societe requis",
          not consultant_reqs, str(consultant_reqs))

    # Chaque type de candidat doit avoir au moins un champ requis, sinon le
    # score de completude diviserait par zero.
    for ctype in CANDIDATE_TYPES:
        n = sum(1 for f in fields
                if requis_for(f["Applies_To"], f["Required"], ctype) == "Requis")
        check(f"{ctype} a des champs requis ({n})", n > 0)

    # ------------------------------------------------------- documents ----
    print("\n[4] Documents et expiration")
    docs = wb["DOCUMENTS"]
    first = 4
    names = [docs.cell(row=first + i, column=1).value
             for i in range(len(DOCUMENT_TYPES))]
    check(f"les {len(DOCUMENT_TYPES)} types de documents sont pre-listes",
          names == [d[0] for d in DOCUMENT_TYPES])
    check("aucune date n'est pre-remplie (pas de donnee inventee)",
          all(docs.cell(row=first + i, column=c).value is None
              for i in range(len(DOCUMENT_TYPES)) for c in (4, 5)))
    check("la disponibilite par defaut est 'Non fourni'",
          docs.cell(row=first, column=3).value == NOT_PROVIDED)

    status = docs.cell(row=first, column=7).value or ""
    for state in ("Manquant", "Expire", "Expire sous 15 jours",
                  "Expire sous 30 jours", "Valide", "Sans expiration"):
        check(f"le statut '{state}' est calcule", f'"{state}"' in status)
    check("un document non disponible est Manquant avant tout calcul de date",
          status.index('"Manquant"') < status.index('"Expire"'))
    last_doc = first + MAX_DOC_ROWS - 1
    check(f"la formule de statut couvre la ligne {last_doc}",
          bool(docs.cell(row=last_doc, column=7).value))

    # ---------------------------------------------------- completude ------
    print("\n[5] Completude distincte de l'eligibilite")
    comp = wb["COMPLETENESS"]
    texts = [str(c.value) for row in comp.iter_rows() for c in row if c.value]
    check("l'onglet avertit que ce n'est PAS une eligibilite",
          any("PAS" in t and "eligibilite" in t for t in texts))
    check("il renvoie vers le module 04 pour l'eligibilite",
          any("module 04" in t for t in texts))

    score = next((str(comp.cell(row=r, column=2).value)
                  for r in range(1, 40)
                  if comp.cell(row=r, column=1).value == "SCORE DE COMPLETUDE"), "")
    check("le score de completude existe", bool(score))
    check("le score gere la division par zero",
          "Choisir le type de candidat" in score, score)

    only_required = next((str(comp.cell(row=r, column=2).value)
                          for r in range(1, 40)
                          if comp.cell(row=r, column=1).value
                          == SECTION_TITLES["IDENTITY"]), "")
    check("le compte par section ne retient que les champs Requis",
          '"Requis"' in only_required, only_required)

    # --------------------------------------------------------- securite ---
    print("\n[6] Donnees")
    all_texts = [str(c.value) for sheet in wb.worksheets
                 for row in sheet.iter_rows() for c in row if c.value]
    urls = [t for t in all_texts if t.startswith("http")]
    check("aucune URL reelle pre-remplie", not urls, str(urls[:3]))

    filled_values = [ident.cell(row=r, column=2).value
                     for r in range(5, 5 + len(ident_fields))]
    check("aucune valeur de profil pre-remplie",
          all(v is None for v in filled_values),
          str([v for v in filled_values if v is not None]))

    # ---------------------------------------------------------- resultat --

    # -------------------------------------------- validite des formules ---
    print("\n""[7] Validite syntaxique des formules")
    from formula_check import invalid_references, formula_count
    cassees = invalid_references(OUT_FILE)
    total = formula_count(OUT_FILE)
    check(f"les {total} formules du classeur ont des references valides",
          not cassees,
          str(cassees[:3]) + (f" (+{len(cassees) - 3})" if len(cassees) > 3 else ""))
    check("le classeur contient bien des formules", total > 0)

    print(f"\n{'-' * 62}")
    if FAILURES:
        print(f"ECHEC : {len(FAILURES)}/{CHECKS} verifications en echec")
        for f in FAILURES:
            print(f"  - {f}")
        return 1
    print(f"SUCCES : {CHECKS}/{CHECKS} verifications passees")
    return 0


if __name__ == "__main__":
    sys.exit(main())
