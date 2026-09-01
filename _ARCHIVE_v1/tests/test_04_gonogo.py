"""
Tests du module 04 - Go / No-Go Engine.

Le classeur contient une cascade de decision ecrite en formules Excel. Le
builder contient la meme cascade ecrite en Python. Si les deux divergent, le
produit ment a son utilisateur.

Ces tests parsent la formule VERDICT reellement ecrite dans le .xlsx, la
transforment en arbre de decision, puis comparent son resultat a celui de
verdict_for() sur une serie de scenarios metier.

    python tests/test_04_gonogo.py
"""

import copy
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tests"))

from schema import columns as S  # noqa: E402
from builders.build_04_gonogo import (  # noqa: E402
    DEFAULT_GO_THRESHOLD, MAX_ROWS, OUT_FILE, TECHNICAL_COLS, col, read_demo,
    verdict_for,
)

FAILURES = []
CHECKS = 0


def check(label, condition, detail=""):
    global CHECKS
    CHECKS += 1
    if condition:
        print(f"  ok   {label}")
    else:
        print(f"  FAIL {label}  {detail}")
        FAILURES.append(label)


# --------------------------------------------------------------------------
# Parseur de formule : =IF(cond,alors,IF(cond,alors,...)) devient une liste
# ordonnee de (condition, resultat) plus un resultat par defaut.
# --------------------------------------------------------------------------
def split_top_level(text):
    parts, depth, quoted, current = [], 0, False, ""
    for ch in text:
        if ch == '"':
            quoted = not quoted
        if not quoted:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif ch == "," and depth == 0:
                parts.append(current)
                current = ""
                continue
        current += ch
    parts.append(current)
    return parts


def parse_cascade(formula):
    expr = str(formula).lstrip("=").strip()
    branches = []
    while expr.startswith("IF(") and expr.endswith(")"):
        args = split_top_level(expr[3:-1])
        if len(args) != 3:
            raise AssertionError(f"IF a {len(args)} arguments : {expr[:80]}")
        cond, then, other = (a.strip() for a in args)
        branches.append((cond, then.strip('"')))
        expr = other
    return branches, expr.strip('"')


COND_NUM = re.compile(r"^B(\d+)(=|>=|<=|>|<)(-?\d+)$")
COND_CELL = re.compile(r"^B(\d+)(=|>=|<=|>|<)B(\d+)$")
OPS = {"=": lambda a, b: a == b, ">": lambda a, b: a > b,
       "<": lambda a, b: a < b, ">=": lambda a, b: a >= b,
       "<=": lambda a, b: a <= b}


def eval_cond(cond, values):
    m = COND_CELL.match(cond)
    if m:
        return OPS[m.group(2)](values[int(m.group(1))], values[int(m.group(3))])
    m = COND_NUM.match(cond)
    if m:
        return OPS[m.group(2)](values[int(m.group(1))], int(m.group(3)))
    raise AssertionError(f"Condition non interpretable : {cond!r}")


def run_cascade(branches, default, values):
    for cond, result in branches:
        if eval_cond(cond, values):
            return result
    return default


def counters_for(criteria, threshold):
    """Valeurs que les COUNTIFS du classeur produiraient pour ce scenario."""
    assessed = [c for c in criteria if c.get("Criterion_ID")]
    unclear = ("Partiellement satisfait", "A verifier")
    gaps = ("Non satisfait", "Partiellement satisfait", "A verifier")

    weight = sum(int(c["Weight"]) for c in assessed
                 if c.get("Result") and c.get("Weight"))
    points = sum(int(c["Weight"]) * S.RESULT_SCORES[c["Result"]]
                 for c in assessed if c.get("Result") and c.get("Weight"))

    return {
        "Criteres evalues": len(assessed),
        "Criteres sans resultat": sum(1 for c in assessed if not c.get("Result")),
        "Eliminatoires non satisfaits": sum(
            1 for c in assessed
            if c["Eliminatory"] == "OUI" and c.get("Result") == "Non satisfait"),
        "Eliminatoires non tranches": sum(
            1 for c in assessed
            if c["Eliminatory"] == "OUI" and c.get("Result") in unclear),
        "Obligatoires non satisfaits": sum(
            1 for c in assessed
            if c["Mandatory"] == "OUI" and c["Eliminatory"] == "NON"
            and c.get("Result") in gaps),
        "Eligibility Score": points / weight if weight else 0,
        "Seuil de GO (score d'eligibilite)": threshold,
    }


def scenarios(demo):
    def with_result(rows, cid, result):
        out = copy.deepcopy(rows)
        for c in out:
            if c["Criterion_ID"] == cid:
                c["Result"] = result
        return out

    all_ok = copy.deepcopy(demo)
    for c in all_ok:
        c["Result"] = "Satisfait"

    sans_resultat = copy.deepcopy(all_ok)
    sans_resultat[4]["Result"] = ""

    # Tout ce qui n'est pas eliminatoire echoue : le score s'effondre sans
    # qu'aucun eliminatoire ne soit en cause.
    score_bas = copy.deepcopy(demo)
    for c in score_bas:
        c["Result"] = "Satisfait" if c["Eliminatory"] == "OUI" else "Non satisfait"

    return [
        ("jeu DEMO livre", demo, "NO_GO"),
        ("DEMO avec CRIT-003 corrige",
         with_result(demo, "CRIT-003", "Satisfait"), "GO_WITH_ACTIONS"),
        ("tous les criteres satisfaits", all_ok, "GO"),
        ("eliminatoire echoue mais tout le reste parfait",
         with_result(all_ok, "CRIT-003", "Non satisfait"), "NO_GO"),
        ("eliminatoire non tranche (A verifier)",
         with_result(all_ok, "CRIT-003", "A verifier"), "NO_GO_CONDITIONAL"),
        ("eliminatoire partiellement satisfait",
         with_result(all_ok, "CRIT-002", "Partiellement satisfait"),
         "NO_GO_CONDITIONAL"),
        ("un critere sans resultat", sans_resultat, "NO_GO_CONDITIONAL"),
        ("score sous le seuil, aucun eliminatoire en cause", score_bas, "NO_GO"),
        ("aucun critere saisi", [], "Aucun critere saisi"),
    ]


def main():
    print(f"\nFichier teste : {OUT_FILE}")
    if not OUT_FILE.exists():
        print("FAIL : le fichier n'existe pas. Lancer d'abord le builder.")
        return 1

    wb = load_workbook(OUT_FILE, data_only=False)
    demo = read_demo()

    # ---------------------------------------------------------- structure --
    print("\n[1] Structure")
    check("les 4 onglets attendus",
          wb.sheetnames == ["START_HERE", "DECISION", "CRITERIA", "LISTS"],
          str(wb.sheetnames))
    check("l'onglet technique LISTS est masque",
          wb["LISTS"].sheet_state == "hidden")

    crit = wb["CRITERIA"]
    headers = [c.value for c in crit[1][:len(S.GO_NOGO)]]
    check("les colonnes CRITERIA correspondent au schema",
          headers == S.GO_NOGO, f"ecart {set(S.GO_NOGO) ^ set(headers)}")
    check("les 4 colonnes techniques sont masquees",
          all(crit.column_dimensions[col(n)].hidden for n in TECHNICAL_COLS))

    # --------------------------------------------- Score non saisissable ---
    print("\n[2] La colonne Score est derivee, pas saisie")
    idx_score = S.GO_NOGO.index("Score") + 1
    score_formula = str(crit.cell(row=2, column=idx_score).value or "")
    check("Score est une formule", score_formula.startswith("="))
    check("Score depend du Result", f"${col('Result')}2" in score_formula)
    for result, value in (("Satisfait", "100"), ("Partiellement satisfait", "50")):
        check(f"{result} vaut {value}",
              f'"{result}"' in score_formula and value in score_formula)
    check("Score est calcule jusqu'a la derniere ligne",
          bool(crit.cell(row=MAX_ROWS + 1, column=idx_score).value))
    check("aucun score n'est saisi en dur sur les lignes DEMO",
          all(str(crit.cell(row=r, column=idx_score).value).startswith("=")
              for r in range(2, 2 + len(demo))))

    # ------------------------------------------------------------ verdict --
    print("\n[3] Cascade de decision (formule relue depuis le fichier)")
    dec = wb["DECISION"]
    labels = {}
    for r in range(1, 60):
        v = dec.cell(row=r, column=1).value
        if isinstance(v, str):
            labels.setdefault(v, r)

    check("l'onglet DECISION porte un VERDICT", "VERDICT" in labels)
    verdict_formula = dec.cell(row=labels["VERDICT"] + 1, column=2).value
    branches, default = parse_cascade(verdict_formula)
    check("la cascade a 6 conditions et un cas par defaut",
          len(branches) == 6, f"{len(branches)} conditions")

    produced = [b[1] for b in branches] + [default]
    for verdict in S.VERDICTS:
        check(f"le verdict {verdict} est atteignable", verdict in produced)

    elim_pos = next(i for i, (c, _r) in enumerate(branches)
                    if c == f"B{labels['Eliminatoires non satisfaits']}>0")
    score_pos = next(i for i, (c, _r) in enumerate(branches) if "<B" in c)
    check("les criteres eliminatoires sont evalues AVANT le score",
          elim_pos < score_pos,
          f"eliminatoire en {elim_pos}, score en {score_pos}")
    check("un eliminatoire non satisfait donne NO_GO",
          branches[elim_pos][1] == "NO_GO")

    # ------------------------------------ formules du classeur vs Python ---
    print("\n[4] Le classeur et la logique Python decident a l'identique")
    row_of = {label: labels[label] for label in [
        "Criteres evalues", "Criteres sans resultat",
        "Eliminatoires non satisfaits", "Eliminatoires non tranches",
        "Obligatoires non satisfaits", "Eligibility Score",
        "Seuil de GO (score d'eligibilite)"]}

    for name, criteria, expected in scenarios(demo):
        values = {row_of[k]: v
                  for k, v in counters_for(criteria, DEFAULT_GO_THRESHOLD).items()}
        from_sheet = run_cascade(branches, default, values)
        from_python, _motif, _score = verdict_for(criteria)
        check(f"{name} -> {expected}",
              from_sheet == expected and from_python == expected,
              f"classeur={from_sheet} python={from_python}")

    # ------------------------------------------------- le cas qui compte ---
    print("\n[5] Un score eleve ne rachete jamais un eliminatoire")
    parfait = copy.deepcopy(demo)
    for c in parfait:
        c["Result"] = "Satisfait"
    for c in parfait:
        if c["Criterion_ID"] == "CRIT-003":
            c["Result"] = "Non satisfait"
    verdict, motif, score = verdict_for(parfait)
    check("scenario calcule", score is not None)
    check(f"score de {score}/100 et pourtant NO_GO", verdict == "NO_GO",
          f"verdict {verdict}")
    check("le motif nomme le critere bloquant", "CRIT-003" in motif, motif)
    check("le motif dit qu'aucun score ne compense", "compenser" in motif)

    demo_verdict, _m, demo_score = verdict_for(demo)
    check(f"le jeu DEMO passe le seuil ({demo_score} >= {DEFAULT_GO_THRESHOLD})",
          demo_score >= DEFAULT_GO_THRESHOLD)
    check("et reste malgre tout en NO_GO", demo_verdict == "NO_GO")

    # ------------------------------------------------- scores separes -----
    print("\n[6] Les trois scores restent separes")
    for label in ("Relevance Score", "Eligibility Score", "Readiness Score"):
        check(f"{label} est affiche", label in labels)
    check("Relevance est une saisie, pas un calcul",
          dec.cell(row=labels["Relevance Score"], column=2).value is None)

    elig = str(dec.cell(row=labels["Eligibility Score"], column=2).value)
    ready = str(dec.cell(row=labels["Readiness Score"], column=2).value)
    check("Eligibility est une moyenne ponderee",
          col("Points") in elig and col("Poids_Effectif") in elig)
    check("Readiness utilise ses propres colonnes",
          col("Points_Readiness") in ready and col("Poids_Readiness") in ready)
    check("Eligibility et Readiness sont deux formules distinctes", elig != ready)

    all_dec = [str(c.value) for row in dec.iter_rows() for c in row if c.value]
    check("aucun score global ne fusionne les trois",
          not any(re.search(r"score\s+(global|total|unique)", t, re.I)
                  for t in all_dec))
    check("l'onglet avertit de ne pas les moyenner",
          any("moyenne" in t.lower() for t in all_dec))

    ready_formula = str(crit.cell(
        row=2, column=S.GO_NOGO.index("Poids_Readiness") + 1).value)
    for category in S.READINESS_CATEGORIES:
        check(f"Readiness inclut la categorie {category}",
              f'"{category}"' in ready_formula)
    hors = [c for c in S.CRITERION_CATEGORIES
            if c not in S.READINESS_CATEGORIES and f'"{c}"' in ready_formula]
    check("Readiness exclut les autres categories", not hors, str(hors))

    # ---------------------------------------------------------- donnees ---
    print("\n[7] Donnees DEMO")
    check("les criteres DEMO sont identifies comme fictifs",
          all(str(crit.cell(row=r, column=3).value).startswith("DEMO")
              for r in range(2, 2 + len(demo))))
    urls = [str(c.value) for row in crit.iter_rows()
            for c in row if c.value and str(c.value).startswith("http")]
    check("aucune URL reelle dans le jeu DEMO", not urls, str(urls[:3]))


    # -------------------------------------------- validite des formules ---
    print("\n""[8] Validite syntaxique des formules")
    from formula_check import invalid_references, formula_count
    cassees = invalid_references(OUT_FILE)
    total = formula_count(OUT_FILE)
    check(f"les {total} formules du classeur ont des references valides",
          not cassees,
          str(cassees[:3]) + (f" (+{len(cassees) - 3})" if len(cassees) > 3 else ""))
    check("le classeur contient bien des formules", total > 0)

    print(f"\n{'-' * 62}")
    if FAILURES:
        print(f"ECHEC : {len(FAILURES)}/{CHECKS} verifications en echec")
        for f in FAILURES:
            print(f"  - {f}")
        return 1
    print(f"SUCCES : {CHECKS}/{CHECKS} verifications passees")
    return 0


if __name__ == "__main__":
    sys.exit(main())
