"""
Tests du module 12 - coherence entre le script et le classeur.

La logique metier est testee sous Node (tests/test_12_apps_script.js). Ici on
verifie ce que Node ne peut pas voir : que le JavaScript et les classeurs
Python parlent bien des memes onglets, des memes colonnes et des memes
libelles de reglages.

C'est la panne la plus probable de ce module : renommer une colonne d'un
cote et casser l'autre silencieusement.

    python tests/test_12_apps_script.py
"""

import json
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from schema import columns as S  # noqa: E402
from builders.build_02_command_center import OUT_FILE as CC_FILE  # noqa: E402
from builders.build_12_apps_script import (  # noqa: E402
    OUT_DIR, SCHEMA_FILE, SOURCE_DIR, SOURCE_FILES,
)

FAILURES = []
CHECKS = 0

HANDWRITTEN = ["Core.gs", "Rss.gs", "Sheets.gs", "Menu.gs"]


def check(label, condition, detail=""):
    global CHECKS
    CHECKS += 1
    if condition:
        print(f"  ok   {label}")
    else:
        print(f"  FAIL {label}  {detail}")
        FAILURES.append(label)


def read(name):
    return (SOURCE_DIR / name).read_text(encoding="utf-8")


def main():
    print(f"\nSources : {SOURCE_DIR}")

    # ---------------------------------------------------------- livrable --
    print("\n[1] Livrable et autorisations")
    for name in SOURCE_FILES + ["INSTALLATION_CLIENT.md",
                                "DEPLOIEMENT_OPERATEUR.md",
                                "GUIDE_UTILISATEUR.md", "MISE_EN_LIGNE.md",
                                "README.md"]:
        check(f"{name} present dans le livrable", (OUT_DIR / name).exists())

    manifest = json.loads(read("appsscript.json"))
    check("le manifeste est un JSON valide", isinstance(manifest, dict))
    check("le fuseau horaire est declare", bool(manifest.get("timeZone")))
    scopes = manifest.get("oauthScopes", [])
    check("les autorisations sont declarees explicitement", len(scopes) > 0)
    check("l acces au classeur est limite au fichier courant",
          any("spreadsheets.currentonly" in s for s in scopes), str(scopes))
    check("l acces Gmail est en lecture seule",
          all("gmail" not in s or s.endswith("gmail.readonly") for s in scopes),
          str([s for s in scopes if "gmail" in s]))

    print("\n[1b] Bundle clasp")
    clasp_dir = OUT_DIR / "clasp"
    check("le dossier clasp existe", clasp_dir.is_dir())

    ecarts = []
    for name in SOURCE_FILES:
        target = name[:-3] + ".js" if name.endswith(".gs") else name
        copie = clasp_dir / target
        if not copie.exists():
            ecarts.append(f"{target} absent")
        elif copie.read_bytes() != (SOURCE_DIR / name).read_bytes():
            ecarts.append(f"{target} differe de {name}")
    check(f"les {len(SOURCE_FILES)} fichiers sont copies a l identique",
          not ecarts, str(ecarts))

    config_path = clasp_dir / ".clasp.json.example"
    check(".clasp.json.example est fourni", config_path.exists())
    if config_path.exists():
        config = json.loads(config_path.read_text(encoding="utf-8"))
        check("l identifiant de script reste a renseigner",
              config.get("scriptId") == "VOTRE_SCRIPT_ID")
        order = config.get("filePushOrder", [])
        check("Schema est pousse avant le code qui l utilise",
              order and order.index("Schema.js") < order.index("Core.js"),
              str(order))
        check("le manifeste est pousse en premier",
              order and order[0] == "appsscript.json", str(order))

    # ------------------------------------------------- Schema.gs genere ---
    print("\n[2] Schema.gs est genere, pas ecrit a la main")
    schema_js = SCHEMA_FILE.read_text(encoding="utf-8")
    check("le fichier annonce qu il est genere",
          "NE PAS MODIFIER A LA MAIN" in schema_js)
    check("il nomme son generateur", "build_12_apps_script.py" in schema_js)
    check("il porte la version du schema", S.SCHEMA_VERSION in schema_js)

    for sheet in S.SHEETS.values():
        check(f"l onglet {sheet} est expose au script", f'"{sheet}"' in schema_js)
    for status in S.STATUSES_CLOSED:
        check(f"le statut clos {status} est expose", f'"{status}"' in schema_js)

    # ------------------------- le JS ne code aucun nom de colonne en dur ---
    print("\n[3] Aucun nom d onglet ni de colonne ecrit en dur dans le JS")
    sheet_names = set(S.SHEETS.values())
    columns = set(S.OPPORTUNITIES) | set(S.LOGS) | set(S.SOURCES)

    for name in HANDWRITTEN:
        source = read(name)
        hard_sheets = [s for s in sheet_names
                       if f'"{s}"' in source or f"'{s}'" in source]
        check(f"{name} ne nomme aucun onglet en dur", not hard_sheets,
              str(hard_sheets))
        hard_columns = [c for c in columns
                        if f"'{c}'" in source or f'"{c}"' in source]
        check(f"{name} ne nomme aucune colonne en dur", not hard_columns,
              str(hard_columns))

    # ------------------------ les cles utilisees existent dans le schema --
    print("\n[4] Les cles utilisees par le script existent")
    joined = "\n".join(read(n) for n in HANDWRITTEN)

    for pattern, known, label in [
        (r"SCHEMA\.OPP\.(\w+)", set(S.OPP_KEYS), "SCHEMA.OPP"),
        (r"SCHEMA\.LOG\.(\w+)", set(S.LOG_KEYS), "SCHEMA.LOG"),
        (r"SCHEMA\.SHEETS\.(\w+)", set(S.SHEETS), "SCHEMA.SHEETS"),
        (r"SETTINGS_LABELS\[[\"'](\w+)[\"']\]", set(S.SETTINGS_LABELS),
         "les cles de reglage"),
    ]:
        used = set(re.findall(pattern, joined))
        unknown = sorted(used - known)
        check(f"toutes {label} existent ({len(used)} utilisees)",
              not unknown, str(unknown))

    setup_html = read("Setup.html")
    for key in ("org_name", "timezone", "notify_email", "reminders_enabled",
                "reminder_days", "gmail_label"):
        check(f"le formulaire de configuration expose {key}",
              f'id="{key}"' in setup_html)

    # -------------------------------- les colonnes visees existent bien ---
    print("\n[5] Les colonnes visees existent dans le Command Center")
    check("le Command Center a ete genere", CC_FILE.exists())
    if CC_FILE.exists():
        wb = load_workbook(CC_FILE, data_only=False)

        opp_headers = [c.value for c in wb["OPPORTUNITIES"][1]]
        missing = sorted(c for c in S.OPP_KEYS.values() if c not in opp_headers)
        check(f"les {len(S.OPP_KEYS)} colonnes visees par SCHEMA.OPP existent",
              not missing, str(missing))

        log_headers = [c.value for c in wb["LOGS"][1]]
        missing = sorted(c for c in S.LOG_KEYS.values() if c not in log_headers)
        check("les colonnes visees par SCHEMA.LOG existent", not missing,
              str(missing))

        check("tous les onglets attendus par le script existent",
              sheet_names <= set(wb.sheetnames),
              str(sorted(sheet_names - set(wb.sheetnames))))

        # Le script ecrit la liste des deadlines a partir d une ligne fixe :
        # elle doit rester libre dans le classeur genere.
        deadlines = wb["DEADLINES"]
        occupied = [r for r in range(S.DEADLINES_LIST_ROW, deadlines.max_row + 1)
                    if any(c.value is not None for c in deadlines[r])]
        check(f"la ligne {S.DEADLINES_LIST_ROW} et les suivantes sont libres "
              f"pour la liste generee", not occupied, f"occupees : {occupied}")

        settings_labels = [str(c.value) for row in wb["SETTINGS"].iter_rows()
                           for c in row if c.value]
        missing = sorted(k for k, v in S.SETTINGS_LABELS.items()
                         if v not in settings_labels)
        check("tous les reglages attendus existent dans le classeur",
              not missing, str(missing))

        watchlist_labels = [str(c.value) for row in wb["WATCHLIST"].iter_rows()
                            for c in row if c.value]
        missing = sorted(k for k, v in S.WATCHLIST_LABELS.items()
                         if v not in watchlist_labels)
        check("tous les parametres de watchlist existent dans le classeur",
              not missing, str(missing))
        check("le titre de la zone de cibles existe",
              S.WATCHLIST_TARGETS_TITLE in watchlist_labels)

        lists_headers = [c.value for c in wb["LISTS"][1]]
        for name in ("Country", "Sector", "Opportunity_Type", "Currency",
                     "Status"):
            check(f"la liste {name} est lisible par le formulaire",
                  name in lists_headers)

    # ---------------------------------------------------------- securite --
    print("\n[6] Securite et cablage")
    every_source = "\n".join(read(n) for n in SOURCE_FILES)
    suspicious = re.compile(r"(sk-[A-Za-z0-9]{16,}|AIza[A-Za-z0-9_-]{20,}|"
                            r"[0-9]{8,10}:AA[A-Za-z0-9_-]{30,})")
    hits = suspicious.findall(every_source)
    check("aucune cle ou jeton dans le code livre", not hits, str(hits))
    check("le script ne cherche pas dans toute la boite Gmail",
          "GmailApp.search" not in every_source)
    check("il ne lit que le libelle configure",
          "getUserLabelByName" in read("Menu.gs"))

    defined = set(re.findall(r"^function (\w+)\s*\(", joined, re.M))
    handlers = set(re.findall(r"\.addItem\('[^']*',\s*'(\w+)'\)", read("Menu.gs")))
    missing = sorted(handlers - defined)
    check(f"les {len(handlers)} entrees de menu pointent sur des fonctions "
          f"existantes", not missing, str(missing))

    # On ne cible que la fin de chaine, seule a designer une fonction serveur :
    #     .getFormOptions();
    #     .submitOpportunity(form);
    called = set()
    for html in ("Setup.html", "AddOpportunity.html"):
        called |= set(re.findall(r"^\s*\.(\w+)\(\w*\);", read(html), re.M))
    called -= {"withSuccessHandler", "withFailureHandler"}
    missing = sorted(called - defined)
    check(f"les {len(called)} appels des barres laterales pointent sur des "
          f"fonctions existantes", not missing, str(missing))

    trigger = re.search(r"newTrigger\((\w+)\)", read("Menu.gs"))
    check("le declencheur quotidien vise une fonction existante",
          trigger is not None and "dailyCheck" in defined)

    # ------------------------------------------- la logique metier passe --
    print("\n[7] Tests de logique metier (Node)")
    node = subprocess.run(["node", str(ROOT / "tests" / "test_12_apps_script.js")],
                          cwd=ROOT, capture_output=True, text=True)
    lines = [ln for ln in node.stdout.strip().splitlines() if ln.strip()]
    check("la suite Node passe", node.returncode == 0,
          lines[-1] if lines else node.stderr[:200])
    if lines:
        print(f"       {lines[-1]}")

    print(f"\n{'-' * 62}")
    if FAILURES:
        print(f"ECHEC : {len(FAILURES)}/{CHECKS} verifications en echec")
        for f in FAILURES:
            print(f"  - {f}")
        return 1
    print(f"SUCCES : {CHECKS}/{CHECKS} verifications passees")
    return 0


if __name__ == "__main__":
    sys.exit(main())
