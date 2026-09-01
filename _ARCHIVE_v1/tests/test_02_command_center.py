"""
Tests du module 02 - Command Center.

Ces tests ne se contentent pas de verifier que le fichier existe : ils
relisent les formules et les regles de mise en forme REELLEMENT ecrites
dans le .xlsx, puis les evaluent sur le jeu DEMO.

Aucune dependance externe. Lancement :
    python tests/test_02_command_center.py
"""

import csv
import datetime as dt
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tests"))

from schema import columns as S  # noqa: E402
from builders.build_02_command_center import MAX_ROWS, OUT_FILE  # noqa: E402

FAILURES = []
CHECKS = 0

GREY = "ECECEC"
GREEN = "D8F3DC"
YELLOW = "FFF3BF"
ORANGE = "FFE0C2"
RED = "FFD6D6"


def check(label, condition, detail=""):
    global CHECKS
    CHECKS += 1
    if condition:
        print(f"  ok   {label}")
    else:
        print(f"  FAIL {label}  {detail}")
        FAILURES.append(label)


# --------------------------------------------------------------------------
# Mini-evaluateur : interprete les regles de couleur telles qu'ecrites dans
# le fichier. Il ne devine rien - une formule d'une forme inconnue fait
# echouer le test au lieu d'etre ignoree silencieusement.
# --------------------------------------------------------------------------
RE_CLOSED = re.compile(r'^=OR\((?:\$[A-Z]+2="[^"]+",?)+\)$')
RE_STATUS = re.compile(r'\$[A-Z]+2="([^"]+)"')
RE_CMP = re.compile(r'^=AND\(\$([A-Z]+)2<>"",\$([A-Z]+)2(<=|<|>)(-?\d+)\)$')


def eval_rule(formula, row, closed_statuses):
    """Retourne True si la regle s'applique a `row`."""
    if RE_CLOSED.match(formula):
        # La liste des statuts clos est desormais ecrite dans la formule :
        # on la relit plutot que de la supposer.
        return row["status"] in RE_STATUS.findall(formula)

    m = RE_CMP.match(formula)
    if m:
        _dl, _days, op, threshold = m.groups()
        if row["deadline"] is None or row["days"] is None:
            return False
        t = int(threshold)
        return {"<=": row["days"] <= t,
                "<": row["days"] < t,
                ">": row["days"] > t}[op]

    raise AssertionError(f"Forme de formule inconnue, non testable : {formula!r}")


def main():
    print(f"\nFichier teste : {OUT_FILE}")
    if not OUT_FILE.exists():
        print("FAIL : le fichier n'existe pas. Lancer d'abord le builder.")
        return 1

    wb = load_workbook(OUT_FILE, data_only=False)
    today = dt.date.today()

    # ---------------------------------------------------------- structure --
    print("\n[1] Structure du classeur")
    expected = ["START_HERE", "DASHBOARD", "OPPORTUNITIES", "PIPELINE",
                "DEADLINES", "WATCHLIST", "SETTINGS", "SOURCES", "LOGS", "LISTS"]
    check("les 10 onglets sont presents et dans l'ordre",
          wb.sheetnames == expected, f"trouve {wb.sheetnames}")
    check("l'onglet technique LISTS est masque",
          wb["LISTS"].sheet_state == "hidden")

    ws = wb["OPPORTUNITIES"]
    headers = [c.value for c in ws[1][:len(S.OPPORTUNITIES)]]
    check("les 30 colonnes OPPORTUNITIES correspondent au schema",
          headers == S.OPPORTUNITIES,
          f"ecart : {set(S.OPPORTUNITIES) ^ set(h for h in headers if h)}")

    src_headers = [c.value for c in wb["SOURCES"][1][:len(S.SOURCES)]]
    check("les colonnes SOURCES correspondent au schema",
          src_headers == S.SOURCES)

    # ------------------------------------------------------ Days_Remaining --
    print("\n[2] Formule Days_Remaining")
    i_days = S.OPPORTUNITIES.index("Days_Remaining") + 1
    dl = S.col_letter(S.OPPORTUNITIES, "Deadline_Date")
    f_first = ws.cell(row=2, column=i_days).value
    f_last = ws.cell(row=MAX_ROWS + 1, column=i_days).value
    check("presente des la ligne 2",
          f_first == f'=IF(${dl}2="","",${dl}2-TODAY())', str(f_first))
    check(f"presente jusqu'a la ligne {MAX_ROWS + 1} (saisie future couverte)",
          f_last == f'=IF(${dl}{MAX_ROWS + 1}="","",'
                    f"${dl}{MAX_ROWS + 1}-TODAY())", str(f_last))
    check("renvoie du vide et non 0 quand la deadline est absente",
          "=IF(" in str(f_first) and '""' in str(f_first))

    # --------------------------------------------------- listes deroulantes --
    print("\n[3] Listes deroulantes et bornes")
    validated = set()
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list":
            for rng in str(dv.sqref).split():
                validated.add(re.match(r"([A-Z]+)", rng).group(1))
    for field in ("Status", "Priority", "Opportunity_Type", "Sector",
                  "Country", "Region", "Currency", "Language",
                  "Eliminatory_Criterion"):
        check(f"{field} a une liste deroulante",
              S.col_letter(S.OPPORTUNITIES, field) in validated)

    bounded = [dv for dv in ws.data_validations.dataValidation if dv.type == "whole"]
    check("les 3 scores sont bornes 0-100", len(bounded) == 3, f"trouve {len(bounded)}")

    # ------------------------------------------------- couleurs deadlines ---
    print("\n[4] Regles de couleur (relues depuis le fichier)")
    last_col = S.col_letter(S.OPPORTUNITIES, S.OPPORTUNITIES[-1])
    target = f"A2:{last_col}{MAX_ROWS + 1}"
    row_rules = []
    for rng, rules in ws.conditional_formatting._cf_rules.items():
        if str(rng.sqref) == target:
            row_rules = rules
    check("6 regles de couleur sur la plage de lignes", len(row_rules) == 6,
          f"trouve {len(row_rules)}")
    check("toutes les regles sont exclusives (stopIfTrue)",
          bool(row_rules) and all(r.stopIfTrue for r in row_rules))
    if row_rules:
        first = row_rules[0].formula[0]
        check("la 1re regle (priorite max) est celle des dossiers clos",
              RE_CLOSED.match(first) is not None, first)
        check("la 1re regle peint en gris",
              row_rules[0].dxf.fill.bgColor.rgb.endswith(GREY))

    if row_rules:
        listes = RE_STATUS.findall(row_rules[0].formula[0])
        check("la regle des dossiers clos enumere les 6 statuts du schema",
              sorted(listes) == sorted(S.STATUSES_CLOSED),
              f"trouve {listes}")
        check("aucune regle de couleur ne reference une autre feuille",
              all("!" not in r.formula[0] for r in row_rules),
              "Google Sheets abandonne ces regles a l import")

    # Une couleur dont le canal alpha vaut 00 est TRANSPARENTE : la regle
    # existe mais ne peint rien. C'est passe une fois, plus jamais.
    alpha = [r.dxf.fill.bgColor.rgb for r in row_rules]
    check("les 6 couleurs sont opaques (alpha FF)",
          all(str(a).upper().startswith("FF") for a in alpha), str(alpha))
    entete = ws.cell(row=1, column=1).fill.fgColor.rgb
    check("le remplissage des en-tetes est opaque",
          str(entete).upper().startswith("FF"), str(entete))

    # ------------------------------- evaluation des regles sur le jeu DEMO --
    print("\n[5] Evaluation des regles sur les 10 lignes DEMO")
    demo = list(csv.DictReader(
        open(ROOT / "data" / "demo" / "opportunities_demo.csv", encoding="utf-8")))
    fills = [r.dxf.fill.bgColor.rgb[-6:] for r in row_rules]

    expected_colors = {
        "DEMO-001": (GREEN, "J+28 ouvert"),
        "DEMO-002": (YELLOW, "J+12 ouvert"),
        "DEMO-003": (ORANGE, "J+5 ouvert"),
        "DEMO-004": (RED, "J+1 ouvert"),
        "DEMO-005": (GREY, "deadline depassee"),
        "DEMO-006": (GREY, "Soumis a J+2 : ne doit PAS etre rouge"),
        "DEMO-007": (GREY, "Gagne"),
        "DEMO-008": (GREY, "Perdu"),
        "DEMO-009": (GREY, "NO-GO malgre un score de 85"),
        "DEMO-010": (GREEN, "J+40 nouveau"),
    }

    for d in demo:
        offset = d["Deadline_Offset_Days"]
        row = {
            "status": d["Status"],
            "deadline": today + dt.timedelta(days=int(offset)) if offset else None,
            "days": int(offset) if offset else None,
        }
        applied = None
        for rule, fill in zip(row_rules, fills):
            if eval_rule(rule.formula[0], row, S.STATUSES_CLOSED):
                applied = fill
                break
        want, why = expected_colors[d["Opportunity_ID"]]
        check(f"{d['Opportunity_ID']} -> {want} ({why})", applied == want,
              f"obtenu {applied}")

    # --------------------------------------------- critere eliminatoire ----
    print("\n[6] Critere eliminatoire")
    el = S.col_letter(S.OPPORTUNITIES, "Eliminatory_Criterion")
    el_rng = f"{el}2:{el}{MAX_ROWS + 1}"
    el_rules = [rules for rng, rules in ws.conditional_formatting._cf_rules.items()
                if str(rng.sqref) == el_rng]
    check("la colonne a sa propre regle visuelle, independante des scores",
          len(el_rules) == 1 and len(el_rules[0]) == 1)
    d9 = next(d for d in demo if d["Opportunity_ID"] == "DEMO-009")
    check("le cas DEMO score eleve + critere eliminatoire est bien un NO-GO",
          d9["Eliminatory_Criterion"] == "OUI"
          and int(d9["Relevance_Score"]) >= 80
          and d9["Status"] == "NO-GO")

    # ------------------------------------------------------- dashboard -----
    print("\n[7] Dashboard")
    dash = wb["DASHBOARD"]
    win_formula = None
    for r in range(1, 60):
        if dash.cell(row=r, column=1).value == "Win rate":
            win_formula = dash.cell(row=r, column=2).value
            break
    check("le KPI Win rate existe", win_formula is not None)
    if win_formula:
        rows = sorted({int(x) for x in re.findall(r"B(\d+)", win_formula)})
        labels = [dash.cell(row=n, column=1).value for n in rows]
        check("le win rate reference bien les lignes Gagne et Perdu",
              labels == ["Gagne", "Perdu"], f"references {labels}")
        check("division par zero geree", '"n/a"' in win_formula)

    deadline_kpi = None
    for r in range(1, 40):
        if dash.cell(row=r, column=1).value == "Deadlines dans moins de 7 jours":
            deadline_kpi = dash.cell(row=r, column=2).value
    check("le KPI deadlines exclut les statuts clos",
          deadline_kpi is not None and "COUNTIF(LISTS!" in deadline_kpi
          and "=0)" in deadline_kpi)

    # -------------------------------------------------------- securite -----
    print("\n[8] Securite et donnees")
    settings_texts = [str(c.value) for row in wb["SETTINGS"].iter_rows()
                      for c in row if c.value]
    check("SETTINGS avertit de ne pas stocker de cle API",
          any("NE PAS stocker de cle API" in t for t in settings_texts))

    all_texts = [str(c.value) for sheet in wb.worksheets
                 for row in sheet.iter_rows() for c in row if c.value]
    suspicious = re.compile(r"(sk-[A-Za-z0-9]{16,}|AIza[A-Za-z0-9_-]{20,}|"
                            r"[0-9]{8,10}:AA[A-Za-z0-9_-]{30,})")
    hits = [t for t in all_texts if suspicious.search(t)]
    check("aucun token/cle present dans le fichier livre", not hits, str(hits))

    urls = [t for t in all_texts if t.startswith("http")]
    check("toutes les URL DEMO sont fictives (example.org)",
          bool(urls) and all("example.org" in u for u in urls),
          str([u for u in urls if "example.org" not in u]))

    # ---------------------------------------------------------- resultat ---

    # -------------------------------------------- validite des formules ---
    print("\n""[9] Validite syntaxique des formules")
    from formula_check import invalid_references, formula_count
    cassees = invalid_references(OUT_FILE)
    total = formula_count(OUT_FILE)
    check(f"les {total} formules du classeur ont des references valides",
          not cassees,
          str(cassees[:3]) + (f" (+{len(cassees) - 3})" if len(cassees) > 3 else ""))
    check("le classeur contient bien des formules", total > 0)

    print(f"\n{'-' * 62}")
    if FAILURES:
        print(f"ECHEC : {len(FAILURES)}/{CHECKS} verifications en echec")
        for f in FAILURES:
            print(f"  - {f}")
        return 1
    print(f"SUCCES : {CHECKS}/{CHECKS} verifications passees")
    return 0


if __name__ == "__main__":
    sys.exit(main())
