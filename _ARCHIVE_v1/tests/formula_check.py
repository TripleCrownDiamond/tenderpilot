"""
Validation syntaxique des references de plage dans les classeurs generes.

Ce module existe a cause d'un bug reel : les formules SUMPRODUCT du Dashboard
et de l'onglet DECISION etaient generees avec une borne de fin amputee de sa
lettre de colonne - `$P$2:$1001` au lieu de `$P$2:$P$1001`. Excel et Sheets
refusent cette reference et affichent #ERROR!.

53 formules etaient touchees, et aucun test ne l'a vu : ils verifiaient la
presence de fragments (`COUNTIF(LISTS!`, `=0)`) sans jamais valider la
syntaxe des references produites.

Desormais chaque classeur genere est scanne : toute plage dont les deux
bornes ne sont pas du meme type - cellule/cellule, colonne/colonne,
ligne/ligne - fait echouer le build.
"""

import re

from openpyxl import load_workbook

CELL = r"\$?[A-Z]{1,3}\$?[0-9]+"
COLUMN = r"\$?[A-Z]{1,3}"
ROW = r"\$?[0-9]+"

# L'ordre compte : `$A$2` doit etre reconnu comme une cellule, pas comme la
# colonne `$A` suivie d'autre chose.
TOKEN = f"(?:{CELL}|{COLUMN}|{ROW})"
SHEET = r"(?:[A-Za-z0-9_]+!)?"

PAIR = re.compile(rf"(?<![A-Z0-9$])({SHEET}{TOKEN})\s*:\s*({SHEET}{TOKEN})")
STRING_LITERAL = re.compile(r'"[^"]*"')


def _kind(token):
    """cell, col, row - ou None si le jeton n'est pas une reference valide."""
    token = token.split("!")[-1]
    if re.fullmatch(CELL, token):
        return "cell"
    if re.fullmatch(COLUMN, token):
        return "col"
    if re.fullmatch(ROW, token):
        return "row"
    return None


def malformed_ranges(formula):
    """Fragments de plage syntaxiquement invalides dans une formule."""
    # Les chaines litterales contiennent des deux-points ("Motif : ...") qui
    # ne sont pas des plages.
    naked = STRING_LITERAL.sub('""', formula)
    bad = []
    for match in PAIR.finditer(naked):
        left, right = _kind(match.group(1)), _kind(match.group(2))
        if left is None or right is None or left != right:
            bad.append(match.group(0))
    return bad


def invalid_references(path):
    """Retourne [(onglet, cellule, fragment), ...] pour un classeur."""
    workbook = load_workbook(path, data_only=False)
    found = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    for fragment in malformed_ranges(cell.value):
                        found.append((sheet.title, cell.coordinate, fragment))
    return found


def formula_count(path):
    """Nombre de cellules contenant une formule - garde-fou de couverture.

    Un classeur dont les formules auraient disparu passerait le test de
    validite sans rien signaler : on verifie donc aussi qu'il y en a.
    """
    workbook = load_workbook(path, data_only=False)
    return sum(
        1
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )


# --------------------------------------------------------------------------
# Auto-test du validateur : un detecteur qui ne detecte rien passerait
# inapercu. Ces cas tournent a chaque import depuis un test.
# --------------------------------------------------------------------------
_VALID = [
    "=SUM(OPPORTUNITIES!$P$2:$P$1001)",
    "=COUNTIF(LISTS!$D$2:$D$7,OPPORTUNITIES!$X$2:$X$1001)",
    "=SUM(A:A)",
    "=SUM(2:2)",
    '=IF(B12=0,"n/a",B10/(B10+B11))',
    '=IF($N2="","",$N2-TODAY())',
    '=IF(A1="","Evaluation incomplete : des criteres sont sans resultat.","")',
]
_INVALID = [
    "=SUM(OPPORTUNITIES!$P$2:$1001)",
    "=COUNTIF(LISTS!$D$2:$D$7,OPPORTUNITIES!$X$2:$1001)",
    "=SUM(CRITERIA!$O$2:$301)",
]

for _formula in _VALID:
    assert not malformed_ranges(_formula), f"faux positif : {_formula}"
for _formula in _INVALID:
    assert malformed_ranges(_formula), f"reference cassee non detectee : {_formula}"
