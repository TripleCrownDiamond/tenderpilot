"""
TenderPilot - conformite du classeur et coherence avec le script.

Les scenarios metier sont testes sous Node (tests/test_logic.js). Ici on
verifie ce que Node ne voit pas : la structure du classeur livre, et le fait
que le script et le classeur parlent des memes colonnes.

    python tests/test_sheet.py
"""

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from schema import columns as S  # noqa: E402
from builders.toolkit import (  # noqa: E402
    OUT_DIR, OUT_XLSX, SCHEMA_GS, SCRIPT_FILES,
)

ECHECS = []
CONTROLES = 0

# Fichiers ecrits a la main, par opposition a Schema.gs qui est genere.
MANUELS = ["Core.gs", "Rss.gs", "Html.gs", "Json.gs", "Sheet.gs",
           "Sources.gs", "Telegram.gs", "Run.gs"]


def check(libelle, condition, detail=""):
    global CONTROLES
    CONTROLES += 1
    if condition:
        print(f"  ok   {libelle}")
    else:
        print(f"  FAIL {libelle}  {detail}")
        ECHECS.append(libelle)


def lire(nom):
    return (ROOT / "apps_script" / nom).read_text(encoding="utf-8")


def main():
    print(f"\nClasseur teste : {OUT_XLSX}")
    if not OUT_XLSX.exists():
        print("FAIL : le classeur n'existe pas. Lancer d'abord le build.")
        return 1

    wb = load_workbook(OUT_XLSX, data_only=False)

    # ---------------------------------------------------------- structure --
    print("\n[1] Structure du classeur")
    attendus = ["LISEZ_MOI", S.SHEETS["opportunities"], S.SHEETS["sources"],
                S.SHEETS["config"], S.SHEETS["logs"]]
    check("cinq onglets, dans l'ordre", wb.sheetnames == attendus,
          str(wb.sheetnames))

    opp = wb[S.SHEETS["opportunities"]]
    entetes = [c.value for c in opp[1][:len(S.OPPORTUNITIES)]]
    check(f"les {len(S.OPPORTUNITIES)} colonnes du cahier des charges",
          entetes == S.OPPORTUNITIES,
          f"ecart : {set(S.OPPORTUNITIES) ^ set(e for e in entetes if e)}")
    check("les colonnes de notification sont masquees",
          all(opp.column_dimensions[S.col_letter(S.OPPORTUNITIES, n)].hidden
              for n in S.HIDDEN_COLUMNS))
    check("aucune opportunite pre-remplie",
          all(opp.cell(row=r, column=1).value is None
              for r in range(2, min(opp.max_row, 10) + 1)))

    src = wb[S.SHEETS["sources"]]
    check("les colonnes SOURCES correspondent au schema",
          [c.value for c in src[1][:len(S.SOURCES)]] == S.SOURCES)
    # On s'arrete a la premiere ligne vide : la note de bas de tableau
    # occupe elle aussi la colonne A et ne doit pas etre lue comme une source.
    lignes_src = []
    for r in range(2, src.max_row + 1):
        if not src.cell(row=r, column=1).value:
            break
        lignes_src.append({nom: src.cell(row=r, column=i + 1).value
                           for i, nom in enumerate(S.SOURCES)})
    check("des sources sont livrees pre-remplies", len(lignes_src) >= 10,
          f"{len(lignes_src)} sources")
    check("chaque source a un identifiant et un pays",
          all(l["Source_ID"] and l["Pays_Defaut"] for l in lignes_src))
    def methode_valide(m):
        return m in S.METHODES or str(m).startswith(S.METHODE_PREFIXES)
    check("chaque methode est RSS, MANUAL, HTML:<site> ou JSON:<site>",
          all(methode_valide(l["Methode"]) for l in lignes_src),
          str({l["Methode"] for l in lignes_src
               if not methode_valide(l["Methode"])}))

    rss = [l for l in lignes_src if l["Methode"] == "RSS"
           or str(l["Methode"]).startswith(S.METHODE_PREFIXES)]
    check("toutes les sources RSS ont une adresse https",
          all(str(l["URL"]).startswith("https://") for l in rss))
    check("toutes les sources RSS portent une trace de verification",
          all("Verifie" in str(l["Statut"]) for l in rss),
          str([l["Source_ID"] for l in rss if "Verifie" not in str(l["Statut"])]))

    manuelles = [l for l in lignes_src if l["Methode"] == "MANUAL"]
    check("une source sans flux n'est jamais active",
          all(l["Active"] == "NON" for l in manuelles),
          str([l["Source_ID"] for l in manuelles if l["Active"] != "NON"]))
    # Une frappe comme "ON" au lieu de "OUI", ou "OU" au lieu de "OUI",
    # eteint une source sans le moindre message : estVrai() ne reconnait que
    # true/vrai/oui/yes/1, et le moteur web ne compare qu'a "OUI". Le
    # registre a deja porte un "ON" pendant plusieurs jours.
    check("la colonne Active ne contient que OUI ou NON",
          all(str(l["Active"]).strip() in ("OUI", "NON") for l in lignes_src),
          str([(l["Source_ID"], l["Active"]) for l in lignes_src
               if str(l["Active"]).strip() not in ("OUI", "NON")]))
    check("les sources sont reparties sur plusieurs pays",
          len({l["Pays_Defaut"] for l in lignes_src}) >= 10,
          str(len({l["Pays_Defaut"] for l in lignes_src})) + " pays")

    cfg = wb[S.SHEETS["config"]]
    check("les colonnes CONFIG correspondent au schema",
          [c.value for c in cfg[1][:3]] == S.CONFIG_COLUMNS)
    cles = [cfg.cell(row=r, column=1).value for r in range(2, len(S.CONFIG) + 2)]
    check(f"les {len(S.CONFIG_KEYS)} cles de configuration sont presentes",
          cles == S.CONFIG_KEYS, str(cles))
    check("chaque cle porte une description",
          all(cfg.cell(row=r, column=3).value
              for r in range(2, len(S.CONFIG) + 2)))
    check("l'email de notification est vide par defaut",
          cfg.cell(row=2, column=2).value in ("", None))

    check("les colonnes LOGS correspondent au schema",
          [c.value for c in wb[S.SHEETS["logs"]][1][:len(S.LOGS)]] == S.LOGS)

    # ------------------------------------------------------ pas de formule --
    print("\n[2] Le classeur ne contient aucune formule")
    formules = [f"{f.title}!{c.coordinate}"
                for f in wb.worksheets for row in f.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    check("tout est calcule par le script, rien par le tableur",
          not formules, str(formules[:5]))

    # ------------------------------------------------- Schema.gs genere ----
    print("\n[3] Schema.gs est genere depuis le schema")
    schema_js = SCHEMA_GS.read_text(encoding="utf-8")
    check("le fichier annonce qu'il est genere",
          "NE PAS MODIFIER A LA MAIN" in schema_js)
    check("il nomme son generateur", "builders/toolkit.py" in schema_js)
    check("il porte la version du schema", S.SCHEMA_VERSION in schema_js)
    for feuille in S.SHEETS.values():
        check(f"l'onglet {feuille} est expose au script",
              f'"{feuille}"' in schema_js)
    for statut in S.DELAI_STATUTS:
        check(f"le statut {statut} est expose", f'"{statut}"' in schema_js)

    # --------------------------- le script ne code aucun nom en dur --------
    print("\n[4] Aucun nom d'onglet ni de colonne en dur dans le script")
    noms = set(S.SHEETS.values()) | set(S.OPPORTUNITIES) | set(S.SOURCES)
    for fichier in MANUELS:
        source = lire(fichier)
        en_dur = [n for n in noms if f'"{n}"' in source or f"'{n}'" in source]
        check(f"{fichier} passe par SCHEMA", not en_dur, str(en_dur))

    joint = "\n".join(lire(f) for f in MANUELS)
    for motif, connues, libelle in [
        (r"SCHEMA\.OPP\.(\w+)", set(S.OPP_KEYS), "SCHEMA.OPP"),
        (r"SCHEMA\.SRC\.(\w+)", set(S.SOURCE_KEYS), "SCHEMA.SRC"),
        (r"SCHEMA\.SHEETS\.(\w+)", set(S.SHEETS), "SCHEMA.SHEETS"),
    ]:
        utilisees = set(re.findall(motif, joint))
        inconnues = sorted(utilisees - connues)
        check(f"toutes les cles {libelle} existent ({len(utilisees)} utilisees)",
              not inconnues, str(inconnues))

    manquantes = sorted(c for c in S.OPP_KEYS.values() if c not in entetes)
    check("chaque cle SCHEMA.OPP vise une colonne existante",
          not manquantes, str(manquantes))

    # ----------------------------------------------------- perimetre MVP ---
    print("\n[5] Perimetre : ce qui a ete retire n'est pas revenu")
    retires = ["relevanceScore", "verdict", "GoNoGo", "watchlist", "GmailApp",
               # Telegram a ete reintroduit le 2026-09-01 comme second
               # canal de notification : voir Telegram.gs.
               "PostgreSQL", "Supabase", "getUserLabelByName",
               "HtmlService"]
    presents = [m for m in retires if m in joint]
    check("aucun module hors MVP dans le code", not presents, str(presents))

    manifeste = json.loads(lire("appsscript.json"))
    scopes = manifeste.get("oauthScopes", [])
    check("aucune autorisation Gmail demandee",
          not any("gmail" in s for s in scopes), str(scopes))
    check("l'acces au classeur est limite au fichier courant",
          any("spreadsheets.currentonly" in s for s in scopes))
    check("l'acces reseau est declare (lecture des flux)",
          any("script.external_request" in s for s in scopes))
    check("l'envoi d'email est declare",
          any("script.send_mail" in s for s in scopes))

    # ---------------------------------------------------------- livrable ---
    print("\n[6] Livrable")
    for nom in SCRIPT_FILES + ["TenderPilot.xlsx", "README.md"]:
        check(f"{nom} present", (OUT_DIR / nom).exists())
    readme = (OUT_DIR / "README.md").read_text(encoding="utf-8")
    for section in ["Installation", "Configuration des emails",
                    "Ajouter une source", "Lancer manuellement",
                    "Execution automatique"]:
        check(f"le README couvre : {section}", section in readme)

    # Le guide d'installation doit nommer CHAQUE fichier de script livre.
    # Un fichier oublie ne provoque aucune erreur visible : le classeur
    # s'installe, et une partie des sources cesse simplement de repondre.
    for nom in SCRIPT_FILES:
        if not nom.endswith(".gs") or nom == "Schema.gs":
            continue
        court = nom[:-3]
        check(f"le guide fait creer le fichier {court}",
              f"| `{court}` |" in readme, nom)
    check("le guide annonce le bon nombre de fichiers",
          f"les {len([f for f in SCRIPT_FILES if f.endswith('.gs')])} fichiers"
          in readme)



    # ------------------- les quatre copies du registre concordent ----------
    #
    # data/sources.csv est la reference. Trois copies en derivent : l'onglet
    # SOURCES du classeur, le catalogue embarque dans Schema.gs, et le
    # registre TypeScript du web. Une copie qui derive n'est signalee nulle
    # part : la source est simplement absente d'un des deux moteurs. C'est
    # exactement ce qui etait arrive a la source BAD.
    print("\n[5] Le registre de sources est identique partout")

    reference = {l["Source_ID"].strip(): l for l in lignes_src}

    # -- copie 1 : l'onglet SOURCES du classeur --
    onglet = wb[S.SHEETS["sources"]]
    dans_onglet = set()
    for ligne in onglet.iter_rows(min_row=2, values_only=True):
        code = str(ligne[0]).strip() if ligne and ligne[0] else ""
        # La derniere ligne est une note d'aide, pas une source.
        if code in reference:
            dans_onglet.add(code)
    check("le classeur porte toutes les sources du CSV",
          dans_onglet == set(reference),
          str(sorted(set(reference) - dans_onglet)[:5]))

    # -- copie 2 : le catalogue embarque dans Schema.gs --
    schema_txt = SCHEMA_GS.read_text(encoding="utf-8")
    check("Schema.gs embarque le catalogue", "SOURCES_LIVREES" in schema_txt)
    debut = schema_txt.index("[", schema_txt.index("SOURCES_LIVREES:"))
    profondeur, fin = 0, debut
    for i in range(debut, len(schema_txt)):
        if schema_txt[i] == "[":
            profondeur += 1
        elif schema_txt[i] == "]":
            profondeur -= 1
            if profondeur == 0:
                fin = i
                break
    catalogue = {l[0].strip(): l for l in json.loads(schema_txt[debut:fin + 1])}
    check("le catalogue embarque couvre tout le CSV",
          set(catalogue) == set(reference),
          str(sorted(set(reference) ^ set(catalogue))[:5]))
    check("chaque ligne du catalogue a autant de colonnes que l onglet",
          all(len(l) == len(S.SOURCES) for l in catalogue.values()))

    # -- copie 3 : le registre TypeScript du web --
    ts = ROOT / "web" / "src" / "data" / "sources-defaut.ts"
    if ts.exists():
        texte = ts.read_text(encoding="utf-8")
        # Le "[" de "SourceDefaut[]" precede le vrai tableau : on part
        # apres le signe egal, et on equilibre les crochets.
        d = texte.index("[", texte.index("SOURCES_DEFAUT: SourceDefaut[] =") + 32)
        web = {x["code"]: x for x in json.loads(texte[d:texte.rindex("]") + 1])}
        check("le registre web couvre tout le CSV",
              set(web) == set(reference),
              str(sorted(set(reference) ^ set(web))[:5]))
        divergences = [
            code for code, r in reference.items()
            if code in web and (
                web[code]["url"] != r["URL"].strip()
                or web[code]["methode"] != r["Methode"].strip()
                or (web[code]["typeDefaut"] or "") != (r["Type_Defaut"] or "").strip()
                or web[code]["active"] != (r["Active"].strip().upper() == "OUI"))
        ]
        check("aucune divergence de contenu entre le web et le CSV",
              not divergences, str(divergences[:5]))
    else:
        check("le registre web existe", False, str(ts))

    # -- l'onglet technique est livre masque --
    check("l onglet SOURCES est livre masque",
          onglet.sheet_state == "hidden", str(onglet.sheet_state))
    for cle in ("opportunities", "config", "logs"):
        feuille = wb[S.SHEETS[cle]]
        check("l onglet " + feuille.title + " reste visible",
              feuille.sheet_state == "visible", str(feuille.sheet_state))


    # ------------------------------------------- les guides PDF generes ---
    print("\n[6] Les guides PDF sont complets et lisibles")

    guides = OUT_DIR / "guides"
    attendus = {
        "client/1_Guide_Demarrage.pdf": "le client : demarrer avec le lien",
        "client/2_Catalogue_des_Sources.pdf": "le client : le catalogue",
        "operateur/1_Guide_Operateur.pdf": "vous : preparer et vendre",
        "operateur/2_Installation_Manuelle.pdf": "vous : fabriquer le maitre",
        "operateur/3_Guide_Application_Web.pdf": "l autre produit",
    }
    for nom, quoi in attendus.items():
        fichier = guides / nom
        check("le guide pour " + quoi + " est genere", fichier.exists(), nom)
        # Un PDF vide fait quelques centaines d octets : il aurait l air
        # present sans rien contenir.
        if fichier.exists():
            check("le guide pour " + quoi + " n est pas vide",
                  fichier.stat().st_size > 4000,
                  str(fichier.stat().st_size) + " octets")

    # Chaque jeu se numerote a partir de 1 dans son dossier : un client qui
    # recoit un "guide 3" se demande ou sont les deux premiers.
    for dossier in ("client", "operateur"):
        numeros = sorted(int(f.name.split("_")[0])
                         for f in (guides / dossier).glob("*.pdf"))
        check("les guides " + dossier + " se numerotent depuis 1",
              numeros == list(range(1, len(numeros) + 1)), str(numeros))

    try:
        import pdfplumber
        import re as _re

        manuel = guides / "operateur" / "2_Installation_Manuelle.pdf"
        with pdfplumber.open(manuel) as pdf:
            texte = chr(10).join((p.extract_text() or "") for p in pdf.pages)
        nb = len([f for f in SCRIPT_FILES if f.endswith(".gs")])
        check("le guide PDF annonce le bon nombre de fichiers",
              "les " + str(nb) + " fichiers" in texte)
        check("aucune balise de mise en forme n a fuite dans le PDF",
              not any(m in texte for m in ("<b>", "<font", "<link")))

        # Les numeros doivent se suivre.
        #
        # Une etape suivie d un bloc de code ou d un tableau coupe la liste
        # en deux. Sans precaution, chaque morceau repart a 1 et le guide
        # affiche "1, 2, 3, 1, 1, 1" : le lecteur perd le fil au milieu
        # d une installation.
        motif = _re.compile("^(\d+)\s+\S")
        for chemin in sorted(guides.rglob("*.pdf")):
            with pdfplumber.open(chemin) as pdf:
                rendu = chr(10).join((p.extract_text() or "")
                                     for p in pdf.pages)
            suites, courante = [], []
            for ligne in rendu.split(chr(10)):
                m = motif.match(ligne)
                if m:
                    courante.append(int(m.group(1)))
                elif courante:
                    suites.append(courante)
                    courante = []
            if courante:
                suites.append(courante)
            cassees = [x for x in suites
                       if len(x) > 1 and x != list(range(x[0], x[0] + len(x)))]
            check("numerotation continue dans " + chemin.name,
                  not cassees, str(cassees[:2]))
    except ImportError:
        print("  (pdfplumber absent : controle du contenu PDF saute)")

    # ------------------------------- les archives pretes a la vente -------
    print("\n[7] Les archives de livraison sont completes")

    import zipfile
    from builders.toolkit import VERSION

    vente = (ROOT / "dist" / "A_VENDRE"
             / ("TenderPilot_Sheets_v" + VERSION + ".zip"))
    prive = (ROOT / "dist" / "PRIVE_NE_PAS_ENVOYER"
             / ("TenderPilot_OPERATEUR_v" + VERSION + ".zip"))

    for etiquette, archive in (("_a_vendre", vente), ("_privee", prive)):
        check("l archive" + etiquette + " existe", archive.exists(),
              archive.name)
        if archive.exists():
            with zipfile.ZipFile(archive) as z:
                check("l archive" + etiquette + " n est pas corrompue",
                      z.testzip() is None)

    # Le client ne recoit AUCUN element du produit lui-meme. Sans les
    # fichiers, le produit ne peut etre ni revendu ni redistribue.
    if vente.exists():
        with zipfile.ZipFile(vente) as z:
            noms_vente = z.namelist()
        for interdit, quoi in ((".gs", "fichier de script"),
                               (".xlsx", "classeur"),
                               ("Installation_Manuelle", "methode manuelle"),
                               ("Guide_Operateur", "guide operateur")):
            check("aucun " + quoi + " dans l archive a vendre",
                  not any(interdit in n for n in noms_vente))
        check("l archive a vendre tient en trois fichiers",
              len(noms_vente) == 3, str(len(noms_vente)))
        for attendu in ("1_Guide_Demarrage.pdf", "2_Catalogue_des_Sources.pdf",
                        "COMMENCEZ_ICI.txt"):
            check(attendu + " est dans l archive a vendre",
                  any(n.endswith(attendu) for n in noms_vente))

    if prive.exists():
        with zipfile.ZipFile(prive) as z:
            noms_prive = z.namelist()
        absents = [n for n in SCRIPT_FILES
                   if not any(x.endswith("/script/" + n) for x in noms_prive)]
        check("tous les scripts sont dans l archive privee",
              not absents, str(absents))
        check("le guide operateur est dans l archive privee",
              any("1_Guide_Operateur" in n for n in noms_prive))
        # L operateur doit avoir sous les yeux ce que le client lit.
        check("l archive privee contient la copie des docs client",
              any("/docs_client/" in n for n in noms_prive))

    check("un seul fichier a vendre dans A_VENDRE",
          len(list((ROOT / "dist" / "A_VENDRE").glob("*.zip"))) == 1)
    check("le dossier prive porte son avertissement",
          (ROOT / "dist" / "PRIVE_NE_PAS_ENVOYER"
           / "NE_PAS_ENVOYER_AU_CLIENT.txt").exists())

    # dist/ est reconstruit a chaque build : sans copie datee, la version
    # livree hier disparait des qu on relance.
    archives_v = ROOT / "dist" / "ARCHIVES" / ("v" + VERSION)
    check("la version publiee est archivee", archives_v.is_dir(),
          str(archives_v))
    if archives_v.is_dir():
        check("les deux archives sont conservees",
              len(list(archives_v.glob("*.zip"))) == 2)


    # ------------------------- la reference des agents est unique ---------
    #
    # AGENTS.md est lu par Cursor, Codex, Aider. Claude Code lit une skill.
    # Les deux doivent dire exactement la meme chose : deux guides qui
    # divergent, c'est deux agents qui travaillent selon des regles
    # differentes sur le meme depot.
    print("\n[8] Les agents lisent tous la meme reference")

    agents = ROOT / "AGENTS.md"
    skill = ROOT / ".claude" / "skills" / "tenderpilot" / "SKILL.md"
    claude = ROOT / "CLAUDE.md"

    check("AGENTS.md existe", agents.exists())
    check("la skill Claude existe", skill.exists())
    check("CLAUDE.md renvoie vers AGENTS.md",
          claude.exists() and "AGENTS.md" in claude.read_text(encoding="utf-8"))

    if agents.exists() and skill.exists():
        texte_agents = agents.read_text(encoding="utf-8")
        texte_skill = skill.read_text(encoding="utf-8")
        check("la skill porte son frontmatter",
              texte_skill.startswith("---\nname: tenderpilot"))
        # Le corps doit etre identique, commentaire d en-tete mis a part.
        corps = texte_agents
        if corps.startswith("<!--"):
            corps = corps[corps.index("-->") + 3:].lstrip()
        check("la skill dit exactement ce que dit AGENTS.md",
              corps.strip() in texte_skill)
        # Les regles de fond doivent y etre.
        # La convention "sans accents" vaut pour le code livre a Google
        # Sheets, pas pour la documentation : on cherche le texte reel.
        for regle in ("Vérifier, jamais supposer",
                      "Ne jamais inventer une date",
                      "Parité entre les deux moteurs"):
            check("AGENTS.md porte la regle : " + regle.split(",")[0],
                  regle in corps, regle)


    print(f"\n{'-' * 58}")
    if ECHECS:
        print(f"ECHEC : {len(ECHECS)}/{CONTROLES} controles en echec")
        for e in ECHECS:
            print(f"  - {e}")
        return 1
    print(f"SUCCES : {CONTROLES}/{CONTROLES} controles passes")
    return 0


if __name__ == "__main__":
    sys.exit(main())
