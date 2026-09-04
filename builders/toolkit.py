"""
TenderPilot MVP - generateur unique.

Produit dans dist/TenderPilot/ :
  - TenderPilot.xlsx        le classeur a importer dans Google Sheets
  - les fichiers Apps Script (dont Schema.gs, genere depuis le schema)
  - README.md               installation, configuration, exploitation

Un seul builder : le produit tient en un classeur et un script.
"""

import csv
import datetime as dt
import json
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schema import columns as S

ROOT = Path(__file__).resolve().parent.parent
SCRIPT_DIR = ROOT / "apps_script"
OUT_DIR = ROOT / "dist" / "TenderPilot"
OUT_XLSX = OUT_DIR / "TenderPilot.xlsx"
SCHEMA_GS = SCRIPT_DIR / "Schema.gs"

VERSION = "1.0.0"
MAX_ROWS = 2000

SCRIPT_FILES = ["appsscript.json", "Schema.gs", "Core.gs", "Rss.gs",
                "Html.gs", "Json.gs", "Sheet.gs", "Sources.gs",
                "Telegram.gs", "Llm.gs",
                "Run.gs"]

# Couleurs en ARGB opaque. Un code a 6 chiffres est complete par openpyxl
# avec "00", soit un remplissage TRANSPARENT que ni Excel ni Sheets
# n'affichent.
INK = "FF16202D"
HEAD_BG = "FF1F3A5F"
MUTED = "FF6B7280"
RULE = "FFD5DBE3"

F_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
F_BODY = Font(name="Calibri", size=10, color=INK)
F_MUTED = Font(name="Calibri", size=9, color=MUTED, italic=True)
F_TITLE = Font(name="Calibri", size=13, bold=True, color=HEAD_BG)
FILL_HEAD = PatternFill("solid", fgColor=HEAD_BG)
THIN = Side(style="thin", color=RULE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WIDTHS = {
    "Opportunite": 52, "Organisation": 26, "Resume": 60, "Lien": 34,
    "PDF": 28, "Source": 14, "Pays": 16, "Secteur": 20, "Type": 18,
    "Date_Ajout": 17, "Derniere_MAJ": 17, "Statut_Delai": 17,
    "Jours_Restants": 14, "Deadline": 13, "Date_Publication": 16, "ID": 13,
    "Nom": 30, "URL": 40, "Methode": 12, "Pays_Defaut": 16,
    "Secteur_Defaut": 20, "Type_Defaut": 18, "Active": 10, "Budget": 18,
    "Derniere_Collecte": 18, "Statut": 12, "Source_ID": 14,
    "Cle": 26, "Valeur": 24, "Description": 68,
    "Date": 17, "Action": 20, "Message": 62,
    "Type": 12, "Valeur": 34, "Annonces": 12, "Suivi": 10,
}

def lire_sources():
    """Sources livrees pre-remplies.

    Chaque flux RSS a ete recupere et verifie avant d'entrer dans ce
    fichier : la colonne Statut porte la date du controle et le nombre
    d'annonces trouvees ce jour-la. Une source sans flux exploitable est
    livree en MANUAL et desactivee, jamais en RSS.
    """
    chemin = ROOT / "data" / "sources.csv"
    if not chemin.exists():
        return []
    with open(chemin, encoding="utf-8", newline="") as fh:
        return [[r[c] for c in S.SOURCES] for r in csv.DictReader(fh)]


def entete(ws, colonnes, row=1):
    for i, nom in enumerate(colonnes, start=1):
        c = ws.cell(row=row, column=i, value=nom)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(nom, 15)
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def feuille_opportunites(wb):
    ws = wb.create_sheet(S.SHEETS["opportunities"])
    entete(ws, S.OPPORTUNITIES)
    ws.auto_filter.ref = (f"A1:{get_column_letter(len(S.OPPORTUNITIES))}"
                          f"{MAX_ROWS}")
    for nom in S.HIDDEN_COLUMNS:
        ws.column_dimensions[S.col_letter(S.OPPORTUNITIES, nom)].hidden = True
    return ws


def feuille_sources(wb):
    ws = wb.create_sheet(S.SHEETS["sources"])
    entete(ws, S.SOURCES)
    sources = lire_sources()
    for r, ligne in enumerate(sources, start=2):
        for i, valeur in enumerate(ligne, start=1):
            ws.cell(row=r, column=i, value=valeur).font = F_BODY
    # La note occupe la colonne B, pas la colonne A : la colonne A est le
    # Source_ID, et le script lit comme une source toute ligne qui en porte
    # un. Le 2026-09-02, cette note apparaissait a chaque execution dans le
    # journal, sous la forme d'une source "Methode : RSS ou JSON..."
    # desactivee.
    ws.cell(row=len(sources) + 3, column=2,
            value="Methode : RSS ou JSON:<site> et HTML:<site> (collecte "
                  "automatique), MANUAL (saisie a la main). Active : OUI pour "
                  "collecter. Pour ecarter une source, mettez NON plutot que "
                  "de supprimer la ligne : la synchronisation la remettrait."
            ).font = F_MUTED
    ws.auto_filter.ref = f"A1:{get_column_letter(len(S.SOURCES))}{len(sources) + 1}"
    # VISIBLE. L'onglet a longtemps ete livre masque - c'etait de la
    # plomberie. Il ne l'est plus : la colonne Active est le seul endroit ou
    # le client choisit ce qu'il surveille, et une commande de menu pour
    # afficher un onglet est une marche de trop devant le reglage le plus
    # utile du produit. Le menu TenderPilot garde de quoi le remasquer.
    ws.sheet_state = "visible"
    return ws


def feuille_config(wb):
    ws = wb.create_sheet(S.SHEETS["config"])
    entete(ws, S.CONFIG_COLUMNS)
    for r, (cle, valeur, description) in enumerate(S.CONFIG, start=2):
        ws.cell(row=r, column=1, value=cle).font = Font(
            name="Consolas", size=10, bold=True, color=INK)
        v = ws.cell(row=r, column=2, value=valeur)
        v.font = F_BODY
        v.border = BORDER
        d = ws.cell(row=r, column=3, value=description)
        d.font = F_MUTED
        d.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=len(S.CONFIG) + 3, column=1,
            value="Ne modifiez que la colonne Valeur. Les cles sont lues par "
                  "le script.").font = F_MUTED
    return ws


def feuille_logs(wb):
    ws = wb.create_sheet(S.SHEETS["logs"])
    entete(ws, S.LOGS)
    ws.cell(row=2, column=1,
            value="Rempli automatiquement a chaque execution.").font = F_MUTED
    return ws


def feuille_profil(wb):
    """L'inventaire des pays et secteurs reellement collectes.

    Livre vide : il se remplit au premier passage, depuis les annonces
    reellement collectees. Une liste ecrite d'avance mentirait - elle
    afficherait des pays qu'aucune source active ne publie.
    """
    ws = wb.create_sheet(S.SHEETS["profil"])
    entete(ws, S.PROFIL)
    ws.cell(row=2, column=1,
            value="Rempli automatiquement a chaque execution : recopiez les "
                  "valeurs qui vous interessent dans PAYS_SUIVIS et "
                  "SECTEURS_SUIVIS, onglet CONFIG.").font = F_MUTED
    ws.auto_filter.ref = f"A1:{get_column_letter(len(S.PROFIL))}{MAX_ROWS}"
    return ws


def feuille_demarrage(wb):
    ws = wb.active
    ws.title = "LISEZ_MOI"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 80
    ws.sheet_view.showGridLines = False

    # Style constants for this page only
    F_SECTION = Font(name="Calibri", size=12, bold=True, color=HEAD_BG)
    F_HIGHLIGHT = Font(name="Calibri", size=10, bold=True, color=INK)
    ROW_HEIGHT_TITLE = 40
    ROW_HEIGHT_VERSION = 24
    ROW_HEIGHT_SECTION = 32
    ROW_HEIGHT_BODY = 80
    ROW_HEIGHT_BULLET = 32
    ROW_HEIGHT_COLOR = 24
    BLANK_HEIGHT = 10

    # (row_type, text) where row_type is:
    #   "title"    = big title (TENDERPILOT)
    #   "version"  = version line
    #   "section"  = section header
    #   "body"     = normal body text
    #   "bullet"   = numbered/bulleted step
    #   "blank"    = empty spacer row
    #   "color"    = color legend line (with color dot placeholder)
    lignes = [
        ("title",    "TENDERPILOT"),
        ("version",  f"Version {VERSION}"),
        ("blank",    ""),
        ("section",  "Ce que fait TenderPilot"),
        ("body",     "TenderPilot surveille les appels d offres, marches "
                     "publics, subventions et bourses. Il collecte les "
                     "annonces, evite les doublons, calcule les jours "
                     "restants avant la date limite, colore les lignes "
                     "et vous envoie un email ou une notification Telegram "
                     "quand une echeance approche."),
        ("blank",    ""),
        ("section",  "Demarrer en 3 etapes"),
        ("bullet",  "1.  Menu TenderPilot > Activer l execution automatique."),
        ("bullet",  "2.  Onglet CONFIG : renseignez votre adresse email "
                     "(NOTIFICATION_EMAIL) pour recevoir les alertes."),
        ("bullet",  "3.  Onglet CONFIG : PAYS_SUIVIS et SECTEURS_SUIVIS "
                     "decrivent votre metier."),
        ("blank",    ""),
        ("section",  "Ce qui vous concerne : la colonne Pertinence"),
        ("body",     "Elle repond a une seule question : est-ce que cela me "
                     "concerne ? Elle croise le pays et le secteur de "
                     "l annonce avec vos PAYS_SUIVIS et SECTEURS_SUIVIS."),
        ("color",    "3 - PRIORITAIRE   votre pays ET votre domaine"),
        ("color",    "2 - A VOIR        l un des deux, ou un appel mondial"),
        ("color",    "1 - POSSIBLE      ni l un ni l autre, rien ne l exclut"),
        ("color",    "0 - HORS PROFIL   ailleurs"),
        ("body",     "Rien n est jamais supprime : une annonce hors profil "
                     "reste dans le tableau, rangee en dernier. Triez la "
                     "colonne de Z vers A pour voir vos opportunites en "
                     "premier."),
        ("body",     "Changez PAYS_SUIVIS ou SECTEURS_SUIVIS : tout le "
                     "tableau se remet a jour au passage suivant, sans rien "
                     "recollecter."),
        ("blank",    ""),
        ("section",  "Regler votre profil sans rien inventer"),
        ("body",     "L onglet PAYS_ET_SECTEURS se remplit tout seul a "
                     "chaque passage : il liste les pays et les secteurs "
                     "REELLEMENT collectes, avec le nombre d annonces, et "
                     "dit lesquels vous suivez aujourd hui."),
        ("body",     "Recopiez de la les valeurs qui vous interessent dans "
                     "PAYS_SUIVIS et SECTEURS_SUIVIS. Un pays ecrit de "
                     "memoire, qu aucune source ne publie, ne remonterait "
                     "rien - et vous ne le verriez jamais."),
        ("blank",    ""),
        ("section",  "Choisir les alertes que vous recevez"),
        ("body",     "NOTIFIER_PERTINENCE, onglet CONFIG : listez les "
                     "niveaux qui doivent vous ecrire. Par exemple 3 - "
                     "PRIORITAIRE seul, ou 3 - PRIORITAIRE, 2 - A VOIR. "
                     "Vide = tout vous est notifie."),
        ("body",     "Ce reglage coupe le bruit dans votre boite, jamais "
                     "dans le tableau : les annonces ecartees restent la, "
                     "avec leur couleur et leur echeance. Elargissez le "
                     "reglage et leurs alertes repartent au passage suivant."),
        ("blank",    ""),
        ("section",  "Choisir ce qui est surveille"),
        ("body",     "L onglet SOURCES liste tout ce que TenderPilot sait "
                     "lire. La colonne Active decide : OUI pour collecter, "
                     "NON pour ignorer. Ecrivez NON plutot que de supprimer "
                     "la ligne : la synchronisation la remettrait."),
        ("body",     "La colonne Statut dit ce que chaque source a donne a "
                     "la derniere execution."),
        ("blank",    ""),
        ("section",  "Lire le tableau"),
        ("body",     "Les lignes sont rangees a chaque passage : celles qui "
                     "vous laissent le plus de temps en haut, les plus "
                     "urgentes ensuite, les expirees plus bas, et celles "
                     "dont la source n a pas publie de date tout en bas. A "
                     "delai egal, la plus pertinente passe devant."),
        ("body",     "Chaque ligne est une opportunite. Les colonnes "
                     "importantes : opportunite, organisation, pays, type, "
                     "secteur, budget, date limite, jours restants, statut, "
                     "pertinence."),
        ("body",     "La colonne Budget n est remplie que quand la source "
                     "publie un montant - le portail europeen et Fundpilote "
                     "le font, les autres non. Une case vide veut dire que "
                     "la source ne l a pas dit, jamais que le marche est "
                     "sans budget."),
        ("blank",    ""),
        ("section",  "Les couleurs"),
        ("body",     "Le systeme colore chaque ligne selon le delai restant. "
                     "La colonne Statut_Delai donne la meme information en "
                     "texte."),
        ("color",    "OUVERT  (vert)      plus de 15 jours"),
        ("color",    "A SURVEILLER (jaune)   8 a 15 jours"),
        ("color",    "BIENTOT (orange)     4 a 7 jours"),
        ("color",    "URGENT (rouge)       0 a 3 jours"),
        ("color",    "EXPIRE (gris)        deadline depassee"),
        ("color",    "DATE A VERIFIER      aucune deadline connue"),
        ("blank",    ""),
        ("section",  "Notifications"),
        ("body",     "Par defaut, les alertes partent par email. Pour "
                     "recevoir aussi sur Telegram, creez un bot via "
                     "@BotFather puis renseignez TELEGRAM_TOKEN et "
                     "TELEGRAM_CHAT_ID dans l onglet CONFIG."),
        ("body",     "Le tout premier passage trouve beaucoup d echeances "
                     "proches d un coup. Deux reglages evitent l avalanche : "
                     "au-dela de DIGEST_THRESHOLD nouveautes, un seul email "
                     "recapitulatif ; et MAX_EMAILS_PAR_EXECUTION limite les "
                     "rappels envoyes par passage."),
        ("body",     "Ce qui depasse le plafond n est pas perdu : les "
                     "alertes repartent au passage suivant, les plus "
                     "pertinentes et les plus urgentes d abord. Google "
                     "n accepte que 100 destinataires par jour sur une "
                     "adresse gmail.com ordinaire."),
        ("blank",    ""),
        ("section",  "Classement intelligent (optionnel)"),
        ("body",     "Sans cle : collecte, dates, couleurs, alertes - le "
                     "produit est complet. Avec une cle LLM : en plus, "
                     "tri par type et secteur, resume lisible, articles et "
                     "FAQ ecartes. Fournissez votre cle dans CONFIG (LLM_CLE) "
                     "et activez USE_LLM. Le modele ne touche jamais aux "
                     "dates."),
        ("blank",    ""),
        ("section",  "A quelle heure ca tourne"),
        ("body",     "Trois passages par jour : 8h, 13h et 18h, a QUINZE "
                     "MINUTES PRES. Google repartit la charge de tous ses "
                     "utilisateurs et ne garantit pas la minute exacte : un "
                     "passage de 8h peut arriver a 7h50 ou 8h10. Ce n est "
                     "pas un reveil, c est une veille."),
        ("body",     "Le classeur n a pas besoin d etre ouvert, ni votre "
                     "ordinateur allume : le passage tourne sur les serveurs "
                     "de Google, sous votre compte."),
        ("blank",    ""),
        ("section",  "Certaines sources sont lues en deux temps"),
        ("body",     "Quelques sites listent leurs avis sans jamais ecrire "
                     "la date limite dans la liste : elle n existe que sur "
                     "la fiche de chaque avis. TenderPilot va donc la "
                     "chercher fiche par fiche, dans la limite de "
                     "MAX_FICHES_PAR_PASSAGE."),
        ("body",     "Ces sources arrivent donc par petits paquets, quelques "
                     "annonces a chaque passage, jusqu a ce que tout soit "
                     "rattrape. Une annonce que nous n avons pas pu dater n "
                     "entre pas : mieux vaut l attendre un passage de plus "
                     "que remplir votre tableau de lignes mortes."),
        ("blank",    ""),
        ("section",  "Si une execution est trop longue"),
        ("body",     "Google arrete toute execution a 6 minutes. Au-dela de "
                     "BUDGET_COLLECTE_SECONDES (4 minutes par defaut), la "
                     "collecte s arrete d elle-meme : ce qui a ete lu est "
                     "enregistre normalement, avec les deadlines, les "
                     "couleurs et les alertes."),
        ("body",     "Les sources non lues passent en tete au passage "
                     "suivant : rien n est oublie, le tour se fait sur "
                     "quelques executions. Le journal vous dit combien ont "
                     "ete reportees et par laquelle on reprendra."),
        ("blank",    ""),
        ("section",  "Quand une case reste vide"),
        ("body",     "Une case vide veut toujours dire LA SOURCE NE L A PAS "
                     "DIT, jamais que l information n existe pas. Nous "
                     "n inventons ni date limite ni montant : une date "
                     "devinee vous ferait manquer un depot."),
        ("body",     "Le portail des marches publics du Benin ne publie "
                     "aucune echeance dans son flux : ses lignes restent en "
                     "DATE A VERIFIER, et la date limite est dans le PDF de "
                     "l avis. Le budget, lui, n est publie que par le portail "
                     "europeen et Fundpilote."),
        ("blank",    ""),
        ("section",  "Pourquoi un lien ouvre parfois une liste"),
        ("body",     "Le portail beninois n a pas de page par avis - meme "
                     "dans un navigateur, une ligne n a pour seul lien que "
                     "son PDF, dont l adresse n est publiee nulle part. Le "
                     "lien vous amene donc a la liste : retrouvez l avis par "
                     "son objet ou son autorite contractante."),
        ("body",     "Quelques sites, comme l UNICEF, refusent les visites "
                     "automatiques mais s ouvrent normalement chez vous. Le "
                     "lien est bon."),
        ("blank",    ""),
        ("section",  "Repartir de zero"),
        ("body",     "Menu TenderPilot > Vider les opportunites et le "
                     "journal efface le tableau ET l onglet des journaux, "
                     "apres confirmation. Vos sources et votre configuration "
                     "ne sont jamais touchees."),
        ("body",     "Les temoins d envoi partent avec les lignes : la "
                     "collecte suivante renverra donc les alertes de ces "
                     "memes opportunites. Pensez au plafond "
                     "MAX_EMAILS_PAR_EXECUTION juste apres un vidage."),
        ("blank",    ""),
        ("section",  "Ce que TenderPilot ne fait pas"),
        ("body",     "Il ne remplit pas vos dossiers. Il n invente aucune "
                     "date limite : quand la source ne l ecrit pas, la case "
                     "reste vide. Il ne garantit pas l exhaustivite. "
                     "Verifiez toujours l avis officiel avant de candidater."),
    ]

    row = 2
    for row_type, texte in lignes:
        if row_type == "blank":
            ws.row_dimensions[row].height = BLANK_HEIGHT
            row += 1
            continue

        c = ws.cell(row=row, column=2, value=texte)
        c.alignment = Alignment(wrap_text=True, vertical="top")

        if row_type == "title":
            c.font = F_TITLE
            ws.row_dimensions[row].height = ROW_HEIGHT_TITLE
        elif row_type == "version":
            c.font = F_MUTED
            ws.row_dimensions[row].height = ROW_HEIGHT_VERSION
        elif row_type == "section":
            c.font = F_SECTION
            ws.row_dimensions[row].height = ROW_HEIGHT_SECTION
            c.border = Border(top=THIN)
        elif row_type == "bullet":
            c.font = F_BODY
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                     indent=1)
            ws.row_dimensions[row].height = ROW_HEIGHT_BULLET
        elif row_type == "color":
            c.font = Font(name="Consolas", size=10, color=INK)
            ws.row_dimensions[row].height = ROW_HEIGHT_COLOR
        else:  # body
            c.font = F_BODY
            ws.row_dimensions[row].height = ROW_HEIGHT_BODY

        row += 1

    return ws


# -------------------------------------------------------------- Schema.gs --
def js(valeur):
    return json.dumps(valeur, ensure_ascii=False, indent=2)


def render_schema_gs():
    seuils = "[" + ", ".join(
        f'["{statut}", {seuil}]' for statut, seuil in S.DELAI_SEUILS) + "]"
    pertinences = "[" + ", ".join(
        f'["{libelle}", {score}]' for libelle, score in S.PERTINENCE_SEUILS) + "]"
    # Le catalogue voyage avec le script : voir SOURCES_LIVREES ci-dessous.
    sources_livrees = lire_sources()
    return f"""/**
 * TenderPilot - FICHIER GENERE. NE PAS MODIFIER A LA MAIN.
 *
 * Genere par builders/toolkit.py depuis schema/columns.py.
 * Pour renommer une colonne : modifier le schema, puis relancer
 * `python build.py`.
 *
 * Schema {S.SCHEMA_VERSION}
 */

var SCHEMA = {{
  VERSION: {js(S.SCHEMA_VERSION)},
  SHEETS: {js(S.SHEETS)},

  /** Cle technique -> nom de colonne de l'onglet OPPORTUNITIES. */
  OPP: {js(S.OPP_KEYS)},

  /** Cle technique -> nom de colonne de l'onglet SOURCES. */
  SRC: {js(S.SOURCE_KEYS)},

  /** Champs compares a chaque collecte pour detecter un changement. */
  UPDATABLE: {js(S.UPDATABLE)},

  ID_PREFIX: {js(S.ID_PREFIX)},
  SUMMARY_MAX: {S.SUMMARY_MAX},

  STATUT_OUVERT: {js(S.STATUT_OUVERT)},
  STATUT_SURVEILLER: {js(S.STATUT_SURVEILLER)},
  STATUT_BIENTOT: {js(S.STATUT_BIENTOT)},
  STATUT_URGENT: {js(S.STATUT_URGENT)},
  STATUT_EXPIRE: {js(S.STATUT_EXPIRE)},
  STATUT_INCONNU: {js(S.STATUT_INCONNU)},

  /** (statut, seuil haut inclus), du plus urgent au plus large. */
  DELAI_SEUILS: {seuils},

  COULEURS: {js(S.COULEURS)},

  /**
   * Pertinence : ce que l'annonce vaut POUR CE CLIENT-LA, d'apres
   * PAYS_SUIVIS et SECTEURS_SUIVIS. Elle etiquette, elle ne supprime pas.
   */
  PERTINENCE_PRIORITAIRE: {js(S.PERTINENCE_PRIORITAIRE)},
  PERTINENCE_A_VOIR: {js(S.PERTINENCE_A_VOIR)},
  PERTINENCE_POSSIBLE: {js(S.PERTINENCE_POSSIBLE)},
  PERTINENCE_HORS_PROFIL: {js(S.PERTINENCE_HORS_PROFIL)},

  /** (libelle, score minimum), du plus pertinent au moins pertinent. */
  PERTINENCE_SEUILS: {pertinences},

  /** Un pays ecrit ainsi n'exclut personne : l'annonce est ouverte a tous. */
  PAYS_OUVERTS: {js(S.PAYS_OUVERTS)},

  /** Une opportunite recoit au maximum un email de chaque type. */
  NOTIFICATIONS: {js(S.NOTIFICATIONS)},

  /** Colonnes de l'onglet PAYS_ET_SECTEURS, reecrit a chaque passage. */
  PROFIL: {js(S.PROFIL)},
  PROFIL_TYPE_PAYS: {js(S.PROFIL_TYPE_PAYS)},
  PROFIL_TYPE_SECTEUR: {js(S.PROFIL_TYPE_SECTEUR)},

  /**
   * Le catalogue de sources livre avec cette version.
   *
   * Il est embarque ici pour que le script puisse comparer l'onglet SOURCES
   * a ce qui devrait s'y trouver. Sans cette copie, un classeur installe il
   * y a six mois n'aurait aucun moyen d'apprendre qu'une source a ete
   * ajoutee ou qu'une adresse a change : il faudrait recreer le classeur et
   * perdre les opportunites deja collectees.
   *
   * Chaque entree est une ligne prete pour l'onglet, dans l'ordre des
   * colonnes de SCHEMA.SRC.
   */
  SOURCES_LIVREES: {js(sources_livrees)}
}};

// Export pour les tests Node. Apps Script n'a pas de module.
if (typeof module !== 'undefined') {{
  module.exports = {{ SCHEMA: SCHEMA }};
}}
"""


# ----------------------------------------------------------------- README --
# A quoi sert chaque fichier de script, pour le guide d'installation.
ROLES_SCRIPTS = {
    "Core": "logique commune : dates, doublons, statuts",
    "Rss": "lecture des flux RSS et Atom",
    "Html": "lecture des sites sans flux",
    "Json": "lecture des API publiques",
    "Sheet": "acces au classeur",
    "Sources": "mise a jour du catalogue de sources",
    "Telegram": "notifications sur Telegram",
    "Run": "collecte, deadlines, emails et menu",
}


def render_readme(date):
    # La liste vient de SCRIPT_FILES : le guide ne peut pas oublier un
    # fichier qu'on aurait ajoute au livrable. Schema est traite a part,
    # il remplace le Code.gs cree par Google.
    autres = [f[:-3] for f in SCRIPT_FILES
              if f.endswith(".gs") and f != "Schema.gs"]
    table_scripts = chr(10).join(
        ["| Nom a taper | Contenu a coller | Role |",
         "|-------------|------------------|------|"]
        + [f"| `{n}` | `{n}.gs` | {ROLES_SCRIPTS.get(n, '')} |" for n in autres])
    nb_scripts = len(autres) + 1
    nb_sources = len(lire_sources())
    nb_autres = len(autres)

    return f"""# TenderPilot

Version {VERSION} - genere le {date}

Collecte les opportunites publiees par {nb_sources} sources - flux RSS,
API publiques et sites sans flux - les enregistre dans Google Sheets, met a
jour les deadlines, colore les lignes et envoie les emails utiles. Rien
d'autre.

```
SOURCES -> COLLECTE -> DEDUPLICATION -> GOOGLE SHEETS
       -> DEADLINE -> COULEUR -> EMAIL
```

## Contenu

| Fichier | Role |
|---------|------|
| `TenderPilot.xlsx` | le classeur : 4 onglets visibles, plus `SOURCES` masque |
| `appsscript.json` | manifeste et autorisations |
| `Schema.gs` | **genere** depuis `schema/columns.py`, ne pas editer |
| `Core.gs` | logique pure, testee hors de Google |
| `Rss.gs` | lecture des flux RSS et Atom |
| `Html.gs` | extraction des pages sans flux (gouv.bj, BAD, Enabel, ARMP) |
| `Json.gs` | lecture des API publiques (Banque mondiale) |
| `Telegram.gs` | notifications sur Telegram |
| `Sources.gs` | synchronisation du catalogue de sources |
| `Sheet.gs` | acces au classeur |
| `Run.gs` | collecte, deadlines, emails, menu, declencheurs |

## 1. Installation

Comptez 15 minutes. Il faut un ordinateur : l'editeur de script n'existe pas
sur telephone.

### Etape 1 - Importer le classeur

1. Ouvrez sheets.google.com et creez une **feuille vierge**.
2. *Fichier > Importer > Importer*, deposez `TenderPilot.xlsx`.
3. Choisissez **Remplacer la feuille de calcul**, puis *Importer les donnees*.

> **Ne double-cliquez pas sur le fichier .xlsx depuis Drive.** Google
> l'ouvrirait en mode compatibilite Office : il resterait un fichier Excel,
> et le menu Apps Script n'existerait pas. Si vous voyez un badge `.XLSX` a
> cote du titre, faites *Fichier > Enregistrer au format Google Sheets* et
> travaillez sur la copie obtenue.

Vous devez voir 4 onglets : `LISEZ_MOI`, `OPPORTUNITIES`,
`CONFIG`, `LOGS`. Un cinquieme, `SOURCES`, existe mais est masque : c'est de
la plomberie, tenue a jour toute seule.

### Etape 2 - Regler le fuseau horaire

*Fichier > Parametres > Fuseau horaire* : celui de votre pays. Ce reglage
determine a quelle heure partent les rappels.

### Etape 3 - Ouvrir l'editeur de script

*Extensions > Apps Script*.

> Dans ce menu, **Apps Script** se trouve juste au-dessus d'**AppSheet**.
> Ne les confondez pas : AppSheet est un autre produit, sans rapport.
>
> Si Apps Script n'apparait pas du tout : soit votre fichier est encore un
> `.xlsx` (voir etape 1), soit votre compte est professionnel et
> l'administrateur a desactive Apps Script.

Un nouvel onglet s'ouvre. En haut a gauche, remplacez "Projet sans titre"
par `TenderPilot`.

### Etape 4 - Creer les {nb_scripts} fichiers de script

Un fichier `Code.gs` existe deja. **Il devient `Schema`** :

1. Selectionnez tout son contenu, supprimez-le, collez celui de `Schema.gs`.
2. Les trois points a cote de son nom > *Renommer* > tapez `Schema`.

Creez ensuite les {nb_autres} autres avec le **+** a cote de "Fichiers", en
choisissant **Script** a chaque fois :

{table_scripts}

**Ne sautez aucun fichier.** Chacun porte une partie du travail : il manque
un fichier, et une partie des sources cesse simplement de repondre, sans
message d'erreur.

> **Tapez le nom sans `.gs`.** Apps Script ajoute l'extension lui-meme :
> taper `Core.gs` produit un fichier `Core.gs.gs`.

L'ordre de creation n'a pas d'importance.

### Etape 5 - Remplacer le manifeste

1. Roue crantee **Parametres du projet**, dans la colonne de gauche.
2. Cochez **Afficher le fichier manifeste "appsscript.json" dans l'editeur**.
3. Revenez aux fichiers (icone `<>`), ouvrez `appsscript.json`, remplacez
   tout son contenu par celui du dossier livre.

### Etape 6 - Autoriser le script

**Cliquez d'abord sur `Run.gs` dans la liste des fichiers.**

> C'est le point ou tout le monde bloque : le bouton **Executer** et le menu
> des fonctions n'apparaissent QUE lorsqu'un fichier `.gs` est ouvert. Sur
> `appsscript.json`, la barre d'outils ne montre que l'enregistrement et le
> journal d'execution, parce qu'un fichier de configuration ne s'execute pas.

Une fois `Run.gs` ouvert, la barre affiche
`Executer` - `Deboguer` - un menu deroulant de fonctions.

1. Dans le menu deroulant, choisissez **`onOpen`**.
2. Cliquez **Executer**.
3. **Examiner les autorisations**, puis choisissez votre compte Google.
4. L'ecran "Google n'a pas valide cette application" apparait. C'est normal
   pour un script non publie : **Parametres avances** en bas a gauche, puis
   **Acceder a TenderPilot (non securise)**, puis **Autoriser**.
5. Le journal affiche "Execution terminee".

L'icone d'enregistrement grisee signifie que tout est deja enregistre : il
n'y a rien a faire de plus.

### Etape 7 - Retrouver le menu

Retournez sur l'onglet du classeur et **rechargez la page** (F5).

Le menu **TenderPilot** apparait dans la barre, entre *Extensions* et
*Aide*. S'il n'y est pas, attendez dix secondes et rechargez a nouveau : le
menu se charge apres la feuille.

## 2. Configuration des emails

Onglet `CONFIG`, colonne **Valeur** uniquement :

| Cle | Effet |
|-----|-------|
| `NOTIFICATION_EMAIL` | adresse qui recoit tout. Vide = aucun email. |
| `SEND_NEW_OPPORTUNITY` | email a chaque nouvelle opportunite |
| `SEND_J7` / `SEND_J3` / `SEND_J1` | rappels d'echeance |
| `SEND_EXPIRED` | email quand la deadline est passee (`false` par defaut) |
| `DIGEST_THRESHOLD` | au-dela de ce nombre de nouveautes, un seul email recapitulatif |
| `TIMEZONE` | fuseau des calculs de deadline |
| `MAX_ITEMS_PER_SOURCE` | annonces lues par source et par execution |

Une opportunite recoit **au maximum un email de chaque type**. Les colonnes
`Notif_*`, masquees, retiennent ce qui a deja ete envoye : ne les videz pas,
sinon les emails repartent.

## 3. Ajouter une source

Onglet `SOURCES`, une ligne par source :

| Colonne | Valeur |
|---------|--------|
| `Source_ID` | identifiant libre, par exemple `SRC-003` |
| `Nom` | nom lisible du site |
| `Methode` | voir le tableau ci-dessous |
| `URL` | adresse du flux, de la page ou de l'API |
| `Pays_Defaut`, `Secteur_Defaut`, `Type_Defaut` | valeurs appliquees aux annonces de cette source |
| `Active` | `OUI` pour l'inclure dans la collecte |

Quatre methodes existent :

| Methode | Ce que c'est |
|---------|--------------|
| `RSS` | un flux RSS ou Atom. **La seule que vous pouvez ajouter vous-meme.** |
| `HTML:<site>` | une page lue par un analyseur dedie, ecrit pour ce site |
| `JSON:<site>` | une API publique lue par un adaptateur dedie |
| `MANUAL` | rien n'est collecte, la saisie est manuelle |

Vous pouvez ajouter librement une source `RSS`. Pour trouver son flux :
icone RSS sur le site, ou essayer `/rss` ou `/feed` apres l'adresse de la
page des appels d'offres.

Les methodes `HTML:` et `JSON:` demandent du code : chaque site a son propre
analyseur. Les sources livrees couvrent deja le portail national beninois,
la Banque africaine de developpement, Enabel, l'ARMP et la Banque mondiale.
Pour en ajouter une autre, il faut nous la demander.

Une source qui exige un login, un captcha ou un navigateur automatise se
marque `MANUAL` et se saisit a la main. C'est un choix assume : ces sources
cassent en permanence.

Attention : une methode `HTML:` ou `JSON:` dont l'analyseur n'existe pas ne
provoque pas d'erreur. La source est simplement ignoree, et l'onglet `LOGS`
indique une saisie manuelle.

## 4. Mettre a jour les sources

Menu **TenderPilot > Synchroniser les sources**.

Les sources bougent : une adresse change, un site est ajoute, une extraction
est reparee. La synchronisation aligne votre classeur sur le catalogue livre
avec la derniere version du script, sans rien vous faire perdre :

- **aucune ligne n'est effacee** - vos propres sources restent ;
- **la colonne Active n'est jamais touchee** - ce que vous avez desactive
  reste desactive ;
- seuls le nom, l'adresse, la methode et le statut de verification sont
  rafraichis.

Un message resume ce qui a ete ajoute et mis a jour. Le detail est dans
l'onglet `LOGS`.

L'onglet `SOURCES` est livre **masque** : c'est de la plomberie, pas un
tableau de bord, et la synchronisation le tient a jour sans que vous ayez a
l'ouvrir. Pour le voir : menu **TenderPilot > Afficher / masquer l onglet
SOURCES**.

## 5. Lancer manuellement

Menu **TenderPilot > Executer maintenant**. Un message resume ce qui a ete
fait ; le detail est dans l'onglet `LOGS`.

## 6. Execution automatique

Menu **TenderPilot > Activer l execution automatique** : trois passages par
jour, a 8h, 13h et 18h. Les deadlines et les couleurs sont recalculees a
chaque passage, meme sans nouvelle opportunite.

## 7. Recevoir les alertes sur Telegram

En plus des emails. Un email se perd dans une boite deja pleine ; une
notification Telegram arrive sur le telephone.

1. Dans Telegram, ecrivez a **@BotFather**, envoyez `/newbot` et suivez les
   questions. Il vous donne un **jeton**.
2. Ecrivez a **@userinfobot** : il vous donne votre **identifiant de salon**.
   Pour un groupe, ajoutez-y d'abord votre bot.
3. Onglet `CONFIG`, renseignez trois lignes :

| Cle | Valeur |
|-----|--------|
| `SEND_TELEGRAM` | `true` |
| `TELEGRAM_TOKEN` | le jeton donne par @BotFather |
| `TELEGRAM_CHAT_ID` | votre identifiant de salon |

4. Menu **TenderPilot > Tester la notification Telegram**. Un message doit
   arriver dans les secondes qui suivent.

> **Le jeton permet d'ecrire a votre place.** Ne le partagez pas, et ne
> laissez pas le classeur ouvert en modification a n'importe qui.

Les deux canaux suivent les memes regles : une opportunite ne vous previent
jamais deux fois par le meme canal. Mais ils sont independants - si Telegram
tombe, les emails partent quand meme, et l'inverse est vrai aussi.

Vous pouvez n'utiliser que Telegram : laissez `NOTIFICATION_EMAIL` vide.

## 8. Ce que le script ne fait pas

Pas d'IA, pas de scoring, pas de Go/No-Go, pas de base de donnees, pas de
comptes utilisateurs. Une deadline absente reste vide : elle n'est jamais
devinee.
"""


# ------------------------------------------------------------- AGENTS.md --
#
# Un seul texte de reference pour tous les agents.
#
# AGENTS.md est la convention que lisent Cursor, Codex, Aider et les autres.
# Claude Code, lui, charge une skill. Plutot que d'entretenir deux copies -
# et de les laisser diverger comme le registre de sources l'avait fait - la
# skill est GENEREE depuis AGENTS.md, frontmatter en tete.

SKILL_FRONTMATTER = """---
name: tenderpilot
description: Travailler sur TenderPilot - ajouter une source de marches publics (pays ou organisme), corriger un bug de collecte, ou auditer le registre (types, secteurs, coherence entre les quatre copies). A charger avant toute modification de data/sources.csv, des analyseurs (html.ts, json.ts, Html.gs, Json.gs), du schema ou des guides. Contient les invariants du projet et les pieges qui ont deja coute du temps.
---

<!-- FICHIER GENERE depuis AGENTS.md par builders/toolkit.py.
     Modifiez AGENTS.md, puis relancez `python build.py`. -->

"""


def generer_skill():
    """Ecrit la skill Claude a partir d'AGENTS.md."""
    source = ROOT / "AGENTS.md"
    if not source.exists():
        return None
    texte = source.read_text(encoding="utf-8")
    # Le commentaire d'en-tete d'AGENTS.md parle de la skill : inutile de le
    # repeter dans la skill elle-meme.
    if texte.startswith("<!--"):
        texte = texte[texte.index("-->") + 3:].lstrip()

    cible = ROOT / ".claude" / "skills" / "tenderpilot" / "SKILL.md"
    cible.parent.mkdir(parents=True, exist_ok=True)
    cible.write_text(SKILL_FRONTMATTER + texte, encoding="utf-8",
                      newline=chr(10))
    return cible


# ------------------------------------------------------------------ build --
def build():
    date = dt.date.today().isoformat()

    # La skill Claude suit AGENTS.md : une seule reference pour tous les
    # agents, humains ou non.
    generer_skill()

    SCHEMA_GS.write_text(render_schema_gs(), encoding="utf-8")

    wb = Workbook()
    feuille_demarrage(wb)
    feuille_opportunites(wb)
    feuille_sources(wb)
    feuille_config(wb)
    feuille_profil(wb)
    feuille_logs(wb)
    # PAYS_ET_SECTEURS se lit juste apres CONFIG : on y regle son profil,
    # on vient ici verifier ce qui existe.
    wb._sheets = [wb[n] for n in ["LISEZ_MOI", S.SHEETS["opportunities"],
                                  S.SHEETS["sources"], S.SHEETS["config"],
                                  S.SHEETS["profil"], S.SHEETS["logs"]]]
    wb.active = 0

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    manquants = [n for n in SCRIPT_FILES if not (SCRIPT_DIR / n).exists()]
    if manquants:
        raise FileNotFoundError(f"Fichiers Apps Script absents : {manquants}")
    for nom in SCRIPT_FILES:
        shutil.copy2(SCRIPT_DIR / nom, OUT_DIR / nom)

    (OUT_DIR / "README.md").write_text(render_readme(date), encoding="utf-8")
    return OUT_DIR


if __name__ == "__main__":
    print(f"OK  {build()}")
